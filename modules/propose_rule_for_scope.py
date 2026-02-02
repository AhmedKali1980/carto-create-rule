#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
propose_rule_for_scope_fix11.py — Carto NG

What this script does
---------------------
Build a "Flow ↔ Applicable Rules" analysis sheet for one RUN scope (app/env):
- Read egress / ingress flow CSV extracts (raw or zone-filtered).
- Factorize flows into proposal rows (direction + anchor labels + peer identity + proto/port).
- Load "Scope Applicable Rules" using the *frozen* logic in scope_rules_applicability.py (NO change).
- For each factorized flow row, try to identify which applicable rule(s) can authorize it.
- Classify + style rows:
  - Deleted workload (peer is a single IP and peer name contains "Deleted workload") → Grey
  - Matched Bouquets Infra rule → Light blue
  - Matched Business Rule in Scope → Light green
  - Matched Business Rule in other Scope → Light pink
  - No match → Light yellow + Action=Create
  - Multiple matches (redundancy / optimization candidate) → Light orange + Action=Optimize
- Add multi-line "redundant_*" columns to quickly locate extra rules authorizing the same flow.
- Write:
  - CSV (in RUN/derived/)
  - Excel sheets:
      * "Flow-Rule Match" (and also "Proposed Rules" for backward compatibility)
      * "Action Plan" (only Create/Optimize lines)

Hard constraints
----------------
- Do NOT modify the scope applicability logic (scope_rules_applicability.py). We only *consume* its output.
- No regressions on flow factorization or IPList election (kept from previous fixes).

Notes about "All workloads"
---------------------------
A rule side with src_all_workloads/dst_all_workloads = true is NOT globally permissive.
It is only meaningful inside a ruleset with a non-empty ruleset_scope, and it applies only to
workloads matching that ruleset_scope (typically app/env). If ruleset_scope is empty, we ignore
all_workloads for matching purposes to avoid false positives.

"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
import time
import ipaddress
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Set, Tuple, Iterator

# -----------------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------------

logger = logging.getLogger("propose_rule_for_scope")

def setup_logging(level: str) -> None:
    lvl = getattr(logging, (level or "INFO").upper(), logging.INFO)
    logging.basicConfig(
        level=lvl,
        format="[%(levelname)s %(asctime)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

def info(msg: str) -> None:
    logger.info(msg)

def warn(msg: str) -> None:
    logger.warning(msg)

def dbg(enabled: bool, msg: str) -> None:
    if enabled:
        logger.debug(msg)

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------

LABEL_KEYS = ["app", "env", "role", "loc", "OS"]

CAT_BOUQUET = "Bouquets Infra rule"
CAT_BIZ_IN  = "Business Rule in Scope"
CAT_BIZ_OUT = "Business Rule in other Scope"

INFO_DELETED = "Deleted workload"
INFO_BOUQUET = "Bouquets Infra"
INFO_BIN     = "Business in scope"
INFO_BOUT    = "Business in other scope"
INFO_NOMATCH = "No matching rule"

ACTION_NO     = "No Action"
ACTION_CREATE = "Create"
ACTION_OPTIM  = "Optimize"

# -----------------------------------------------------------------------------
# Import frozen logic (no change)
# -----------------------------------------------------------------------------

def _import_scope_module():
    """
    Import scope_rules_applicability from either:
    - sibling module (when running from carto-bash/modules)
    - project root (when running from carto-bash)
    """
    try:
        # running from modules/
        import scope_rules_applicability as sra  # type: ignore
        return sra
    except Exception:
        # running from root: ensure modules/ is on path
        here = Path(__file__).resolve()
        sys.path.insert(0, str(here.parent))
        try:
            import scope_rules_applicability as sra  # type: ignore
            return sra
        except Exception as e:
            raise ImportError("Cannot import scope_rules_applicability.py from modules/.") from e

SRA = _import_scope_module()

# We reuse their token parsing + labelgroup loader to stay consistent.
parse_kv_tokens = getattr(SRA, "parse_kv_tokens")
load_labelgroups = getattr(SRA, "load_labelgroups")
get_applicable_rules = getattr(SRA, "get_applicable_rules")

# -----------------------------------------------------------------------------
# Path helpers
# -----------------------------------------------------------------------------

def run_root_from(p: Optional[Path]) -> Optional[Path]:
    """
    If p is inside RUNS/<run_id>/{raw,derived,excel,log}, return RUNS/<run_id>.
    """
    if not p:
        return None
    cur = p if p.is_dir() else p.parent
    cur = cur.resolve()
    for _ in range(12):
        if (cur / "raw").exists() and (cur / "derived").exists():
            return cur
        if cur.parent == cur:
            break
        cur = cur.parent
    return None

def cohere_paths(raw_dir: Path, derived_dir: Path, excel_path: Optional[Path], debug: bool=False) -> Tuple[Path, Path, Optional[Path]]:
    rr = run_root_from(raw_dir) or run_root_from(derived_dir) or run_root_from(excel_path)
    if rr:
        raw_guess = rr / "raw"
        der_guess = rr / "derived"
        if raw_guess.exists():
            raw_dir = raw_guess
        if der_guess.exists():
            derived_dir = der_guess
        if excel_path:
            # keep as-is unless relative
            pass
    raw_dir = raw_dir.resolve()
    derived_dir = derived_dir.resolve()
    if excel_path:
        excel_path = excel_path.resolve()
    dbg(debug, f"cohere_paths => raw_dir={raw_dir} derived_dir={derived_dir} excel={excel_path}")
    return raw_dir, derived_dir, excel_path

def find_conf_path(conf_arg: str, debug: bool=False) -> Optional[Path]:
    """
    Find carto.conf from:
      - explicit arg path
      - CWD/carto.conf
      - parent dirs
    """
    p = Path(conf_arg).expanduser()
    if p.exists():
        return p.resolve()
    p2 = Path.cwd() / conf_arg
    if p2.exists():
        return p2.resolve()

    cur = Path.cwd().resolve()
    for _ in range(8):
        candidate = cur / conf_arg
        if candidate.exists():
            return candidate.resolve()
        if cur.parent == cur:
            break
        cur = cur.parent
    dbg(debug, f"conf not found via arg={conf_arg}")
    return None

# -----------------------------------------------------------------------------
# CSV helpers
# -----------------------------------------------------------------------------

def _sniff_delimiter(sample: str) -> str:
    # tolerate ill-formed samples
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        return dialect.delimiter
    except Exception:
        return ","

def _iter_csv_rows(path: Path, debug: bool=False) -> Tuple[List[Dict[str, str]], List[str]]:
    if not path.exists():
        return [], []
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        sample = f.read(4096)
        delim = _sniff_delimiter(sample)
        f.seek(0)
        reader = csv.DictReader(f, delimiter=delim)
        rows = [{(k or "").strip(): (v or "").strip() for k, v in r.items()} for r in reader if any((v or "").strip() for v in r.values())]

        cols = list(reader.fieldnames or [])
    dbg(debug, f"read csv: {path.name} rows={len(rows)} delim={repr(delim)} cols={len(cols)}")
    return rows, cols


def _iter_csv_rows_stream(path: Path, debug: bool = False) -> Iterator[Dict[str, str]]:
    """
    Stream CSV rows as dicts (low memory).

    Unlike _iter_csv_rows(), this does NOT materialize the full file in memory.
    Intended for large flow files (Flow-in/Flow-out) to avoid OOM kills.
    """
    if not path.exists():
        return iter(())

    def _gen() -> Iterator[Dict[str, str]]:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(4096)
            delim = _sniff_delimiter(sample)
            f.seek(0)

            reader = csv.DictReader(f, delimiter=delim)
            if debug:
                dbg(
                    f"csv stream: {path.name} delimiter={delim!r} "
                    f"cols={len(reader.fieldnames or [])}"
                )

            for raw in reader:
                if not raw:
                    continue
                row = {str(k).strip(): (v or "").strip() for k, v in raw.items() if k is not None}
                if not row:
                    continue
                # Skip fully empty lines
                if not any((v or "").strip() for v in row.values()):
                    continue
                yield row

    return _gen()


def _as_text(x: Any, joiner: str = "|") -> str:
    """Coerce potentially aggregated values (set/list/tuple) to deterministic text.

    Some intermediate rows keep aggregated fields (peer_value, redundant_* columns) as sets.
    Many helpers in this file assume strings and call `.strip()`. This helper makes the
    conversion explicit and deterministic.
    """
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    if isinstance(x, (set, list, tuple)):
        parts = [str(p) for p in x if str(p).strip()]
        parts = sorted(set(parts))
        return joiner.join(parts)
    return str(x)

# -----------------------------------------------------------------------------
# Safe wrapper around frozen parse_kv_tokens (handles non-string inputs)
# -----------------------------------------------------------------------------
# Some aggregation steps can produce sets/lists for fields like peer_value. The frozen
# parse_kv_tokens assumes a string and calls .strip() on its input. We keep the frozen
# contract untouched by wrapping it here.
_parse_kv_tokens_frozen = parse_kv_tokens  # keep reference to frozen implementation

def parse_kv_tokens(x: Any) -> Dict[str, Set[str]]:
    """Wrapper over frozen parse_kv_tokens that tolerates set/list/tuple inputs.

    The internal "compact labels" representation uses pipes (app=...|env=...|...).
    The frozen parser expects ';' separators, so we normalize pipes to ';' here.
    """
    s = _as_text(x, joiner=";")
    s = s.replace("|", ";")
    return _parse_kv_tokens_frozen(s)


def _debug_dump_non_string_fields(proposals: List[Dict[str, Any]], limit: int = 40) -> None:
    """Emit debug logs for any proposal fields that are not plain strings.

    This helps troubleshooting cases where a `.strip()` is accidentally called on a set/list.
    """
    seen = 0
    for i, p in enumerate(proposals):
        for k, v in p.items():
            if isinstance(v, (set, list, tuple, dict)):
                try:
                    if isinstance(v, dict):
                        sample = list(v.items())[:5]
                        desc = f"dict(len={len(v)}) sample={sample}"
                    else:
                        vv = list(v)  # type: ignore
                        sample = [str(x) for x in vv[:5]]
                        desc = f"{type(v).__name__}(len={len(v)}) sample={sample}"
                except Exception:
                    desc = f"{type(v).__name__}(unprintable)"
                warn(f"[debug-proposed-rules] proposals[{i}] field '{k}' is {desc}")
                seen += 1
                if seen >= limit:
                    warn(f"[debug-proposed-rules] ... truncated after {limit} entries")
                    return


def _pick(cols: List[str], *names: str) -> str:
    """
    Pick the first column in cols matching any 'names' case-insensitively, with minor normalization.
    """
    norm = {c.lower().strip().replace("_", " ").replace("-", " "): c for c in cols}
    for n in names:
        key = n.lower().strip().replace("_", " ").replace("-", " ")
        if key in norm:
            return norm[key]
    # fallback: contains
    for n in names:
        key = n.lower().strip()
        for c in cols:
            if key in (c or "").lower():
                return c
    return ""

def parse_semicolon_list(s: str) -> List[str]:
    return [p.strip() for p in (s or "").split(";") if p.strip()]

def _norm_label_key(k: str) -> str:
    k = (k or "").strip()
    if not k:
        return ""
    if k.lower() == "os":
        return "OS"
    return k.lower()

def _parse_rule_labels_kv(raw: str) -> Dict[str, Set[str]]:
    """
    Parse a rule-side label expression into a kv map with **OR semantics per key**
    and **AND semantics across keys**.

    Examples accepted:
      - "env:PRD;app:FOO;role:PSM;role:PSMP"
      - "OS:LINUX;OS:AIX;OS:WINDOWS"
      - "app=FOO|env=PRD" (pipes and '=' are also accepted)
    """
    if not raw:
        return {}
    # normalize separators to ';'
    s = (raw or "")
    s = s.replace("|", ";").replace(",", ";")
    out: Dict[str, Set[str]] = {}
    for tok in parse_semicolon_list(s):
        t = (tok or "").strip()
        if not t:
            continue
        t = t.strip()
        # tolerate wrapping parentheses fragments (e.g., from copied cells)
        t = t.strip("()")
        if not t:
            continue
        if ":" in t:
            k, v = t.split(":", 1)
        elif "=" in t:
            k, v = t.split("=", 1)
        else:
            continue
        nk = _norm_label_key(k)
        nv = (v or "").strip()
        if not nk or not nv:
            continue
        out.setdefault(nk, set()).add(nv)
    return out


# -----------------------------------------------------------------------------
# carto.conf parsing (IPList election patterns)
# -----------------------------------------------------------------------------

def load_conf(conf_path: Optional[Path], debug: bool=False) -> Tuple[List[str], List[str]]:
    """
    Returns (allowed_prefixes, priority_prefixes) from carto.conf.
    Both are returned as list of prefixes in order.
    """
    if not conf_path or not conf_path.exists():
        warn("carto.conf not found: IPList election will accept any name and no priority.")
        return [], []
    allowed: List[str] = []
    prio: List[str] = []
    with conf_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip()
            if k == "IPLIST_ALLOWED_PREFIXES":
                allowed = [x for x in v.split(";") if x]
            elif k == "IPLIST_NAME_PRIORITY":
                prio = [x for x in v.split(";") if x]
    dbg(debug, f"conf: IPLIST_ALLOWED_PREFIXES={allowed} IPLIST_NAME_PRIORITY={prio}")
    return allowed, prio

def _match_prefix_or_glob(name: str, pat: str) -> bool:
    if not pat:
        return False
    if pat.endswith("*"):
        return name.startswith(pat[:-1])
    return name.startswith(pat)

def elect_iplist_from_tokens(iplists_raw: str,
                             allowed_pats: List[str],
                             prio_pats: List[str],
                             debug: bool=False) -> Tuple[str, str]:
    """
    Given a semicolon list of IPList names (and sometimes hrefs), pick a single "best" IPList.
    - keep only allowed prefixes if allowed_pats is provided
    - apply priority order (first match wins)
    - else fallback to lexical min for determinism
    Returns (elected_name, reason)
    """
    toks = [t.strip() for t in re.split(r"[;,\s]+", iplists_raw or "") if t.strip()]
    if not toks:
        return "", "no_iplist"
    # Keep only "names" (strip href tails)
    names: List[str] = []
    for t in toks:
        # href -> take last part
        if "/" in t:
            t2 = t.rsplit("/", 1)[-1]
            names.append(t2)
        else:
            names.append(t)
    # Filter allowed
    if allowed_pats:
        kept = [n for n in names if any(_match_prefix_or_glob(n, p) for p in allowed_pats)]
        if kept:
            names = kept
    # Priority
    for p in prio_pats or []:
        for n in names:
            if _match_prefix_or_glob(n, p):
                return n, f"priority({p})"
    # Deterministic fallback
    chosen = sorted(set(names))[0]
    return chosen, "fallback_lexical"

# -----------------------------------------------------------------------------
# Flow file selection
# -----------------------------------------------------------------------------

def _find_flow_files(raw_dir: Path, derived_dir: Path, start: str, end: str, prefer_raw: bool, debug: bool=False) -> Tuple[Optional[Path], Optional[Path]]:
    """
    Find (out_path, in_path).
    Priority:
      1) zone-filtered files in derived_dir if present (and not prefer_raw):
         - flows_out.zone.csv / flows_in.zone.csv
      2) exact raw names:
         - flows_out_{start}_{end}.csv / flows_in_{start}_{end}.csv
      3) auto select latest matching pair in raw_dir: flows_{in|out}_YYYY-MM-DD_YYYY-MM-DD.csv
    """
    zone_out = derived_dir / "flows_out.zone.csv"
    zone_in  = derived_dir / "flows_in.zone.csv"
    if not prefer_raw and zone_out.exists() and zone_in.exists():
        info(f"using zone-filtered flows: {zone_out.name} / {zone_in.name}")
        return zone_out, zone_in

    exact_out = raw_dir / f"flows_out_{start}_{end}.csv"
    exact_in  = raw_dir / f"flows_in_{start}_{end}.csv"
    if exact_out.exists() and exact_in.exists():
        info(f"using exact raw flows: {exact_out.name} / {exact_in.name}")
        return exact_out, exact_in

    pat = re.compile(r"^flows_(in|out)_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})\.csv$")
    ins: Dict[Tuple[str, str], Path] = {}
    outs: Dict[Tuple[str, str], Path] = {}
    for p in raw_dir.glob("flows_in_*.csv"):
        m = pat.match(p.name)
        if m:
            ins[(m.group(2), m.group(3))] = p
    for p in raw_dir.glob("flows_out_*.csv"):
        m = pat.match(p.name)
        if m:
            outs[(m.group(2), m.group(3))] = p
    keys = sorted(set(ins.keys()) & set(outs.keys()))
    if not keys:
        warn(f"No matching in/out pairs found in {raw_dir}")
        return None, None
    # choose the most recent (by end date then start date)
    keys_sorted = sorted(keys, key=lambda k: (k[1], k[0]))
    chosen = keys_sorted[-1]
    out_path = outs[chosen]
    in_path = ins[chosen]
    info(f"auto-selected raw flows: {out_path.name} / {in_path.name} (requested start={start}, end={end})")
    dbg(debug, f"available pairs={keys_sorted}")
    return out_path, in_path

def load_flows(raw_dir: Path, derived_dir: Path, start: str, end: str, prefer_raw: bool, debug: bool=False) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    out_path, in_path = _find_flow_files(raw_dir, derived_dir, start, end, prefer_raw, debug=debug)
    if not out_path or not in_path:
        return [], []
    out_rows, _ = _iter_csv_rows(out_path, debug=debug)
    in_rows, _ = _iter_csv_rows(in_path, debug=debug)
    return out_rows, in_rows

# -----------------------------------------------------------------------------
# Extract flow side labels / iplist / names
# -----------------------------------------------------------------------------

def _labels_from_flow_row(row: Dict[str, str], side_prefix: str) -> Dict[str, str]:
    """
    side_prefix in {"source", "destination"}.
    Returns dict with keys LABEL_KEYS.
    """
    # The raw CSV headers may be: "Source Application", "Source Environment", etc.
    # We do not assume exact naming; caller supplies columns.
    out = {k: "" for k in LABEL_KEYS}
    # We accept several variants; we will resolve from the row keys directly.
    # Using contains match to avoid hard dependency on exact exports.
    for k in LABEL_KEYS:
        # map internal keys to header tokens
        if k == "app":
            candidates = [f"{side_prefix} application", f"{side_prefix} app"]
        elif k == "env":
            candidates = [f"{side_prefix} environment", f"{side_prefix} env"]
        elif k == "role":
            candidates = [f"{side_prefix} role"]
        elif k == "loc":
            candidates = [f"{side_prefix} location", f"{side_prefix} loc"]
        elif k == "OS":
            candidates = [f"{side_prefix} os", f"{side_prefix} operating system"]
        else:
            candidates = [f"{side_prefix} {k}"]
        v = ""
        for ck in row.keys():
            low = (ck or "").lower()
            if any(c in low for c in candidates):
                v = row.get(ck, "") or ""
                break
        out[k] = (v or "").strip()
    return out

def _get_by_contains(row: Dict[str, str], *contains: str) -> str:
    for ck in row.keys():
        low = (ck or "").lower()
        if any(c in low for c in contains):
            return (row.get(ck, "") or "").strip()
    return ""

def _get_by_contains_priority(row: Dict[str, str], *contains: str) -> str:
    """
    Like _get_by_contains(), but respects the priority/order of patterns.

    Example: _get_by_contains(row, "protocol", "transmission") can accidentally return
    'Transmission' if that column appears earlier in the CSV. This helper ensures
    'protocol' is preferred when both exist.
    """
    for c in contains:
        c = (c or "").lower()
        if not c:
            continue
        for ck in row.keys():
            low = (ck or "").lower()
            if c in low:
                return (row.get(ck, "") or "").strip()
    return ""


def _to_int(x: Any) -> int:
    try:
        return int(str(x).strip())
    except Exception:
        return 0



def _to_int_local(v: Any) -> int:
    """Local int conversion helper (kept for backward-compatibility).

    Some builder functions (e.g., Proposed rules egress) use _to_int_local while others
    use _to_int. This global helper prevents NameError when the local version is only
    defined inside a different function.
    """
    try:
        return int(str(v or "0").strip() or "0")
    except Exception:
        return 0

def _norm_proto(x: str) -> str:
    p = (x or "").strip().lower()
    if p in ("tcp", "udp", "icmp", "igmp"):
        return p
    # workloader sometimes provides "Transmission" or "Protocol"
    if p.startswith("tcp"):
        return "tcp"
    if p.startswith("udp"):
        return "udp"
    if p.startswith("icmp"):
        return "icmp"
    if p.startswith("igmp"):
        return "igmp"
    return p
# -----------------------------------------------------------------------------
# Rule matching (flows ↔ applicable rules)
# -----------------------------------------------------------------------------

@dataclass(frozen=True)
class ServiceSpec:
    proto: str
    port_start: int
    port_end: int

def _parse_services_raw(services_raw: str) -> List[ServiceSpec]:
    """
    Parse Illumio service strings from exports.

    Supports semicolon-separated tokens in multiple formats, e.g.:
      - tcp/443 ; udp/53 ; icmp ; igmp
      - tcp 443 ; tcp 80-81
      - 443 tcp ; 80-81 tcp
      - ANY
      - Service object wrappers like: NAME ( ... ; ... ) ; 21 TCP
        (the wrapper name is ignored; tokens inside parentheses are parsed too)
    """
    if services_raw is None:
        return []

    s = str(services_raw)

    # Strip service object names like "NAME (...)" -> "(...)"
    # so we can parse the inner token list.
    s = re.sub(r"(?<![A-Za-z0-9_.-])[A-Za-z][A-Za-z0-9_.-]*\s*\(", "(", s)

    # Turn parentheses into separators; the inner list uses semicolons as delimiters.
    s = s.replace("(", ";").replace(")", ";")

    specs: List[ServiceSpec] = []
    for tok in parse_semicolon_list(s):
        t = (tok or "").strip().strip(",")
        if not t:
            continue

        tu = t.upper()

        # ANY
        if tu == "ANY":
            specs.append(ServiceSpec(proto="any", port_start=0, port_end=65535))
            continue

        # IGMP / ICMP (no port notion)
        if tu in ("IGMP", "ICMP"):
            specs.append(ServiceSpec(proto=tu.lower(), port_start=0, port_end=65535))
            continue

        # proto/port or proto/port-range
        m = re.match(r"(?i)^(tcp|udp|icmp|igmp)\s*/\s*(\d+)(?:\s*-\s*(\d+))?$", t)
        if m:
            proto = m.group(1).lower()
            a = int(m.group(2))
            b = int(m.group(3) or m.group(2))
            specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
            continue

        # proto only
        m2 = re.match(r"(?i)^(tcp|udp|icmp|igmp)\s*$", t)
        if m2:
            proto = m2.group(1).lower()
            specs.append(ServiceSpec(proto=proto, port_start=0, port_end=65535))
            continue

        # proto port / proto port-range  (e.g., "tcp 443" or "tcp 80-81")
        m3 = re.match(r"(?i)^(tcp|udp|icmp|igmp)\s+(\d+)(?:\s*-\s*(\d+))?$", t)
        if m3:
            proto = m3.group(1).lower()
            a = int(m3.group(2))
            b = int(m3.group(3) or m3.group(2))
            specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
            continue

        # port proto / port-range proto (e.g., "443 tcp" or "80-81 tcp")
        m4 = re.match(r"(?i)^(\d+)(?:\s*-\s*(\d+))?\s+(tcp|udp|icmp|igmp)\s*$", t)
        if m4:
            a = int(m4.group(1))
            b = int(m4.group(2) or m4.group(1))
            proto = m4.group(3).lower()
            specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
            continue

        # Some exports may include tokens like "22 TCP)" — be permissive:
        t2 = re.sub(r"[^0-9A-Za-z/\-\s]", "", t).strip()
        if t2 and t2 != t:
            m4b = re.match(r"(?i)^(\d+)(?:\s*-\s*(\d+))?\s+(tcp|udp|icmp|igmp)\s*$", t2)
            if m4b:
                a = int(m4b.group(1))
                b = int(m4b.group(2) or m4b.group(1))
                proto = m4b.group(3).lower()
                specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
                continue
            m3b = re.match(r"(?i)^(tcp|udp|icmp|igmp)\s+(\d+)(?:\s*-\s*(\d+))?$", t2)
            if m3b:
                proto = m3b.group(1).lower()
                a = int(m3b.group(2))
                b = int(m3b.group(3) or m3b.group(2))
                specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
                continue
            m1b = re.match(r"(?i)^(tcp|udp|icmp|igmp)\s*/\s*(\d+)(?:\s*-\s*(\d+))?$", t2)
            if m1b:
                proto = m1b.group(1).lower()
                a = int(m1b.group(2))
                b = int(m1b.group(3) or m1b.group(2))
                specs.append(ServiceSpec(proto=proto, port_start=min(a, b), port_end=max(a, b)))
                continue
            if t2.upper() in ("IGMP", "ICMP"):
                specs.append(ServiceSpec(proto=t2.lower(), port_start=0, port_end=65535))
                continue

        # Unknown token: ignore (do not make matching stricter than Illumio)
        continue

    return specs
def _service_matches(flow_proto: str, flow_port: int, specs: List[ServiceSpec]) -> bool:
    if not specs:
        return True  # conservative: if rule has no service info, do not block match here
    fp = _norm_proto(flow_proto)

    try:
        if flow_port is None:
            pnum = 0
        else:
            # workloader CSV may yield floats (e.g., 1522.0) or strings
            pnum = int(float(flow_port))
    except Exception:
        pnum = 0

    for s in specs:
        if s.proto == "any":
            return True
        if fp == s.proto and s.port_start <= pnum <= s.port_end:
            return True
    return False
def _merge_kv(dst: Dict[str, Set[str]], src: Dict[str, Set[str]]) -> Dict[str, Set[str]]:
    out = {k: set(v) for k, v in dst.items()}
    for k, vals in src.items():
        out.setdefault(k, set()).update(vals)
    return out

def _group_kv(group_names: List[str], lg_map: Dict[Tuple[str, str], Set[str]]) -> Dict[str, Set[str]]:
    """
    Convert tuple-key labelgroup map to kv map.
    """
    out: Dict[str, Set[str]] = {}
    for g in group_names:
        g = g.strip()
        if not g:
            continue
        for k in LABEL_KEYS:
            members = lg_map.get((k, g))
            if members:
                out.setdefault(k, set()).update(members)
    return out

def _flow_matches_kv(flow_side: Dict[str, str], rule_kv: Dict[str, Set[str]]) -> bool:
    if not rule_kv:
        return False
    for k, allowed in rule_kv.items():
        fv = (flow_side.get(k, "") or "").strip()
        if not fv or fv not in allowed:
            return False
    return True

def _flow_hits_any_group(flow_side: Dict[str, str], group_name: str, lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    for k in LABEL_KEYS:
        fv = (flow_side.get(k, "") or "").strip()
        if not fv:
            continue
        members = lg_map.get((k, group_name))
        if members and fv in members:
            return True
    return False

def _flow_excluded(flow_side: Dict[str, str],
                   labels_excl_raw: str,
                   groups_excl_raw: str,
                   lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    """
    Return True if flow_side matches any exclusion.
    """
    excl_kv: Dict[str, Set[str]] = _parse_rule_labels_kv(labels_excl_raw or "")
    for k, bad in excl_kv.items():
        fv = (flow_side.get(k, "") or "").strip()
        if fv and fv in bad:
            return True
    for g in parse_semicolon_list(groups_excl_raw or ""):
        if _flow_hits_any_group(flow_side, g, lg_map):
            return True
    return False

def _side_matches(flow_side: Dict[str, str],
                  flow_iplist: str,
                  rule_all: str,
                  rule_iplists_raw: str,
                  rule_labels_raw: str,
                  rule_groups_raw: str,
                  rule_labels_excl_raw: str,
                  rule_groups_excl_raw: str,
                  ruleset_scope_raw: str,
                  lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    """
    Side selection match with safe handling for all_workloads and exclusions.

    We treat rule side selectors (labels + labelgroups) as a *combined constraint*
    (AND across keys), consistent with the frozen scope logic used to compute applicability.

    Match succeeds if:
      - all_workloads is true AND ruleset_scope is non-empty AND flow matches ruleset_scope AND not excluded
      OR
      - flow_iplist matches a selected iplist
      OR
      - flow labels satisfy (labels + expanded labelgroups) AND not excluded
    """
    if _flow_excluded(flow_side, rule_labels_excl_raw, rule_groups_excl_raw, lg_map):
        return False

    all_true = str(rule_all or "").strip().lower() in ("true", "1", "yes", "y")

    # 1) all_workloads (scoped)
    if all_true:
        rs_kv: Dict[str, Set[str]] = parse_kv_tokens(ruleset_scope_raw or "")
        # if no ruleset_scope, all_workloads is meaningless here => ignore to avoid false positives
        if not rs_kv:
            return False
        return _flow_matches_kv(flow_side, rs_kv)

    # 2) iplist
    rule_iplists = set(parse_semicolon_list(rule_iplists_raw or ""))
    if flow_iplist and rule_iplists and flow_iplist in rule_iplists:
        return True

    # 3) labels + groups
    rule_kv: Dict[str, Set[str]] = _parse_rule_labels_kv(rule_labels_raw or "")
    groups = parse_semicolon_list(rule_groups_raw or "")
    if groups:
        rule_kv = _merge_kv(rule_kv, _group_kv(groups, lg_map))
    return _flow_matches_kv(flow_side, rule_kv)

def _flow_sides_from_proposal(p: Dict[str, Any]) -> Tuple[Dict[str, str], str, Dict[str, str], str]:
    """
    Return (src_labels, src_iplist, dst_labels, dst_iplist) in *rule* sense.
    direction:
      - egress: anchor -> peer
      - ingress: peer -> anchor
    Note: if peer_type == 'iplist', we match by peer_value.
    """
    direction = _as_text(p.get("direction_raw") or p.get("direction") or "").strip().lower()
    anchor = {k: str(p.get(f"anchor_{k}", "") or "").strip() for k in LABEL_KEYS}
    peer = {k: str(p.get(f"peer_{k}", "") or "").strip() for k in LABEL_KEYS}

    if bool(p.get("peer_labels_disabled")):
        peer = {k: "" for k in LABEL_KEYS}

    # These fields may be aggregated as sets in intermediate rows.
    anchor_iplist = _as_text(p.get("anchor_iplist", ""), joiner="|").strip()
    peer_iplist = _as_text(p.get("peer_iplist", ""), joiner="|").strip()
    if _as_text(p.get("peer_type") or "").strip() == "iplist":
        peer_iplist = _as_text(p.get("peer_value") or "", joiner="|").strip()

    if direction == "egress":
        return anchor, anchor_iplist, peer, peer_iplist
    # ingress
    return peer, peer_iplist, anchor, anchor_iplist

def proposal_matches_rule(p: Dict[str, Any], rule: Dict[str, str], lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    # service check
    services_raw = rule.get("services_raw") or rule.get("services") or ""
    specs = _parse_services_raw(services_raw)
    if not _service_matches(p.get("proto", ""), _to_int(p.get("port", 0)), specs):
        return False

    src_labels, src_iplist, dst_labels, dst_iplist = _flow_sides_from_proposal(p)

    rs_scope = rule.get("ruleset_scope", "") or ""

    src_ok = _side_matches(
        flow_side=src_labels,
        flow_iplist=src_iplist,
        rule_all=rule.get("src_all_raw", ""),
        rule_iplists_raw=rule.get("src_iplists_raw", ""),
        rule_labels_raw=rule.get("src_labels_raw", ""),
        rule_groups_raw=rule.get("src_groups_raw", ""),
        rule_labels_excl_raw=rule.get("src_labels_excl_raw", ""),
        rule_groups_excl_raw=rule.get("src_groups_excl_raw", ""),
        ruleset_scope_raw=rs_scope,
        lg_map=lg_map,
    )
    if not src_ok:
        return False

    dst_ok = _side_matches(
        flow_side=dst_labels,
        flow_iplist=dst_iplist,
        rule_all=rule.get("dst_all_raw", ""),
        rule_iplists_raw=rule.get("dst_iplists_raw", ""),
        rule_labels_raw=rule.get("dst_labels_raw", ""),
        rule_groups_raw=rule.get("dst_groups_raw", ""),
        rule_labels_excl_raw=rule.get("dst_labels_excl_raw", ""),
        rule_groups_excl_raw=rule.get("dst_groups_excl_raw", ""),
        ruleset_scope_raw=rs_scope,
        lg_map=lg_map,
    )
    return dst_ok

# -----------------------------------------------------------------------------
# Core factorization + classification
# -----------------------------------------------------------------------------

ProposalKey = Tuple[str, str, str, str, str, str, str, str, str, int]  # direction + anchor labels + peer + proto/port

def _compact_labels(d: Dict[str, str]) -> str:
    parts = []
    for k in LABEL_KEYS:
        v = (d.get(k) or "").strip()
        if v:
            parts.append(f"{k}={v}")
    return "|".join(parts)

def _is_intra_app(anchor_labels: Dict[str, str], peer_labels: Dict[str, str], peer_value: str) -> bool:
    """Return True if peer side belongs to same app/env as anchor (in-scope intra-app)."""
    a_app = (anchor_labels.get("app") or "").strip()
    a_env = (anchor_labels.get("env") or "").strip()
    if not a_app or not a_env:
        return False
    p_app = (peer_labels.get("app") or "").strip()
    p_env = (peer_labels.get("env") or "").strip()
    if p_app and p_env and p_app == a_app and p_env == a_env:
        return True
    # fallback: parse from compact 'app=...|env=...' representation
    try:
        kv = parse_kv_tokens(_as_text(peer_value or ""))
        apps = kv.get("app") or set()
        envs = kv.get("env") or set()
        # parse_kv_tokens returns Dict[str, Set[str]] (frozen contract). Keep it robust anyway.
        if isinstance(apps, str):
            apps = {apps.strip()} if apps.strip() else set()
        if isinstance(envs, str):
            envs = {envs.strip()} if envs.strip() else set()
        if (a_app in apps) and (a_env in envs):
            return True
    except Exception:
        pass
    return False


def detect_peer_type(peer_labels: Dict[str, str],
                     peer_ip: str,
                     peer_hostname: str,
                     peer_name: str,
                     iplist_raw: str,
                     group_by: str,
                     allowed_pats: List[str],
                     prio_pats: List[str],
                     debug: bool=False) -> Tuple[str, str, str]:
    """
    Returns (peer_type, peer_value, reason).
    """
    if iplist_raw:
        elected, reason = elect_iplist_from_tokens(iplist_raw, allowed_pats, prio_pats, debug=debug)
        if elected:
            return "iplist", elected, reason

    if group_by == "labels":
        compact = _compact_labels(peer_labels)
        if compact:
            return "labels", compact, "by_peer_labels"

    if peer_ip:
        return "ip", peer_ip, "by_ip"
    if peer_hostname:
        return "hostname", peer_hostname, "by_hostname"
    if peer_name:
        return "ip", peer_name, "by_name"
    return "unknown", "", "no_peer_identity"

def propose(raw_dir: Path,
            derived_dir: Path,
            start: str,
            end: str,
            prefer_raw: bool,
            min_flows: int,
            group_by: str,
            allowed_pats: List[str],
            prio_pats: List[str],
            rules_applicables: List[Dict[str, str]],
            lg_map: Dict[Tuple[str, str], Set[str]],
            debug: bool=False,
            enable_avoid_label_pairs: bool=False,
            avoid_label_pairs: Optional[Set[str]]=None,
            kub_iplist_resolver: Optional[Callable[[str], str]]=None,
    network_zone_nets: Optional[List[ipaddress._BaseNetwork]] = None,
    network_zone_name: str = "",
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    out_path, in_path = _find_flow_files(raw_dir, derived_dir, start, end, prefer_raw=prefer_raw, debug=debug)
    info(f"flows source files: out={out_path.name} in={in_path.name}")

    agg: Dict[ProposalKey, Dict[str, Any]] = {}

    # Optional: avoid some peer app labels by replacing them with KUB_* IPLISTS (deterministic).
    # This is gated behind CLI option --enable-avoid-label-pairs (non-regression).
    avoid_label_pairs = set(avoid_label_pairs or [])
    _kub_resolve_cache: Dict[str, str] = {}
    _avoid_stats = {"hits": 0, "resolved_to_iplist": 0, "fallback_to_ip": 0, "no_ip": 0}




    # Apply --network-zone handling
    #  - Main matching is EAST/WEST strictly in-zone (both endpoints).
    #  - Additionally, to avoid North/South being blocked by remote managed destinations,
    #    we keep a limited set of out-of-zone egress flows for proposal purposes:
    #       egress + destination Managed labels (app prefix ends with 'M')
    #    Those rows are marked nz_ns=True and will be:
    #       - excluded from Flow-Rule Match sheet
    #       - only used for proposals when no existing rule match exists.
    nets_v4: List[ipaddress.IPv4Network] = []
    nets_v6: List[ipaddress.IPv6Network] = []
    using_zone_files: bool = False
    raw_out_path_for_ns: Optional[Path] = None

    if network_zone_nets:
        nets_v4, nets_v6 = _split_nets_by_family(network_zone_nets)

        # If pre-filtered zone files are used, they are EW-only, so we must scan RAW egress
        # to discover NS -> remote managed destinations.
        using_zone_files = (
            (out_path == (derived_dir / "flows_out.zone.csv")) and
            (in_path  == (derived_dir / "flows_in.zone.csv")) and
            (not prefer_raw)
        )
        if using_zone_files:
            raw_out_path_for_ns, _raw_in_ignored = _find_flow_files(
                raw_dir, derived_dir, start, end, prefer_raw=True, debug=debug
            )

    def add_flow(direction: str, row: Dict[str, str]) -> None:
        # identify anchor/peer sides depending on direction
        if direction == "egress":
            anchor_labels = _labels_from_flow_row(row, "source")
            peer_labels   = _labels_from_flow_row(row, "destination")
            anchor_iplist_raw = _get_by_contains(row, "source iplist")
            peer_iplist_raw   = _get_by_contains(row, "destination iplist")
            peer_ip = _get_by_contains(row, "destination ip")
            peer_hostname = _get_by_contains(row, "destination hostname")
            peer_name = _get_by_contains(row, "destination name")
            is_deleted = "deleted workload" in (peer_name or "").lower()
        else:  # ingress
            anchor_labels = _labels_from_flow_row(row, "destination")
            peer_labels   = _labels_from_flow_row(row, "source")
            anchor_iplist_raw = _get_by_contains(row, "destination iplist")
            peer_iplist_raw   = _get_by_contains(row, "source iplist")
            peer_ip = _get_by_contains(row, "source ip")
            peer_hostname = _get_by_contains(row, "source hostname")
            peer_name = _get_by_contains(row, "source name")
            is_deleted = "deleted workload" in (peer_name or "").lower()

        proto_raw = _get_by_contains_priority(row, "protocol") or _get_by_contains(row, "proto")
        proto_norm = _norm_proto(proto_raw)
        proto = proto_norm.upper() if proto_norm else ""
        port = _to_int(_get_by_contains(row, "port"))

        # choose peer identity / type (with optional avoid-label-pairs override)
        peer_labels_disabled = False
        peer_type: str
        peer_value: str
        peer_reason: str

        peer_app = (peer_labels.get("app") or "").strip()
        if enable_avoid_label_pairs and avoid_label_pairs and peer_app and (peer_app in avoid_label_pairs):
            _avoid_stats["hits"] += 1
            ip = (peer_ip or "").strip()
            if ip:
                iplist_name = ""
                if kub_iplist_resolver:
                    if ip in _kub_resolve_cache:
                        iplist_name = _kub_resolve_cache[ip]
                    else:
                        try:
                            iplist_name = (kub_iplist_resolver(ip) or "").strip()
                        except Exception:
                            iplist_name = ""
                        _kub_resolve_cache[ip] = iplist_name
                if iplist_name:
                    peer_type, peer_value, peer_reason = "iplist", iplist_name, "avoid_label_pairs:kub_iplist"
                    peer_labels_disabled = True
                    _avoid_stats["resolved_to_iplist"] += 1
                else:
                    peer_type, peer_value, peer_reason = "ip", ip, "avoid_label_pairs:ip_fallback"
                    peer_labels_disabled = True
                    _avoid_stats["fallback_to_ip"] += 1
            else:
                _avoid_stats["no_ip"] += 1
                peer_type, peer_value, peer_reason = detect_peer_type(
                    peer_labels=peer_labels,
                    peer_ip=peer_ip,
                    peer_hostname=peer_hostname,
                    peer_name=peer_name,
                    iplist_raw=peer_iplist_raw,
                    group_by=group_by,
                    allowed_pats=allowed_pats,
                    prio_pats=prio_pats,
                    debug=debug,
                )
        else:
            peer_type, peer_value, peer_reason = detect_peer_type(
                peer_labels=peer_labels,
                peer_ip=peer_ip,
                peer_hostname=peer_hostname,
                peer_name=peer_name,
                iplist_raw=peer_iplist_raw,
                group_by=group_by,
                allowed_pats=allowed_pats,
                prio_pats=prio_pats,
                debug=debug,
            )

        # UI-only direction override for intra-app flows (keep raw direction for matching)
        direction_raw = direction
        if peer_type == "labels" and _is_intra_app(anchor_labels, peer_labels, peer_value):
            direction_ui = "intra-app"
        else:
            direction_ui = direction
        key: ProposalKey = (
            direction_raw,
            anchor_labels.get("app",""), anchor_labels.get("env",""), anchor_labels.get("role",""),
            anchor_labels.get("loc",""), anchor_labels.get("OS",""),
            peer_type, peer_value,
            proto_norm, port,
        )
        cur = agg.get(key)
        if not cur:
            cur = {
                "direction": direction_ui,
                "direction_raw": direction_raw,
                **{f"anchor_{k}": anchor_labels.get(k, "") for k in LABEL_KEYS},
                "anchor_iplist": "",  # not used as selector now; kept for completeness
                "peer_type": peer_type,
                "peer_value": peer_value,
                "peer_labels_disabled": bool(peer_labels_disabled),
                **{f"peer_{k}": peer_labels.get(k, "") for k in LABEL_KEYS},
                "peer_iplist": "",  # elected may be in peer_value when peer_type=iplist
                "peer_deleted_workload": False,
                "proto": proto,
                "port": port,
                "num_flows": 0,
                "nz_class": "",
                "nz_ns": False,
                "num_flows_true": 0,
                "first_detected": "",
                "last_detected": "",
                "suggested_ruleset": "",
                "matched_rule_category": "",
                # matching output columns
                "Info": "",
                "Action": "",
                # primary rule selectors
                "primary_src_iplists": "",
                "primary_src_labels": "",
                "primary_src_groups": "",
                "primary_dst_iplists": "",
                "primary_dst_labels": "",
                "primary_dst_groups": "",
                # redundant (multiline)
                "redundant_rule_categories": "",
                "redundant_ruleset_names": "",
                "redundant_rule_descriptions": "",
                "redundant_src_iplists": "",
                "redundant_src_labels": "",
                "redundant_src_groups": "",
                "redundant_dst_iplists": "",
                "redundant_dst_labels": "",
                "redundant_dst_groups": "",
                # always last:
                "ruleset_name": "",
                "rule_description": "",
            }
            agg[key] = cur



        # Tag North/South rows (network-zone only): egress where src is in zone and dst is outside zone.
        # Those rows are kept ONLY for proposal purposes (see filtering above) and will be filtered out
        # from the Flow-Rule Match sheet later.
        if network_zone_nets:
            try:
                src_ip_raw0 = _get_by_contains(row, "source ip") or _get_by_contains(row, "src ip") or ""
                dst_ip_raw0 = _get_by_contains(row, "destination ip") or _get_by_contains(row, "dst ip") or ""
                src_ip0 = _extract_first_ip_token(str(src_ip_raw0))
                dst_ip0 = _extract_first_ip_token(str(dst_ip_raw0))
                if direction_raw == "egress" and src_ip0 and dst_ip0:
                    if _ip_in_any_net(src_ip0, nets_v4, nets_v6) and (not _ip_in_any_net(dst_ip0, nets_v4, nets_v6)):
                        cur["nz_class"] = "NS"
                        cur["nz_ns"] = True
                    else:
                        cur["nz_class"] = "EW"
            except Exception:
                # keep defaults
                pass

        cur["num_flows"] = int(cur.get("num_flows", 0)) + 1

        # Real flow count from raw flows column [Num Flows] (Flow-in/out, col AG)
        nf_raw = _get_by_contains(row, "num flows")
        nf_val = 1
        try:
            if nf_raw not in (None, ""):
                nf_val = int(float(str(nf_raw).strip()))
        except Exception:
            nf_val = 1
        cur["num_flows_true"] = int(cur.get("num_flows_true", 0)) + nf_val


        # Detect timestamps if present
        # Accept several headers: first_detected/last_detected or 'First Detected' etc.
        fd = _get_by_contains(row, "first detected")
        ld = _get_by_contains(row, "last detected")
        if fd and (not cur["first_detected"] or fd < cur["first_detected"]):
            cur["first_detected"] = fd
        if ld and (not cur["last_detected"] or ld > cur["last_detected"]):
            cur["last_detected"] = ld

        if is_deleted and peer_type == "ip":
            cur["peer_deleted_workload"] = True
    # Stream flows (avoid loading large CSVs in memory)
    t0 = time.perf_counter()

    out_total_main = 0
    in_total_main = 0
    out_total_raw = 0  # only used when scanning RAW egress for NS managed candidates

    out_kept = 0
    in_kept = 0
    out_kept_zone = 0
    out_kept_ns = 0

    def _iter_rows(p: Path):
        return _iter_csv_rows_stream(p, debug=False)

    if not network_zone_nets:
        # No zone filtering: process all rows from selected sources
        for r in _iter_rows(out_path):
            out_total_main += 1
            add_flow("egress", r)
            out_kept += 1

        for r in _iter_rows(in_path):
            in_total_main += 1
            add_flow("ingress", r)
            in_kept += 1
    else:
        # Network-zone enabled: keep EW strictly in-zone (both endpoints),
        # and a limited NS subset (egress -> managed labels) for proposal purposes.
        def _is_ew(row: Dict[str, str]) -> bool:
            return _flow_row_is_strictly_in_zone(row, nets_v4, nets_v6)

        def _is_ns_managed(row: Dict[str, str]) -> bool:
            return _flow_row_is_ns_egress_to_managed_labels(row, nets_v4, nets_v6)

        if using_zone_files:
            # Zone files are already EW-only
            for r in _iter_rows(out_path):
                out_total_main += 1
                add_flow("egress", r)
                out_kept += 1
                out_kept_zone += 1

            for r in _iter_rows(in_path):
                in_total_main += 1
                add_flow("ingress", r)
                in_kept += 1

            # Scan RAW egress for NS -> remote managed destinations
            if raw_out_path_for_ns and raw_out_path_for_ns.exists():
                for r in _iter_rows(raw_out_path_for_ns):
                    out_total_raw += 1
                    if not _is_ns_managed(r):
                        continue
                    add_flow("egress", r)
                    out_kept += 1
                    out_kept_ns += 1
        else:
            # We are using RAW flows as sources -> filter on the fly
            for r in _iter_rows(out_path):
                out_total_main += 1
                if _is_ew(r):
                    add_flow("egress", r)
                    out_kept += 1
                    out_kept_zone += 1
                    continue
                if _is_ns_managed(r):
                    add_flow("egress", r)
                    out_kept += 1
                    out_kept_ns += 1

            for r in _iter_rows(in_path):
                in_total_main += 1
                if not _is_ew(r):
                    continue
                add_flow("ingress", r)
                in_kept += 1

        # Keep a compact, actionable log line (no branches)
        if using_zone_files:
            info(
                "network-zone filtering applied: "
                f"out_EW={out_kept_zone}/{out_total_main} "
                f"+ ns_managed={out_kept_ns}/{out_total_raw} "
                f"in_EW={in_kept}/{in_total_main}"
            )
        else:
            info(
                "network-zone filtering applied: "
                f"out_kept={out_kept}/{out_total_main} (zone={out_kept_zone} + ns_managed={out_kept_ns}) "
                f"in_kept={in_kept}/{in_total_main}"
            )

    info(f"aggregation done: groups={len(agg)} in {time.perf_counter() - t0:.2f}s")

    rows = [v for v in agg.values() if int(v.get("num_flows", 0)) >= int(min_flows)]
    info(f"after min_flows filter: rows={len(rows)} (min_flows={min_flows})")

    # classification + multi-match handling
    def cat_priority(cat: str) -> int:
        return {CAT_BOUQUET: 0, CAT_BIZ_IN: 1, CAT_BIZ_OUT: 2}.get((cat or "").strip(), 9)

    counts = {
        INFO_DELETED: 0,
        INFO_BOUQUET: 0,
        INFO_BIN: 0,
        INFO_BOUT: 0,
        INFO_NOMATCH: 0,
        "Optimize": 0,
    }

    for p in rows:
        # defaults
        p["Info"] = ""
        p["Action"] = ""

        # 1) Deleted workload override
        if (p.get("peer_type") == "ip") and bool(p.get("peer_deleted_workload")):
            p["Info"] = INFO_DELETED
            p["Action"] = ACTION_NO
            counts[INFO_DELETED] += 1
            continue

        # 2) find ALL matching rules
        matches: List[Dict[str, str]] = []
        for rule in rules_applicables or []:
            try:
                if proposal_matches_rule(p, rule, lg_map):
                    matches.append(rule)
            except Exception as e:
                dbg(debug, f"rule match error ignored: {e}")

        if not matches:
            p["Info"] = INFO_NOMATCH
            p["Action"] = ACTION_CREATE
            counts[INFO_NOMATCH] += 1
            continue

        # 3) choose primary by category priority (and stable tie-breakers)
        matches_sorted = sorted(
            matches,
            key=lambda r: (
                cat_priority(r.get("rule_category", "")),
                (r.get("ruleset_name", "") or ""),
                (r.get("rule_description", "") or ""),
            ),
        )
        primary = matches_sorted[0]
        others = matches_sorted[1:]

        primary_cat = (primary.get("rule_category", "") or "").strip()
        p["matched_rule_category"] = primary_cat

        if primary_cat == CAT_BOUQUET:
            p["Info"] = INFO_BOUQUET
            counts[INFO_BOUQUET] += 1
        elif primary_cat == CAT_BIZ_IN:
            p["Info"] = INFO_BIN
            counts[INFO_BIN] += 1
        elif primary_cat == CAT_BIZ_OUT:
            p["Info"] = INFO_BOUT
            counts[INFO_BOUT] += 1
        else:
            # unknown category, still treat as matched (keep info empty)
            p["Info"] = primary_cat or "Matched"
            counts[p["Info"]] = counts.get(p["Info"], 0) + 1

        p["ruleset_name"] = primary.get("ruleset_name", "") or ""
        p["rule_description"] = primary.get("rule_description", "") or ""

        # primary selectors (quick lookup)
        p["primary_src_iplists"] = primary.get("src_iplists_raw", "") or ""
        p["primary_src_labels"]  = primary.get("src_labels_raw", "") or ""
        p["primary_src_groups"]  = primary.get("src_groups_raw", "") or ""
        p["primary_dst_iplists"] = primary.get("dst_iplists_raw", "") or ""
        p["primary_dst_labels"]  = primary.get("dst_labels_raw", "") or ""
        p["primary_dst_groups"]  = primary.get("dst_groups_raw", "") or ""

        # redundant selectors (multiline)
        if others:
            p["Action"] = ACTION_OPTIM
            counts["Optimize"] += 1

            def join_field(getter) -> str:
                vals = [str(getter(r) or "").strip() for r in others]
                vals = [v for v in vals if v]
                return "\n".join(vals)

            p["redundant_rule_categories"] = join_field(lambda r: r.get("rule_category", ""))
            p["redundant_ruleset_names"] = join_field(lambda r: r.get("ruleset_name", ""))
            p["redundant_rule_descriptions"] = join_field(lambda r: r.get("rule_description", ""))
            p["redundant_src_iplists"] = join_field(lambda r: r.get("src_iplists_raw", ""))
            p["redundant_src_labels"]  = join_field(lambda r: r.get("src_labels_raw", ""))
            p["redundant_src_groups"]  = join_field(lambda r: r.get("src_groups_raw", ""))
            p["redundant_dst_iplists"] = join_field(lambda r: r.get("dst_iplists_raw", ""))
            p["redundant_dst_labels"]  = join_field(lambda r: r.get("dst_labels_raw", ""))
            p["redundant_dst_groups"]  = join_field(lambda r: r.get("dst_groups_raw", ""))
        else:
            p["Action"] = ACTION_NO

    info("classification summary: " + ", ".join(f"{k}={v}" for k, v in counts.items()))

    if enable_avoid_label_pairs and avoid_label_pairs:
        info(
            "avoid-label-pairs: "
            + ", ".join(
                [
                    f"apps={len(avoid_label_pairs)}",
                    f"hits={_avoid_stats.get('hits', 0)}",
                    f"resolved_iplist={_avoid_stats.get('resolved_to_iplist', 0)}",
                    f"fallback_ip={_avoid_stats.get('fallback_to_ip', 0)}",
                    f"no_ip={_avoid_stats.get('no_ip', 0)}",
                    f"kub_resolver={'yes' if kub_iplist_resolver else 'no'}",
                ]
            )
        )

    # suggested_ruleset: simple helper (kept)
    for p in rows:
        a_app = (p.get("anchor_app") or "").strip()
        a_env = (p.get("anchor_env") or "").strip()
        if a_app and a_env:
            p["suggested_ruleset"] = f"{a_app}-{a_env}"
        elif a_app:
            p["suggested_ruleset"] = a_app

    # sort for stability
    rows_sorted = sorted(rows, key=lambda r: (
        r.get("direction",""),
        r.get("anchor_app",""), r.get("anchor_env",""), r.get("anchor_role",""),
        r.get("peer_type",""), r.get("peer_value",""),
        r.get("proto",""), int(r.get("port",0)),
    ))

    action_plan = [r for r in rows_sorted if (r.get("Action") in (ACTION_CREATE, ACTION_OPTIM))]
    info(f"propose() output rows={len(rows_sorted)} action_plan_rows={len(action_plan)}")
    if not rows_sorted:
        warn("No aggregated rows produced — check scope, flow file names, and time window.")
    return rows_sorted, action_plan

# -----------------------------------------------------------------------------
# CSV + Excel output
# -----------------------------------------------------------------------------

OUTPUT_HEADER = [
    "direction",
    "anchor_app", "anchor_env", "anchor_role", "anchor_loc", "anchor_OS",
    "peer_type", "peer_value",
    "proto", "port", "num_flows", "first_detected", "last_detected",
    "suggested_ruleset",
    "Info", "Action",
    "matched_rule_category",
    "primary_src_iplists", "primary_src_labels", "primary_src_groups",
    "primary_dst_iplists", "primary_dst_labels", "primary_dst_groups",
    "redundant_rule_categories",
    "redundant_ruleset_names",
    "redundant_rule_descriptions",
    "redundant_src_iplists", "redundant_src_labels", "redundant_src_groups",
    "redundant_dst_iplists", "redundant_dst_labels", "redundant_dst_groups",
    # always last:
    "ruleset_name", "rule_description",
]

WRAP_COLUMNS = {
    "primary_src_iplists", "primary_src_labels", "primary_src_groups",
    "primary_dst_iplists", "primary_dst_labels", "primary_dst_groups",
    "redundant_rule_categories",
    "redundant_ruleset_names",
    "redundant_rule_descriptions",
    "redundant_src_iplists", "redundant_src_labels", "redundant_src_groups",
    "redundant_dst_iplists", "redundant_dst_labels", "redundant_dst_groups",
    "rule_description",
}

TO_INVESTIGATE_PREFIXES = ("NZ0_", "NZ1_")
TO_INVESTIGATE_EGRESS_PREFIXES = ("KUB_", "LBI_", "LBO_")

def _split_iplist_values(value: str) -> List[str]:
    if not value:
        return []
    tokens = re.split(r"[,\n;|]+", value)
    out: List[str] = []
    for token in tokens:
        cleaned = token.strip()
        if cleaned:
            out.append(cleaned)
    return out

def _row_matches_to_investigate(row: Dict[str, Any]) -> bool:
    direction = str(row.get("direction") or "").strip().lower()
    if direction not in ("egress", "ingress"):
        return False

    prefixes = TO_INVESTIGATE_PREFIXES
    if direction == "egress":
        prefixes = TO_INVESTIGATE_PREFIXES + TO_INVESTIGATE_EGRESS_PREFIXES

    candidates: List[str] = []
    peer_type = str(row.get("peer_type") or "").strip().lower()
    peer_value = str(row.get("peer_value") or "").strip()
    if peer_type == "iplist" and peer_value:
        candidates.append(peer_value)

    if direction == "egress":
        candidates += _split_iplist_values(str(row.get("primary_dst_iplists") or ""))
        candidates += _split_iplist_values(str(row.get("redundant_dst_iplists") or ""))
    else:
        candidates += _split_iplist_values(str(row.get("primary_src_iplists") or ""))
        candidates += _split_iplist_values(str(row.get("redundant_src_iplists") or ""))

    return any(val.startswith(prefixes) for val in candidates)

def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUTPUT_HEADER)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in OUTPUT_HEADER})
    info(f"CSV written: {path} rows={len(rows)}")

def append_excel(excel_path: Path, sheet_name: str, rows: List[Dict[str, Any]]) -> None:
    """
    Replace sheet with styling + row fills.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    wb = load_workbook(filename=str(excel_path))

    # remove sheet if exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    # fills
    GREY   = PatternFill("solid", fgColor="D9D9D9")  # light grey
    BLUE   = PatternFill("solid", fgColor="D9E1F2")  # light blue
    GREEN  = PatternFill("solid", fgColor="E2EFDA")  # light green
    PINK   = PatternFill("solid", fgColor="FCE4D6")  # light pink
    YELLOW = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    ORANGE = PatternFill("solid", fgColor="F8CBAD")  # light orange

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # write header
    ws.append(OUTPUT_HEADER)
    for c_i, name in enumerate(OUTPUT_HEADER, start=1):
        cell = ws.cell(row=1, column=c_i)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # write data
    for r in rows:
        ws.append([r.get(h, "") for h in OUTPUT_HEADER])

    # styling rows
    for r_i in range(2, ws.max_row + 1):
        info_val = (ws.cell(row=r_i, column=OUTPUT_HEADER.index("Info") + 1).value or "").strip()
        action_val = (ws.cell(row=r_i, column=OUTPUT_HEADER.index("Action") + 1).value or "").strip()

        fill = None
        if _row_matches_to_investigate({h: ws.cell(row=r_i, column=OUTPUT_HEADER.index(h) + 1).value for h in OUTPUT_HEADER}):
            fill = YELLOW
        elif info_val == INFO_DELETED:
            fill = GREY
        elif action_val == ACTION_OPTIM:
            fill = ORANGE
        elif action_val == ACTION_CREATE and info_val == INFO_NOMATCH:
            fill = YELLOW
        elif info_val == INFO_BOUQUET:
            fill = BLUE
        elif info_val == INFO_BIN:
            fill = GREEN
        elif info_val == INFO_BOUT:
            fill = PINK

        for c_i, col_name in enumerate(OUTPUT_HEADER, start=1):
            cell = ws.cell(row=r_i, column=c_i)
            if fill:
                cell.fill = fill
            cell.border = border
            if col_name in WRAP_COLUMNS:
                cell.alignment = wrap_align
            else:
                cell.alignment = normal_align

            # numeric formatting
            if col_name in ("port", "num_flows", "sum_num_flows"):
                try:
                    cell.number_format = "0"
                except Exception:
                    pass

    # Auto-fit (simple heuristic)
    for c_i, col_name in enumerate(OUTPUT_HEADER, start=1):
        max_len = len(col_name)
        for r_i in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r_i, column=c_i).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, min(80, max(len(line) for line in s.splitlines()) if s else 0))
        ws.column_dimensions[get_column_letter(c_i)].width = min(60, max(10, max_len + 2))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(str(excel_path))

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

# ------------------------------ Proposed rules (intra-app) ------------------------------
PROPOSED_RULES_SHEET = "Proposed rules"
PROPOSED_RULES_HEADER = ["Direction", "Strategy", "Source", "Destination", "Services", "sum_num_flows", "Rule Section", "Ruleset"]

# New (additive) sheet for evolving proposal formatting without breaking the stable "Proposed rules" contract.
PROPOSED_RULES1_SHEET = "Proposed rules1"
PROPOSED_RULES1_HEADER = [
    "Direction",
    "Strategy",
    "Source",
    "Destination",
    "Services",
    "num_aggregated_rows",
    "sum_num_flows",
    "Rule Section",
    "Comment",
    "Ruleset",
]
def _parse_ports_list(s: str) -> Set[int]:
    """Parse ports list from a string like '22,3389; 8443' -> {22,3389,8443}."""
    out: Set[int] = set()
    if not s:
        return out
    # Accept separators: comma, semicolon, pipe, whitespace
    parts = re.split(r"[\s,;|]+", str(s))
    for p in parts:
        p = (p or "").strip()
        if not p:
            continue
        # tolerate 'tcp/22' etc -> keep only the number
        m = re.search(r"(\d+)", p)
        if not m:
            continue
        try:
            n = int(m.group(1))
        except Exception:
            continue
        if 0 < n <= 65535:
            out.add(n)
    return out


# -----------------------------------------------------------------------------
# Blacklist parsing (carto.conf port lists)
# -----------------------------------------------------------------------------


def _group_pr1_ingress_finegrained_by_src_dst(
    rows: List[Dict[str, Any]],
    finegrained_single_ports: Optional[Dict[str, List[Tuple[int, int]]]] = None,
) -> List[Dict[str, Any]]:
    """
    Proposed rules1 readability improvement.

    For ingress + finegrained, group rows that share the same Source and Destination (and same Rule Section / Ruleset / Comment)
    by concatenating all observed service tokens into one cell (separated by ';') and summing sum_num_flows.
    Ports listed in carto.conf PORTS_ADMIN/PORTS_TO_ERADICATE/PORTS_TO_CONTROL are never aggregated (one port per rule).

    This does NOT affect the stable "Proposed rules" sheet contract; it only changes Proposed rules1 output ordering/aggregation.
    """
    if not rows:
        return rows

    out: List[Dict[str, Any]] = []
    aggs: Dict[Tuple[str, str, str, str, str, str, str], Dict[str, Any]] = {}

    def _to_int(v: Any) -> int:
        try:
            if v is None:
                return 0
            if isinstance(v, bool):
                return int(v)
            if isinstance(v, (int,)):
                return int(v)
            if isinstance(v, float):
                return int(v)
            s = str(v).strip()
            if not s:
                return 0
            return int(float(s))
        except Exception:
            return 0

    def _iter_service_tokens(cell: Any):
        s = str(cell or "").strip()
        if not s:
            return
        if s.strip().lower() == "all services":
            yield "All Services"
            return
        for tok in re.split(r"[;\n]+", s):
            tok = (tok or "").strip()
            if tok:
                yield tok

    def _is_single_port_tok(tok: str) -> bool:
        m = re.match(r"(?i)^(tcp|udp)\/(\d+)$", tok.strip())
        if not m:
            return False
        proto = m.group(1).lower()
        port = int(m.group(2))
        return _is_port_in_intervals(finegrained_single_ports, proto, port)

    for r in rows:
        direction = str(r.get("Direction", "") or "")
        strategy = str(r.get("Strategy", "") or "").strip().lower()
        rule_section = str(r.get("Rule Section", "") or "")
        # We only group ingress finegrained proposals (both intra-scope and Extrascope) where services are proto/port tokens.
        if direction == "ingress" and strategy == "finegrained" and rule_section in {"intra-scope", "Extrascope"}:
            service_tokens = list(_iter_service_tokens(r.get("Services")))
            if any(_is_single_port_tok(tok) for tok in service_tokens):
                out.append(r)
                continue
            source = str(r.get("Source", "") or "")
            dest = str(r.get("Destination", "") or "")
            ruleset = str(r.get("Ruleset", "") or "")
            comment_key = str(r.get("Comment", "") or "")
            key = (direction, "finegrained", source, dest, rule_section, ruleset, comment_key)

            if key not in aggs:
                base = dict(r)
                base["sum_num_flows"] = 0
                base["_svc_all"] = False
                base["_svc_set"] = set()
                aggs[key] = base
                out.append(base)

            a = aggs[key]
            a["sum_num_flows"] = _to_int(a.get("sum_num_flows")) + _to_int(r.get("sum_num_flows"))

            for tok in service_tokens:
                if tok == "All Services":
                    a["_svc_all"] = True
                else:
                    a["_svc_set"].add(tok)
        else:
            out.append(r)

    # Finalize grouped rows in-place (keep first-seen order)
    for r in out:
        if "_svc_set" in r:
            if r.get("_svc_all"):
                r["Services"] = "All Services"
            else:
                toks = sorted(list(r.get("_svc_set") or []), key=_service_sort_key)
                r["Services"] = ";".join(toks)
            r.pop("_svc_set", None)
            r.pop("_svc_all", None)

    return out


def _read_conf_kv(conf_path: Optional[Path], debug: bool=False) -> Dict[str, str]:
    """Parse simple KEY=VALUE pairs from carto.conf (ignores comments and empty lines)."""
    if not conf_path or not conf_path.exists():
        return {}
    kv: Dict[str, str] = {}
    with conf_path.open("r", encoding="utf-8", errors="replace") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            # strip inline comments
            if "#" in line:
                line = line.split("#", 1)[0].strip()
            if "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip()
            if v.startswith('"') and v.endswith('"') and len(v) >= 2:
                v = v[1:-1]
            if k:
                kv[k] = v
    return kv



def load_avoid_label_pairs(conf_path: Optional[Path], debug: bool=False) -> Set[str]:
    """Load AVOID_LABEL_PAIRS from carto.conf.

    Expected format in carto.conf:
      AVOID_LABEL_PAIRS=app=APP_A;app=APP_B
    Only the app values are kept (deterministic set).
    """
    kv = _read_conf_kv(conf_path, debug=debug)
    raw = (kv.get("AVOID_LABEL_PAIRS") or "").strip()
    if not raw:
        return set()

    # Allow comma/semicolon separated tokens.
    tokens = [t.strip() for t in re.split(r"[;,]+", raw) if t.strip()]
    apps: Set[str] = set()
    for t in tokens:
        if "=" in t:
            k, v = t.split("=", 1)
            if k.strip().lower() == "app" and v.strip():
                apps.add(v.strip())
        else:
            # Backward tolerant: bare app value
            apps.add(t)
    dbg(debug, f"AVOID_LABEL_PAIRS apps loaded: {len(apps)}")
    return apps


def _extract_networks_from_include(include_value: str) -> List[ipaddress._BaseNetwork]:
    """Parse export_iplists.csv 'include' column into ip_network objects (best effort).

    We intentionally support the formats seen in Illumio exports, including tokens like:
      - 171.84.192.0/21#GEN1
      - 10.0.0.5
      - !192.168.0.0/16   (negated tokens are ignored for zone-network extraction)

    Supported:
      - IPv4/IPv6 CIDR
      - Single IP (→ /32 or /128)
      - IPv4 range "a.b.c.d-e.f.g.h" (summarized into CIDRs)
    Everything else is ignored.
    """
    s = (include_value or "").strip()
    if not s:
        return []

    # Normalize separators commonly seen in exports.
    s = s.replace("\r", "\n")
    s = s.replace(" and ", ";").replace("\n", ";")

    # Split primarily on ';', ',', '|', tabs. We will also split on spaces later to be safe.
    parts = re.split(r"[;,|\t]+", s)

    nets: Dict[str, ipaddress._BaseNetwork] = {}

    def _add_net(net: ipaddress._BaseNetwork) -> None:
        nets[str(net)] = net  # de-dup deterministically

    for part in parts:
        raw = (part or "").strip().strip('"').strip("'")
        if not raw:
            continue

        # Some exports may have multiple tokens separated by spaces inside one cell chunk.
        for token in re.split(r"\s+", raw):
            tok = (token or "").strip().strip('"').strip("'")
            if not tok:
                continue

            # Negated tokens (e.g. in ZNOT_* lists) should not contribute to zone nets.
            if tok.startswith("!"):
                continue

            # Drop tags/comments (e.g. '#GEN1').
            if "#" in tok:
                tok = tok.split("#", 1)[0].strip()
            if not tok:
                continue

            # IPv4 range: a.b.c.d-e.f.g.h
            m_range = re.fullmatch(r"(\d{1,3}(?:\.\d{1,3}){3})\s*-\s*(\d{1,3}(?:\.\d{1,3}){3})", tok)
            if m_range:
                try:
                    a = ipaddress.ip_address(m_range.group(1))
                    b = ipaddress.ip_address(m_range.group(2))
                    if a.version == 4 and b.version == 4:
                        for net in ipaddress.summarize_address_range(a, b):
                            _add_net(net)
                except Exception:
                    pass
                continue

            # CIDR or single IP
            try:
                if "/" in tok:
                    net = ipaddress.ip_network(tok, strict=False)
                else:
                    ip = ipaddress.ip_address(tok)
                    suffix = "/32" if ip.version == 4 else "/128"
                    net = ipaddress.ip_network(tok + suffix, strict=False)
                _add_net(net)
            except Exception:
                # As a last-resort, try to extract the first IPv4 token from the string.
                m = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})(?:/(\d{1,2}))?", tok)
                if not m:
                    continue
                ip_s = m.group(1)
                mask = m.group(2)
                try:
                    if mask:
                        _add_net(ipaddress.ip_network(f"{ip_s}/{mask}", strict=False))
                    else:
                        ip = ipaddress.ip_address(ip_s)
                        suffix = "/32" if ip.version == 4 else "/128"
                        _add_net(ipaddress.ip_network(ip_s + suffix, strict=False))
                except Exception:
                    continue

    return [nets[k] for k in sorted(nets.keys())]


# ------------------------------ Network-zone (East-West) filtering ------------------------------
def _resolve_network_zone_iplists(raw_dir: Path, zone_iplist_name: str, debug: bool = False) -> Tuple[str, str, List[ipaddress._BaseNetwork]]:
    """
    Resolve the IPLists needed for --network-zone:
      - zone IPList: <zone_iplist_name>  (must exist)
      - complement IPList: ZNOT_<zone_iplist_name> (must exist)

    Returns: (zone_name, znot_name, zone_networks)
    """
    zone = (zone_iplist_name or "").strip()
    if not zone:
        raise ValueError("--network-zone expects a non-empty IPList name")

    ipl_csv = raw_dir / "export_iplists.csv"
    if not ipl_csv.exists():
        raise FileNotFoundError(f"{ipl_csv} not found (required for --network-zone)")

    rows, cols = _iter_csv_rows(ipl_csv)
    if not rows or not cols:
        raise ValueError(f"{ipl_csv} is empty/unreadable (required for --network-zone)")

    c_name = _pick_col(cols, "name", "iplist_name")
    c_inc = _pick_col(cols, "include", "includes", "ip_ranges", "cidrs")
    if not (c_name and c_inc):
        raise ValueError(f"{ipl_csv} missing required columns for --network-zone (need name/include)")

    zone_row = None
    znot_name = f"ZNOT_{zone}"
    znot_found = False
    for r in rows:
        n = str(r.get(c_name, "") or "").strip()
        if n == zone:
            zone_row = r
        elif n == znot_name:
            znot_found = True

    if zone_row is None:
        raise ValueError(f"--network-zone '{zone}' not found in {ipl_csv.name}")
    if not znot_found:
        raise ValueError(f"--network-zone requires complement IPList '{znot_name}' (missing in {ipl_csv.name})")

    nets = _extract_networks_from_include(str(zone_row.get(c_inc, "") or ""))
    if not nets:
        raise ValueError(f"IPList '{zone}' has empty/unsupported include content in {ipl_csv.name}; cannot filter flows")

    dbg(debug, f"network-zone resolved: zone='{zone}' nets={len(nets)} znot='{znot_name}'")
    return zone, znot_name, nets


def _extract_first_ip_token(value: str) -> str:
    """Best-effort extraction of the first valid IP address from a value (may contain multiple IPs or IP:port)."""
    s = (value or "").strip()
    if not s:
        return ""

    s = s.replace(",", " ").replace(";", " ").replace("|", " ").replace("\n", " ")
    for tok in s.split():
        t = tok.strip().strip('"').strip("'").strip()
        if not t:
            continue

        # Handle [IPv6]:port
        if t.startswith("[") and "]" in t:
            inside = t[1:t.index("]")]
            try:
                ipaddress.ip_address(inside)
                return inside
            except Exception:
                pass

        # Handle IPv4:port
        if t.count(".") == 3 and ":" in t:
            t2 = t.split(":", 1)[0].strip()
            try:
                ipaddress.ip_address(t2)
                return t2
            except Exception:
                pass

        # Plain IP
        try:
            ipaddress.ip_address(t)
            return t
        except Exception:
            continue

    return ""


def _split_nets_by_family(nets: List[ipaddress._BaseNetwork]) -> Tuple[List[ipaddress.IPv4Network], List[ipaddress.IPv6Network]]:
    v4: List[ipaddress.IPv4Network] = []
    v6: List[ipaddress.IPv6Network] = []
    for n in nets or []:
        try:
            if getattr(n, "version", 4) == 6:
                v6.append(n)  # type: ignore[arg-type]
            else:
                v4.append(n)  # type: ignore[arg-type]
        except Exception:
            continue
    return v4, v6


def _ip_in_any_net(ip_s: str, nets_v4: List[ipaddress.IPv4Network], nets_v6: List[ipaddress.IPv6Network]) -> bool:
    try:
        ip = ipaddress.ip_address(ip_s)
    except Exception:
        return False

    if ip.version == 6:
        for n in nets_v6:
            if ip in n:
                return True
        return False

    for n in nets_v4:
        if ip in n:
            return True
    return False


def _filter_flows_strictly_in_zone(
    flows: List[Dict[str, str]],
    nets_v4: List[ipaddress.IPv4Network],
    nets_v6: List[ipaddress.IPv6Network],
) -> List[Dict[str, str]]:
    """Keep only flows where BOTH endpoints IPs are inside the provided networks."""
    out: List[Dict[str, str]] = []
    for r in flows:
        src_ip_raw = _get_by_contains(r, "source ip") or _get_by_contains(r, "src ip") or ""
        dst_ip_raw = _get_by_contains(r, "destination ip") or _get_by_contains(r, "dst ip") or ""
        src_ip = _extract_first_ip_token(str(src_ip_raw))
        dst_ip = _extract_first_ip_token(str(dst_ip_raw))
        if not (src_ip and dst_ip):
            continue
        if _ip_in_any_net(src_ip, nets_v4, nets_v6) and _ip_in_any_net(dst_ip, nets_v4, nets_v6):
            out.append(r)
    return out



def _filter_flows_ns_egress_to_managed_labels(
    flows: List[Dict[str, str]],
    nets_v4: List[ipaddress.IPv4Network],
    nets_v6: List[ipaddress.IPv6Network],
) -> List[Dict[str, str]]:
    """Keep only **egress** North/South flows where:
      - Source IP is inside the provided networks (in-zone)
      - Destination IP is outside the provided networks (out-of-zone)
      - Destination peer is **Managed labels** (app prefix ends with 'M')

    Used ONLY when --network-zone is active to avoid missing rules required
    on the remote managed destination side.
    """
    out: List[Dict[str, str]] = []
    for r in flows:
        src_ip_raw = _get_by_contains(r, "source ip") or _get_by_contains(r, "src ip") or ""
        dst_ip_raw = _get_by_contains(r, "destination ip") or _get_by_contains(r, "dst ip") or ""
        src_ip = _extract_first_ip_token(str(src_ip_raw))
        dst_ip = _extract_first_ip_token(str(dst_ip_raw))
        if not (src_ip and dst_ip):
            continue

        # North/South: src in zone, dst outside
        if not _ip_in_any_net(src_ip, nets_v4, nets_v6):
            continue
        if _ip_in_any_net(dst_ip, nets_v4, nets_v6):
            continue

        # Managed destination labels?
        dst_labels = _labels_from_flow_row(r, "destination")
        dst_app = (dst_labels.get("app") or "").strip()
        pref = _app_prefix(dst_app)
        if not (pref and pref.endswith("M")):
            continue

        out.append(r)
    return out

def _flow_row_is_strictly_in_zone(
    r: Dict[str, str],
    nets_v4: List[ipaddress.IPv4Network],
    nets_v6: List[ipaddress.IPv6Network],
) -> bool:
    """Row-level equivalent of _filter_flows_strictly_in_zone()."""
    src_raw = _get_by_contains(r, "source ip") or ""
    dst_raw = _get_by_contains(r, "destination ip") or ""
    src_ip = _extract_first_ip_token(src_raw)
    dst_ip = _extract_first_ip_token(dst_raw)
    if not (src_ip and dst_ip):
        return False
    return _ip_in_any_net(src_ip, nets_v4, nets_v6) and _ip_in_any_net(dst_ip, nets_v4, nets_v6)


def _flow_row_is_ns_egress_to_managed_labels(
    r: Dict[str, str],
    nets_v4: List[ipaddress.IPv4Network],
    nets_v6: List[ipaddress.IPv6Network],
) -> bool:
    """Row-level equivalent of _filter_flows_ns_egress_to_managed_labels()."""
    src_raw = _get_by_contains(r, "source ip") or ""
    dst_raw = _get_by_contains(r, "destination ip") or ""
    src_ip = _extract_first_ip_token(src_raw)
    dst_ip = _extract_first_ip_token(dst_raw)
    if not (src_ip and dst_ip):
        return False

    # North/South: src in zone, dst outside
    if not _ip_in_any_net(src_ip, nets_v4, nets_v6):
        return False
    if _ip_in_any_net(dst_ip, nets_v4, nets_v6):
        return False

    # Managed destination labels?
    dst_labels = _labels_from_flow_row(r, "destination")
    dst_app = (dst_labels.get("app") or "").strip()
    pref = _app_prefix(dst_app)
    return bool(pref and pref.endswith("M"))


def build_kub_iplist_network_index(raw_dir: Path, debug: bool=False) -> List[Tuple[ipaddress._BaseNetwork, str]]:
    """Build a deterministic (network → iplist_name) index for KUB_* IPLISTS."""
    path = raw_dir / "export_iplists.csv"
    if not path.exists():
        warn(f"export_iplists.csv not found: {path}")
        return []

    rows, cols = _iter_csv_rows(path, debug=debug)
    if not rows:
        warn(f"export_iplists.csv is empty or unreadable: {path}")
        return []

    c_name = _pick(cols, "name")
    c_inc = _pick(cols, "include")
    if not c_name or not c_inc:
        warn(f"export_iplists.csv missing required columns (need name/include): {path}")
        return []

    out: List[Tuple[ipaddress._BaseNetwork, str]] = []
    iplist_names: Set[str] = set()
    for r in rows:
        name = (r.get(c_name) or "").strip()
        if not name.startswith("KUB_"):
            continue
        iplist_names.add(name)
        inc = (r.get(c_inc) or "").strip()
        for net in _extract_networks_from_include(inc):
            out.append((net, name))

    out_sorted = sorted(out, key=lambda t: (t[0].version, -t[0].prefixlen, t[1], str(t[0])))
    dbg(debug, f"KUB_* IPLIST index built: {len(iplist_names)} lists, {len(out_sorted)} networks")
    return out_sorted


def make_kub_iplist_resolver(index: List[Tuple[ipaddress._BaseNetwork, str]]) -> Callable[[str], str]:
    """Return a resolver(ip_str) -> iplist_name (best match) or ""."""
    def _resolve(ip_str: str) -> str:
        s = (ip_str or "").strip()
        if not s:
            return ""
        try:
            ip = ipaddress.ip_address(s)
        except Exception:
            return ""
        for net, name in index:
            if net.version != ip.version:
                continue
            if ip in net:
                return name
        return ""
    return _resolve

def _parse_proto_port_token(tok: str) -> List[Tuple[str, int, int]]:
    """
    Parse a token like:
      - TCP/22
      - UDP/69
      - TCP/137-138
    Returns a list of (proto, start, end). 'proto' is lowercase ('tcp'/'udp').
    For a bare number like '22', returns both TCP and UDP ranges.
    Unsupported tokens (ICMP/IGMP/...) return [].
    """
    tok = (tok or "").strip()
    if not tok:
        return []
    if tok.isdigit():
        p = int(tok)
        if 1 <= p <= 65535:
            return [("tcp", p, p), ("udp", p, p)]
        return []
    m = re.match(r"(?i)^(tcp|udp)\s*/\s*(\d+)(?:\s*-\s*(\d+))?$", tok)
    if not m:
        return []
    proto = m.group(1).lower()
    a = int(m.group(2))
    b = int(m.group(3)) if m.group(3) else a
    if a > b:
        a, b = b, a
    if not (1 <= a <= 65535 and 1 <= b <= 65535):
        return []
    return [(proto, a, b)]


def _merge_intervals(intervals: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    """Merge overlapping/adjacent integer intervals."""
    if not intervals:
        return []
    intervals = sorted(intervals, key=lambda t: (t[0], t[1]))
    out: List[Tuple[int, int]] = []
    cur_a, cur_b = intervals[0]
    for a, b in intervals[1:]:
        if a <= cur_b + 1:
            cur_b = max(cur_b, b)
        else:
            out.append((cur_a, cur_b))
            cur_a, cur_b = a, b
    out.append((cur_a, cur_b))
    return out


def _subtract_interval(base: List[Tuple[int, int]], rem: Tuple[int, int]) -> List[Tuple[int, int]]:
    """Subtract rem interval from base intervals."""
    ra, rb = rem
    out: List[Tuple[int, int]] = []
    for a, b in base:
        if rb < a or ra > b:
            out.append((a, b))
            continue
        if a < ra:
            out.append((a, ra - 1))
        if rb < b:
            out.append((rb + 1, b))
    return out


def _complement_from_blacklist(black: Dict[str, List[Tuple[int, int]]]) -> Dict[str, List[Tuple[int, int]]]:
    """Build allow-intervals as complement of blacklist over [1..65535] for tcp/udp."""
    out: Dict[str, List[Tuple[int, int]]] = {}
    for proto in ("tcp", "udp"):
        base = [(1, 65535)]
        for rem in black.get(proto, []):
            base = _subtract_interval(base, rem)
        out[proto] = base
    return out


def _render_intervals_as_services(allow: Dict[str, List[Tuple[int, int]]], include_misc: bool=True) -> str:
    """Render allow intervals as a semicolon-separated services string."""
    parts: List[str] = []
    if include_misc:
        parts.extend(["IGMP", "ICMP"])
    for proto in ("tcp", "udp"):
        pfx = proto.upper()
        for a, b in allow.get(proto, []):
            if a == b:
                parts.append(f"{pfx}/{a}")
            else:
                parts.append(f"{pfx}/{a}-{b}")
    return ";".join(parts)


def _expand_blacklist_sources(conf_path: Optional[Path], spec: str) -> Tuple[Dict[str, List[Tuple[int, int]]], List[str]]:
    """
    Expand --ports-to-blacklist when strategy=blacklist.
    spec may contain multiple list names (e.g. PORTS_TO_CONTROL,PORTS_ADMIN) and/or inline tokens (TCP/22, 3389).
    Returns:
      - intervals_by_proto: {'tcp': [(a,b),...], 'udp': [(a,b),...]} merged
      - used_list_names: list of config keys referenced (for traceability)
    Behavior:
      - If a token matches a KEY in carto.conf -> expands its value
      - If a token looks like TCP/.. or UDP/.. or a bare number -> parsed inline
      - Otherwise: raises ValueError (fail-fast)
    """
    spec = (spec or "").strip()
    kv = _read_conf_kv(conf_path)
    used: List[str] = []
    intervals: Dict[str, List[Tuple[int, int]]] = {"tcp": [], "udp": []}

    if not spec:
        return {"tcp": [], "udp": []}, []

    toks = [t.strip() for t in re.split(r"[;,\s]+", spec) if t.strip()]
    for t in toks:
        if t in kv:
            used.append(t)
            val = kv[t]
            # split list by ';'
            for item in [x.strip() for x in val.split(";") if x.strip()]:
                for proto, a, b in _parse_proto_port_token(item):
                    intervals[proto].append((a, b))
        else:
            parsed = _parse_proto_port_token(t)
            if parsed:
                for proto, a, b in parsed:
                    intervals[proto].append((a, b))
            else:
                raise ValueError(f"Unknown blacklist token '{t}'. Expected a carto.conf key (e.g. PORTS_TO_CONTROL) or a port token like TCP/22.")

    intervals = {k: _merge_intervals(v) for k, v in intervals.items()}
    return intervals, used


FINEGRAINED_PORT_LIST_KEYS = ("PORTS_ADMIN", "PORTS_TO_ERADICATE", "PORTS_TO_CONTROL")
FINEGRAINED_PORT_LIST_COMMENTS = {
    "PORTS_ADMIN": "Admin port",
    "PORTS_TO_ERADICATE": "Dangerous port (to eradicate)",
    "PORTS_TO_CONTROL": "Dangerous port (to control)",
}


def _load_default_finegrained_single_ports(conf_path: Optional[Path]) -> Tuple[Dict[str, List[Tuple[int, int]]], List[str]]:
    if not conf_path:
        return {"tcp": [], "udp": []}, []
    kv = _read_conf_kv(conf_path)
    keys = [k for k in FINEGRAINED_PORT_LIST_KEYS if k in kv]
    keys = [k for k in ("PORTS_ADMIN", "PORTS_TO_ERADICATE", "PORTS_TO_CONTROL") if k in kv]
    if not keys:
        return {"tcp": [], "udp": []}, []
    spec = ";".join(keys)
    return _expand_finegrained_single_ports(conf_path, spec)


def _load_port_list_intervals(conf_path: Optional[Path]) -> Dict[str, Dict[str, List[Tuple[int, int]]]]:
    if not conf_path:
        return {}
    kv = _read_conf_kv(conf_path)
    out: Dict[str, Dict[str, List[Tuple[int, int]]]] = {}
    for key in FINEGRAINED_PORT_LIST_KEYS:
        val = (kv.get(key) or "").strip()
        if not val:
            continue
        intervals: Dict[str, List[Tuple[int, int]]] = {"tcp": [], "udp": []}
        for item in [x.strip() for x in val.split(";") if x.strip()]:
            for proto, a, b in _parse_proto_port_token(item):
                intervals[proto].append((a, b))
        intervals = {k: _merge_intervals(v) for k, v in intervals.items()}
        out[key] = intervals
    return out


def _port_list_comment_for_services(
    services: Any,
    list_intervals: Dict[str, Dict[str, List[Tuple[int, int]]]],
) -> str:
    s = str(services or "").strip()
    if not s or s.lower() == "all services":
        return ""
    comments: List[str] = []
    for tok in re.split(r"[;\n]+", s):
        tok = (tok or "").strip()
        if not tok:
            continue
        m = re.match(r"(?i)^(tcp|udp)\/(\d+)$", tok)
        if not m:
            continue
        proto = m.group(1).lower()
        port = int(m.group(2))
        for key in FINEGRAINED_PORT_LIST_KEYS:
            intervals = list_intervals.get(key)
            if _is_port_in_intervals(intervals, proto, port):
                comment = FINEGRAINED_PORT_LIST_COMMENTS.get(key, "")
                if comment and comment not in comments:
                    comments.append(comment)
    return "; ".join(comments)


def _expand_finegrained_single_ports(conf_path: Optional[Path], spec: str) -> Tuple[Dict[str, List[Tuple[int, int]]], List[str]]:
    """
    Expand finegrained single-port lists from carto.conf.
    spec may contain multiple list names (e.g. PORTS_ADMIN) and/or inline tokens (TCP/22, 3389).
    Returns:
      - intervals_by_proto: {'tcp': [(a,b),...], 'udp': [(a,b),...]} merged
      - used_list_names: list of config keys referenced (for traceability)
    Behavior:
      - If a token matches a KEY in carto.conf -> expands its value
      - If a token looks like TCP/.. or UDP/.. or a bare number -> parsed inline
      - Otherwise: raises ValueError (fail-fast)
    """
    spec = (spec or "").strip()
    kv = _read_conf_kv(conf_path)
    used: List[str] = []
    intervals: Dict[str, List[Tuple[int, int]]] = {"tcp": [], "udp": []}

    if not spec:
        return {"tcp": [], "udp": []}, []

    toks = [t.strip() for t in re.split(r"[;,\s]+", spec) if t.strip()]
    for t in toks:
        if t in kv:
            used.append(t)
            val = kv[t]
            for item in [x.strip() for x in val.split(";") if x.strip()]:
                for proto, a, b in _parse_proto_port_token(item):
                    intervals[proto].append((a, b))
        else:
            parsed = _parse_proto_port_token(t)
            if parsed:
                for proto, a, b in parsed:
                    intervals[proto].append((a, b))
            else:
                raise ValueError(
                    f"Unknown finegrained port token '{t}'. Expected a carto.conf key (e.g. PORTS_ADMIN) "
                    "or a port token like TCP/22."
                )

    intervals = {k: _merge_intervals(v) for k, v in intervals.items()}
    return intervals, used



def _resolve_blacklist_sources_for_direction(args) -> Dict[str, str]:
    """Resolve blacklist source spec per direction (intra-app/egress/ingress).

    Rules (deterministic):
      - If a per-direction flag is provided, it wins.
      - Otherwise, legacy --ports-to-blacklist applies only if exactly one direction is in blacklist mode.
      - If multiple directions are in blacklist mode and only legacy is provided -> fail-fast.
      - If a direction is in blacklist mode and no sources are provided -> fail-fast.
    Returns:
      dict: { 'intra-app': 'PORTS_TO_CONTROL,...', 'egress': '...', 'ingress': '...' } for active blacklist directions only.
    """
    # Which directions are in blacklist mode?
    active: List[str] = []
    if getattr(args, "strategy_intra_app", "none") == "blacklist":
        active.append("intra-app")
    if getattr(args, "strategy_egress", "none") == "blacklist":
        active.append("egress")
    if getattr(args, "strategy_ingress", "none") == "blacklist":
        active.append("ingress")

    legacy = str(getattr(args, "ports_to_blacklist", "") or "").strip()
    per = {
        "intra-app": str(getattr(args, "ports_to_blacklist_intra_app", "") or "").strip(),
        "egress": str(getattr(args, "ports_to_blacklist_egress", "") or "").strip(),
        "ingress": str(getattr(args, "ports_to_blacklist_ingress", "") or "").strip(),
    }

    resolved: Dict[str, str] = {}
    for d in active:
        if per.get(d):
            resolved[d] = per[d]
            continue
        if legacy:
            if len(active) == 1:
                resolved[d] = legacy
                continue
            raise SystemExit(
                "Ambiguous --ports-to-blacklist: more than one direction uses blacklist. "
                "Use --ports-to-blacklist-intra-app / --ports-to-blacklist-egress / --ports-to-blacklist-ingress."
            )
        raise SystemExit(
            f"Missing blacklist sources for {d}. Provide --ports-to-blacklist-{d} "
            "or use legacy --ports-to-blacklist when only one direction is in blacklist mode."
        )
    return resolved


def _is_port_in_intervals(intervals_by_proto: Optional[Dict[str, List[Tuple[int, int]]]], proto: str, port: int) -> bool:
    """Return True if (proto,port) is inside the provided intervals."""
    if not intervals_by_proto:
        return False
    p = (proto or "").strip().lower()
    if not p or int(port or 0) <= 0:
        return False
    for a, b in intervals_by_proto.get(p, []):
        if a <= port <= b:
            return True
    return False


def _is_blacklisted(intervals_by_proto: Optional[Dict[str, List[Tuple[int, int]]]], proto: str, port: int) -> bool:
    """Return True if (proto,port) is inside the blacklist intervals.

    Accepts None/empty intervals (returns False) to keep call-sites simple and avoid crashes.
    """
    if not intervals_by_proto:
        return False
    proto = (proto or "").strip().lower()
    if proto not in ("tcp", "udp") or int(port or 0) <= 0:
        return False
    return _is_port_in_intervals(intervals_by_proto, proto, port)

def _pick_col(cols: List[str], *cands: str) -> str:
    low = {c.lower(): c for c in cols}
    for c in cands:
        if c.lower() in low:
            return low[c.lower()]
    return ""

def _gather_roles_from_wkld_m(wkld_m_path: Path, app: str, env: str) -> Set[str]:
    """Collect distinct role values from export_wkld.m.csv for one app/env."""
    if not wkld_m_path.exists():
        return set()
    rows, cols = _iter_csv_rows(wkld_m_path)
    if not rows or not cols:
        return set()
    c_app = _pick_col(cols, "app", "application")
    c_env = _pick_col(cols, "env", "environment")
    c_role = _pick_col(cols, "role")
    if not (c_app and c_env and c_role):
        return set()

    def eq(a: str, b: str) -> bool:
        return (a or "").strip().lower() == (b or "").strip().lower()

    out: Set[str] = set()
    for r in rows:
        if not eq(r.get(c_app, ""), app):
            continue
        if not eq(r.get(c_env, ""), env):
            continue
        v = (r.get(c_role, "") or "").strip()
        if v:
            out.add(v)
    return out

def _service_token(proto: str, port: Any) -> str:
    p = (proto or "").strip().lower()
    try:
        n = int(port) if str(port).strip() else 0
    except Exception:
        n = 0
    if p in ("icmp",):
        return "icmp"
    if n <= 0:
        return p or "any"
    return f"{p}/{n}"

def _service_sort_key(tok: Any) -> Tuple[int, str, int, int, str]:
    """Deterministic sort key for service tokens.

    Supports:
      - 'tcp/22', 'udp/53'
      - 'icmp'
      - fallback lexicographic for unknown tokens
    """
    t = str(tok or "").strip().lower()
    order = {"tcp": 0, "udp": 1, "icmp": 2, "igmp": 3, "any": 9}

    # tcp/22 or udp/53 (common internal token)
    if "/" in t:
        proto, rest = t.split("/", 1)
        proto = proto.strip()
        m = re.match(r"^(\d+)(?:-(\d+))?$", rest.strip())
        if m:
            a = int(m.group(1))
            b = int(m.group(2) or m.group(1))
            return (order.get(proto, 9), proto, a, b, t)
        # still prioritize by proto
        return (order.get(proto, 9), proto, 0, 0, t)

    # '80 tcp' / '80-90 tcp' / 'tcp 80' (defensive)
    m = re.match(r"^(tcp|udp)\s+(\d+)(?:-(\d+))?$", t)
    if m:
        proto = m.group(1)
        a = int(m.group(2))
        b = int(m.group(3) or m.group(2))
        return (order.get(proto, 9), proto, a, b, t)
    m = re.match(r"^(\d+)(?:-(\d+))?\s+(tcp|udp)$", t)
    if m:
        a = int(m.group(1))
        b = int(m.group(2) or m.group(1))
        proto = m.group(3)
        return (order.get(proto, 9), proto, a, b, t)

    # single proto token (icmp/any/...)
    return (order.get(t, 9), t, 0, 0, t)


def _sort_service_tokens(tokens: List[str]) -> List[str]:
    return sorted(tokens, key=_service_sort_key)


def _service_tokens_from_proposal(p: Dict[str, Any]) -> List[str]:
    """
    Extract a normalized service token list from one proposal row.

    A factorized proposal row represents a single (proto, port) tuple.
    We return a list for API symmetry with other callers.
    """
    proto = _as_text(p.get("proto") or "").strip()
    port = p.get("port", "")
    # Some rows might carry empty/None port for ICMP, etc.
    tok = _service_token(proto, port)
    tok = (tok or "").strip()
    return [tok] if tok else []


def build_intra_app_proposed_rules(
    proposals: List[Dict[str, Any]],
    raw_dir: Path,
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    blacklist_lists_used: Optional[List[str]] = None,
    finegrained_single_ports: Optional[Dict[str, List[Tuple[int, int]]]] = None,
) -> List[Dict[str, str]]:
    """
    Build rows for the 'Proposed rules' sheet (intra-app).

    Strategies:
      - allow: one broad rule across all roles in app/env, Services='All Services'
      - finegrained: strict ports observed in intra-app flows, grouped by (src_role,dst_role)
      - blacklist:
          * Create ONE default rule that allows all ports EXCEPT the blacklisted ports/ranges
          * Create exception rows for the blacklisted ports observed in intra-app flows (to validate)

    NOTE: For blacklist, blacklisted ports are expanded from carto.conf using --ports-to-blacklist list names.
    """
    strategy = (strategy or "none").strip().lower()
    if strategy in ("none", ""):
        return []

    # Resolve app/env for the run (anchor labels are stable in a run)
    run_app = ""
    run_env = ""
    for p in proposals:
        run_app = _as_text(p.get("anchor_app") or "").strip()
        run_env = _as_text(p.get("anchor_env") or "").strip()
        if run_app and run_env:
            break
    if not (run_app and run_env):
        return []

    # Best-effort: roles from managed workloads, fallback to roles seen in flows
    roles = _gather_roles_from_wkld_m(raw_dir / "export_wkld.m.csv", run_app, run_env)
    if not roles:
        for p in proposals:
            arole = _as_text(p.get("anchor_role") or "").strip()
            prole = _as_text(p.get("peer_role") or "").strip()
            if arole:
                roles.add(arole)
            if prole:
                roles.add(prole)

    ruleset_name = f"{run_app}-{run_env}-RS"
    rule_section = "Intrascope"


    intra_all = [p for p in proposals if _as_text(p.get("direction") or "").strip().lower() == "intra-app"]
    intra_sum_all = sum(_to_int(p.get("num_flows")) for p in intra_all)
    intra_sum_all_true = sum(_to_int(p.get("num_flows_true", p.get("num_flows"))) for p in intra_all)
    if strategy == "allow":
        if not roles:
            return []
        roles_str = "|".join(sorted(roles))
        return [{
            "Direction": "intra-app",
            "Strategy": strategy,
            "Source": roles_str,
            "Destination": roles_str,
            "Services": "All Services",
            "sum_num_flows": intra_sum_all,
                "Comment": "Blacklist default rule",
            "Rule Section": rule_section,
            "Ruleset": ruleset_name,
        }]

    # Only consider intra-app proposals for the rest
    intra = intra_all
    if not intra and strategy != "blacklist":
        return []

    # finegrained behavior (unchanged)
    if strategy == "finegrained":
        by_pair: Dict[Tuple[str, str], Set[str]] = {}
        by_pair_sum: Dict[Tuple[str, str], int] = {}
        by_pair_sum_true: Dict[Tuple[str, str], int] = {}
        by_pair_single: Dict[Tuple[str, str, str], int] = {}
        by_pair_single_true: Dict[Tuple[str, str, str], int] = {}
        for p in intra:
            fs = _flow_sides_from_proposal(p)
            if isinstance(fs, tuple) and len(fs) == 4:
                src_labels, _src_iplist, dst_labels, _dst_iplist = fs
            elif isinstance(fs, tuple) and len(fs) == 2:
                src_labels, dst_labels = fs
            else:
                continue

            src_role = (src_labels.get("role") or "").strip()
            dst_role = (dst_labels.get("role") or "").strip()
            if not (src_role and dst_role):
                continue

            toks = _service_tokens_from_proposal(p)
            if not toks:
                continue
            for tok in toks:
                m = re.match(r"(?i)^(tcp|udp)\/(\d+)$", tok.strip())
                if m and _is_port_in_intervals(finegrained_single_ports, m.group(1).lower(), int(m.group(2))):
                    k_single = (src_role, dst_role, tok.lower())
                    by_pair_single[k_single] = by_pair_single.get(k_single, 0) + _to_int(p.get("num_flows"))
                    by_pair_single_true[k_single] = by_pair_single_true.get(k_single, 0) + _to_int(p.get("num_flows_true", p.get("num_flows")))
                    continue
                by_pair.setdefault((src_role, dst_role), set()).add(tok)
                by_pair_sum[(src_role, dst_role)] = by_pair_sum.get((src_role, dst_role), 0) + _to_int(p.get("num_flows"))
                by_pair_sum_true[(src_role, dst_role)] = by_pair_sum_true.get((src_role, dst_role), 0) + _to_int(p.get("num_flows_true", p.get("num_flows")))

        out_rows: List[Dict[str, str]] = []
        for (src_role, dst_role) in sorted(by_pair.keys()):
            toks = _sort_service_tokens(by_pair[(src_role, dst_role)])
            out_rows.append({
                "Direction": "intra-app",
                "Strategy": strategy,
                "Source": src_role,
                "Destination": dst_role,
                "Services": ";".join(toks),
                "sum_num_flows": by_pair_sum.get((src_role, dst_role), 0),
                "sum_num_flows_true": by_pair_sum_true.get((src_role, dst_role), 0),
                "Rule Section": rule_section,
                "Ruleset": ruleset_name,
            })
        for (src_role, dst_role, tok) in sorted(by_pair_single.keys()):
            out_rows.append({
                "Direction": "intra-app",
                "Strategy": strategy,
                "Source": src_role,
                "Destination": dst_role,
                "Services": tok,
                "sum_num_flows": by_pair_single.get((src_role, dst_role, tok), 0),
                "sum_num_flows_true": by_pair_single_true.get((src_role, dst_role, tok), 0),
                "Rule Section": rule_section,
                "Ruleset": ruleset_name,
            })
        return out_rows

    # blacklist: default allow-all-except + exception rows for blacklisted ports observed
    if strategy == "blacklist":
        if not roles:
            return []
        if not blacklist_intervals:
            # Nothing to blacklist -> behave like allow
            roles_str = "|".join(sorted(roles))
            return [{
                "Direction": "intra-app",
                "Strategy": strategy,
                "Source": roles_str,
                "Destination": roles_str,
                "Services": "All Services",
                "sum_num_flows": intra_sum_all,
                "sum_num_flows_true": intra_sum_all_true,
            "Rule Section": rule_section,
                "Ruleset": ruleset_name,
            }]

        roles_str = "|".join(sorted(roles))

        # Default rule = complement of blacklist across TCP/UDP (plus ICMP/IGMP by default)
        allow_intervals = _complement_from_blacklist(blacklist_intervals)
        default_services = _render_intervals_as_services(allow_intervals, include_misc=True)

        out_rows: List[Dict[str, str]] = [{
            "Direction": "intra-app",
            "Strategy": strategy,
            "Source": roles_str,
            "Destination": roles_str,
            "Services": default_services,
            "sum_num_flows": 0,
            "sum_num_flows_true": 0,
            "Comment": "Blacklist default rule",
            "Rule Section": rule_section,
            "Ruleset": ruleset_name,
        }]

        # Exception rows = blacklisted ports observed in intra-app flows
        by_pair_bl: Dict[Tuple[str, str], Set[str]] = {}
        allow_sum_total = 0
        allow_sum_total_true = 0
        bl_sum_pair: Dict[Tuple[str, str], int] = {}
        bl_sum_pair_true: Dict[Tuple[str, str], int] = {}
        for p in intra:
            fs = _flow_sides_from_proposal(p)
            if isinstance(fs, tuple) and len(fs) == 4:
                src_labels, _src_iplist, dst_labels, _dst_iplist = fs
            elif isinstance(fs, tuple) and len(fs) == 2:
                src_labels, dst_labels = fs
            else:
                continue

            src_role = (src_labels.get("role") or "").strip()
            dst_role = (dst_labels.get("role") or "").strip()
            if not (src_role and dst_role):
                continue

            num = _to_int(p.get("num_flows"))
            num_true = _to_int(p.get("num_flows_true", p.get("num_flows")))
            p_proto = _norm_proto(str(p.get("proto", "") or ""))
            p_port = _to_int(p.get("port"))
            if p_proto in ("tcp", "udp") and p_port > 0 and _is_blacklisted(blacklist_intervals, p_proto, p_port):
                bl_sum_pair[(src_role, dst_role)] = bl_sum_pair.get((src_role, dst_role), 0) + num
                bl_sum_pair_true[(src_role, dst_role)] = bl_sum_pair_true.get((src_role, dst_role), 0) + num_true
            else:
                allow_sum_total += num
                allow_sum_total_true += num_true

            toks = _service_tokens_from_proposal(p)
            if not toks:
                continue

            for t in toks:
                m = re.match(r"(?i)^(tcp|udp)\/(\d+)$", t.strip())
                if not m:
                    continue
                proto = m.group(1).lower()
                port = int(m.group(2))
                if _is_blacklisted(blacklist_intervals, proto, port):
                    by_pair_bl.setdefault((src_role, dst_role), set()).add(f"{proto}/{port}")

        # Attach the computed sum of allowed flows to the default rule row
        if out_rows:
            out_rows[0]["sum_num_flows"] = allow_sum_total
            out_rows[0]["sum_num_flows_true"] = allow_sum_total_true

        for (src_role, dst_role) in sorted(by_pair_bl.keys()):
            toks = _sort_service_tokens(by_pair_bl[(src_role, dst_role)])
            out_rows.append({
                "Direction": "intra-app",
                "Strategy": strategy,
                "Source": src_role,
                "Destination": dst_role,
                "Services": ";".join(toks),
                "sum_num_flows": bl_sum_pair.get((src_role, dst_role), 0),
                "sum_num_flows_true": bl_sum_pair_true.get((src_role, dst_role), 0),
                "Rule Section": "Intrascope Exceptions",
                "Ruleset": ruleset_name,
            })

        return out_rows

    # Unknown strategy -> nothing
    return []
def append_excel_simple_table(excel_path: Path, sheet_name: str, header: List[str], rows: List[Dict[str, Any]], wrap_cols: Optional[Set[str]] = None, row_fill_fn: Optional[Callable[[Dict[str, Any]], Optional[str]]] = None, row_bold_fn: Optional[Callable[[Dict[str, Any]], bool]] = None) -> None:
    """Replace/create a simple sheet with bold header + borders."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    wb = load_workbook(filename=str(excel_path))
    # If previous runs created numbered variants (e.g., 'Proposed rules2'), clean them up
    # to avoid confusion/stale content. We only do it for our generated stable sheet.
    # IMPORTANT: we keep 'Proposed rules1' as it is now an intentional (additive) sheet.
    if sheet_name == PROPOSED_RULES_SHEET:
        for sn in list(wb.sheetnames):
            mm = re.fullmatch(re.escape(sheet_name) + r"(\d+)", sn)
            if mm and int(mm.group(1)) >= 2:
                wb.remove(wb[sn])
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    THIN = Side(style="thin", color="666666")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    FILL_H = PatternFill("solid", fgColor="D9E1F2")

    # header
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = FILL_H
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # rows
    wrap = set([x.strip() for x in (wrap_cols or set())])
    for r_i, r in enumerate(rows, start=2):
        fill_color = row_fill_fn(r) if row_fill_fn else None
        row_fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
        is_bold = bool(row_bold_fn(r)) if row_bold_fn else False
        for c, h in enumerate(header, start=1):
            v = r.get(h, "")
            cell = ws.cell(row=r_i, column=c, value=v)
            cell.border = BORDER
            if row_fill is not None:
                cell.fill = row_fill
            if is_bold:
                try:
                    cell.font = cell.font.copy(bold=True)  # type: ignore[attr-defined]
                except Exception:
                    cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=(h in wrap))

    # autosize (capped)
    for c, h in enumerate(header, start=1):
        letter = get_column_letter(c)
        max_len = len(str(h))
        for rr in range(2, ws.max_row + 1):
            vv = ws.cell(row=rr, column=c).value
            if vv is None:
                continue
            max_len = max(max_len, len(str(vv)))
        ws.column_dimensions[letter].width = min(60, max(12, int(max_len * 1.2)))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(str(excel_path))



# ------------------------------------------------------------------------------
# Post-formatting helpers (presentation only, no contract changes)
# ------------------------------------------------------------------------------

def _auto_width_col(ws, col_letter: str, min_row: int, max_row: int, min_w: int = 10, max_w: int = 70) -> None:
    """Auto-fit a column based on cell string lengths within [min_row..max_row], capped."""
    from openpyxl.utils import column_index_from_string

    col_idx = column_index_from_string(col_letter)
    max_len = 0
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None:
            continue
        s = str(v)
        if len(s) > max_len:
            max_len = len(s)
    if max_len <= 0:
        return
    ws.column_dimensions[col_letter].width = min(max_w, max(min_w, int(max_len * 1.1) + 2))


def _post_format_summary_sheet(wb) -> None:
    """Summary sheet presentation tweaks requested by the user."""
    from openpyxl.styles import Border

    if 'Summary' not in wb.sheetnames:
        return

    ws = wb['Summary']

    # 1) Merge B:C:D for rows 2..6 (B:D). Unmerge any previous ranges intersecting this area.
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = rng.bounds
        except Exception:
            continue
        # Intersects rows 2..6 and columns B..E (2..5)
        if max_row < 2 or min_row > 6:
            continue
        if max_col < 2 or min_col > 5:
            continue
        to_unmerge.append(str(rng))

    for r in to_unmerge:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass

    for row in range(2, 7):
        try:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        except Exception:
            pass
        # Clear borders in column E for these rows to avoid stray lines outside merged area
        try:
            ws.cell(row=row, column=5).border = Border()
        except Exception:
            pass

    # 2) Section 2 table: remove borders outside table (columns D/E) and auto-width B/C.
    sec2_title = 'Section 2: Label Groups inclusion (Scope-wide)'
    sec2_row = None
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == sec2_title:
            sec2_row = r
            break

    if sec2_row is None:
        return

    hdr_row = sec2_row + 1

    # Determine the table end row by scanning down until A/B/C are empty.
    end_row = hdr_row
    for r in range(hdr_row, (ws.max_row or hdr_row) + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if r > hdr_row:
            if all(x is None or str(x).strip() == '' for x in (a, b, c)):
                break
        end_row = r

    # Clear borders in D/E (and only D/E) within the section-2 table height.
    empty_border = Border()
    for r in range(hdr_row, end_row + 1):
        for col in (4, 5):
            try:
                ws.cell(row=r, column=col).border = empty_border
            except Exception:
                pass

    # Auto-width columns B & C on section-2 table block (including header)
    _auto_width_col(ws, 'B', hdr_row, end_row, min_w=12, max_w=55)
    _auto_width_col(ws, 'C', hdr_row, end_row, min_w=18, max_w=85)


def _apply_light_border_to_sheet(wb, sheet_name: str) -> None:
    """Apply a light thin border to the used range of a sheet (presentation only)."""
    from openpyxl.styles import Border, Side

    if sheet_name not in wb.sheetnames:
        return

    ws = wb[sheet_name]
    if (ws.max_row or 0) <= 0 or (ws.max_column or 0) <= 0:
        return

    side = Side(style='thin', color='D9D9D9')
    border = Border(left=side, right=side, top=side, bottom=side)

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Apply border only within used range. Avoid touching entirely empty trailing rows.
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            # Keep existing fill/font/alignment; only set border.
            cell.border = border


def post_format_excel_presentation(excel_path: Path) -> None:
    """Apply small, deterministic presentation fixes without touching data contracts."""
    from openpyxl import load_workbook

    if not excel_path.exists():
        return

    wb = load_workbook(filename=str(excel_path))
    try:
        _post_format_summary_sheet(wb)
        _apply_light_border_to_sheet(wb, 'Scope Applicable Rules')
        _apply_light_border_to_sheet(wb, 'Ruleset Effectiveness')
        wb.save(str(excel_path))
    finally:
        try:
            wb.close()
        except Exception:
            pass


def add_flow_rule_match_legend(excel_path: Path, sheet_name: str = "Flow-Rule Match") -> None:
    """Add/refresh a legend block to the right of the Flow-Rule Match table.

    It tries to reuse the same fill colors already present in the sheet (best effort),
    and falls back to standard Excel colors if it cannot infer them.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not excel_path.exists():
        return

    wb = load_workbook(filename=str(excel_path))
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]

    # Detect if we already have a legend; keep it stable (avoid pushing further right on every run).
    legend_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower() == "legend":
            legend_col = c
            break

    # Determine table end column (header row = 1) excluding legend if present
    if legend_col:
        start_col = legend_col
        table_end_col = max(1, legend_col - 2)
    else:
        table_end_col = 1
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v not in (None, ""):
                table_end_col = c
        start_col = table_end_col + 2

    def _solid_rgb(cell) -> Optional[str]:
        try:
            if cell.fill and cell.fill.patternType == "solid":
                rgb = getattr(cell.fill.fgColor, "rgb", None)
                return rgb
        except Exception:
            return None
        return None

    def _guess_color(keywords: List[str], fallback_rgb: str) -> str:
        kws = [k.lower() for k in keywords]
        max_row = min(ws.max_row or 1, 2000)
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=table_end_col):
            for cell in row:
                if cell.value is None:
                    continue
                s = str(cell.value).lower()
                if any(k in s for k in kws):
                    rgb = _solid_rgb(cell)
                    if rgb:
                        return rgb
        return fallback_rgb

    # Best-effort color discovery by keywords; fall back to our standard palette.
    C_GREY = _guess_color(["deleted"], "FFD9D9D9")
    C_YELLOW = _guess_color(["no rule", "no match"], "FFFFFF00")
    C_BLUE = _guess_color(["bouquets"], "FF9DC3E6")
    C_GREEN = _guess_color(["in scope"], "FFC6E0B4")
    C_LIGHT_ORANGE = _guess_color(["other scope", "extrascope", "extra scope"], "FFF8CBAD")
    C_ORANGE_RED = _guess_color(["duplicate"], "FFF4B084")

    items = [
        (C_GREY, "Deleted workload, ignored"),
        (C_YELLOW, "No rule match the flow"),
        (C_BLUE, "Flow authorized with Bouquets Infra rule"),
        (C_GREEN, "Flow authorized with Business rule in scope"),
        (C_LIGHT_ORANGE, "Flow authorized with Business rule in other scope"),
        (C_ORANGE_RED, "Duplicate rules found"),
    ]

    THIN = Side(style="thin", color="666666")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Clear previous legend block (2 columns wide, 1 header + N rows)
    for r in range(1, 2 + len(items) + 1):
        for c in range(start_col, start_col + 2):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill()  # reset
            cell.border = Border()
            cell.font = Font()
            cell.alignment = Alignment()

    # Header
    hdr = ws.cell(row=1, column=start_col, value="Legend")
    hdr.font = Font(bold=True)
    hdr.alignment = Alignment(horizontal="center", vertical="center")

    # Rows
    for i, (rgb, label) in enumerate(items, start=2):
        ccell = ws.cell(row=i, column=start_col, value=None)
        ccell.fill = PatternFill("solid", fgColor=rgb)
        ccell.border = BORDER

        tcell = ws.cell(row=i, column=start_col + 1, value=label)
        tcell.border = BORDER
        tcell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    ws.column_dimensions[get_column_letter(start_col)].width = 6
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 45

    wb.save(str(excel_path))
    wb.close()



# ------------------------------ Ingress IPList prefix grouping ------------------------------
# Never group IPLISTs together if they don't share a significant prefix.
# Longest-prefix-first to avoid CS_ catching CSD_.
IPLIST_PREFIXES_RAW = [
    "CSD_", "DNS_", "SSD_", "U_", "NZ4_", "KUB_", "NZ3_", "LBI_", "LBO_",
    "CS_", "SS_", "NZ2_", "NZ1_", "NZ0"
]
IPLIST_PREFIXES = sorted(IPLIST_PREFIXES_RAW, key=len, reverse=True)

def _iplist_prefix(iplist_name: str) -> str:
    """Return the grouping prefix for an IPList name.

    - If name starts with a known significant prefix -> return that prefix.
    - Else if it contains '_' -> return token-before-first-underscore + '_'
    - Else -> return the full name (prevents accidental grouping).
    """
    n = str(iplist_name or "").strip()
    if not n:
        return ""
    for pref in IPLIST_PREFIXES:
        if n.startswith(pref):
            return pref
    if "_" in n:
        return n.split("_", 1)[0] + "_"
    return n



def _app_prefix(app_value: str) -> str:
    """
    Compute prefix for an 'app' label value: substring from start until first '_' (excluded).
    If no '_' exists, prefix = whole app.
    """
    s = _as_text(app_value or "").strip()
    if not s:
        return ""
    if "_" in s:
        return s.split("_", 1)[0]
    return s

def _source_selector_for_peer_labels(peer_value: str) -> str:
    """
    Build the Source selector for ingress/labels proposals (Extrascope rules).

    Rules:
      - Always include app + env
      - If app prefix ends with 'M' (managed family), also include role
      - Never include loc / OS

    Output format: 'app=...|env=...|role=...' (role optional)
    """
    try:
        kv = parse_kv_tokens(_as_text(peer_value or ""))
    except Exception:
        kv = {}
    app = _as_text(kv.get("app") or "").strip()
    env = _as_text(kv.get("env") or "").strip()
    role = _as_text(kv.get("role") or "").strip()

    pref = _app_prefix(app)
    managed = bool(pref) and pref.endswith("M")

    parts: List[str] = []
    if app:
        parts.append(f"app={app}")
    if env:
        parts.append(f"env={env}")
    if managed and role:
        parts.append(f"role={role}")
    return "|".join(parts).strip()

def build_ingress_labels_extrascope_proposed_rules(
    proposals: List[Dict[str, Any]],
    strategy: str
) -> List[Dict[str, str]]:
    """
    Proposed rules for ingress flows where peer_type='labels' and the matched_rule_category is NOT Bouquets.

    Selection:
      - direction == 'ingress'
      - peer_type == 'labels'
      - matched_rule_category != 'Bouquets Infra rule'

    Aggregation:
      - Group by (anchor_app, anchor_env, anchor_role, proto, port)
      - Source: unique derived selectors from peer_value (one per line)
      - sum_num_flows: sum of num_flows across grouped Flow-Rule Match rows

    Rule Section: 'Extrascope'
    """
    out_rows: List[Dict[str, str]] = []

    # (app, env, role, proto, port) -> {"sources":set[str], "sum":int}
    groups: Dict[Tuple[str, str, str, str, int], Dict[str, Any]] = {}

    for p in proposals:
        if str(p.get("direction", "") or "").strip() != "ingress":
            continue
        if str(p.get("peer_type", "") or "").strip() != "labels":
            continue
        if str(p.get("matched_rule_category", "") or "").strip() == CAT_BOUQUET:
            continue

        anchor_app = _as_text(p.get("anchor_app") or "").strip()
        anchor_env = _as_text(p.get("anchor_env") or "").strip()
        anchor_role = _as_text(p.get("anchor_role") or "").strip()
        if not anchor_app or not anchor_env or not anchor_role:
            continue

        proto = _as_text(p.get("proto") or "").strip().lower()
        port = p.get("port", None)
        try:
            port_i = int(port)
        except Exception:
            # proto-only traffic (ICMP/IGMP) not handled here
            continue
        if port_i <= 0:
            continue

        pv = p.get("peer_value", "")
        pv_list: List[str] = []
        if isinstance(pv, (set, list, tuple)):
            pv_list = [str(x) for x in pv if str(x).strip()]
        else:
            pv_list = [str(pv or "")]

        touched: Set[Tuple[str, str, str, str, int]] = set()
        gk = (anchor_app, anchor_env, anchor_role, proto, port_i)
        for pv_item in sorted(set([x.strip() for x in pv_list if x.strip()])):
            src_sel = _source_selector_for_peer_labels(pv_item)
            if not src_sel:
                continue
            g = groups.get(gk)
            if not g:
                g = {"sources": set(), "sum": 0}
                groups[gk] = g
            g["sources"].add(src_sel)
            touched.add(gk)
        if touched:
            try:
                groups[gk]["sum"] += int(p.get("num_flows") or 0)
            except Exception:
                pass

    for (app, env, role, proto, port_i), g in sorted(groups.items(), key=lambda kv: kv[0]):
        ruleset_name = f"{app}-{env}-RS"
        tok = _service_token(proto, port_i)
        out_rows.append({
            "Direction": "ingress",
            "Strategy": strategy,
            "Source": "\n".join(sorted(g["sources"])),
            "Destination": role,
            "Services": tok,
            "sum_num_flows": int(g.get("sum") or 0),
            "Rule Section": "Extrascope",
            "Ruleset": ruleset_name,
        })

    return out_rows


def _merge_comment(existing: str, addition: str) -> str:
    existing = (existing or "").strip()
    addition = (addition or "").strip()
    if not addition:
        return existing
    if not existing:
        return addition
    # Deterministic ordering: keep existing first, add if not already present.
    parts = [p.strip() for p in existing.split(";") if p.strip()]
    if addition not in parts:
        parts.append(addition)
    return "; ".join(parts)

def _load_bouquets_refs(enabled_rules_csv: Path) -> Tuple[Set[str], Set[str]]:
    """Return (apps, iplists) referenced by rulesets whose name starts with 'BOUQUETS_'.

    - apps are extracted from src_labels/dst_labels tokens that start with 'app:' (semicolon-separated label list)
    - iplists are extracted from src_iplists/dst_iplists tokens (semicolon-separated list), excluding 'Any'
    """
    apps: Set[str] = set()
    iplists: Set[str] = set()

    if not enabled_rules_csv.exists():
        return apps, iplists

    ANY_IPLISTS = {
        "Any (0.0.0.0/0 and ::/0)",
        "Any (0.0.0.0/0)",
        "Any",
    }

    with enabled_rules_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ruleset_name = (row.get("ruleset_name") or "").strip()
            if not ruleset_name.startswith("BOUQUETS_"):
                continue

            for col in ("src_labels", "dst_labels"):
                s = (row.get(col) or "").strip()
                if not s:
                    continue
                # format like: env:PRD;app:APA_SPV_A7XXX;app:CSU_MON...
                for tok in s.split(";"):
                    tok = tok.strip()
                    if tok.startswith("app:"):
                        val = tok.split(":", 1)[1].strip()
                        if val:
                            apps.add(val)

            for col in ("src_iplists", "dst_iplists"):
                s = (row.get(col) or "").strip()
                if not s:
                    continue
                for tok in s.split(";"):
                    tok = tok.strip()
                    if not tok or tok in ANY_IPLISTS:
                        continue
                    iplists.add(tok)

    return apps, iplists

_APP_EQ_RE = re.compile(r"(?:^|\|)\s*app=([^|\n]+)")
_APP_COLON_RE = re.compile(r"(?:^|;)\s*app:([^;\n]+)")

def _extract_apps_from_selector(selector: str) -> Set[str]:
    s = (selector or "")
    out: Set[str] = set()
    for mm in _APP_EQ_RE.finditer(s):
        v = (mm.group(1) or "").strip()
        if v:
            out.add(v)
    for mm in _APP_COLON_RE.finditer(s):
        v = (mm.group(1) or "").strip()
        if v:
            out.add(v)
    return out

def _extract_iplists_from_selector(selector: str) -> Set[str]:
    s = (selector or "").strip()
    if not s:
        return set()
    # Heuristic: if selector contains '=' or common label syntax, it's not an IPList cell.
    if "=" in s or "env=" in s or "role=" in s or "app=" in s or "app:" in s:
        return set()
    parts = re.split(r"[\n\|;]+", s)
    return set(p.strip() for p in parts if p.strip())

def _row_matches_bouquets_refs(row: Dict[str, Any], bouquets_apps: Set[str], bouquets_iplists: Set[str]) -> bool:
    src = str(row.get("Source", "") or "")
    dst = str(row.get("Destination", "") or "")
    apps = _extract_apps_from_selector(src) | _extract_apps_from_selector(dst)
    if apps and any(a in bouquets_apps for a in apps):
        return True
    ipls = _extract_iplists_from_selector(src) | _extract_iplists_from_selector(dst)
    if ipls and any(i in bouquets_iplists for i in ipls):
        return True
    return False

def build_ingress_labels_extrascope_proposed_rules_v1(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
) -> List[Dict[str, Any]]:
    """Build V1 Proposed rules rows for ingress where peer_type=labels (Extrascope only).

    - finegrained: group by (anchor_app, anchor_env, anchor_role, source_selector) and merge services with ';'
    - blacklist: group non-blacklisted services as above, but keep one line per blacklisted port with Comment='Blacklist Exception'
    """
    out_rows: List[Dict[str, Any]] = []

    allow_groups: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
    bl_groups: Dict[Tuple[str, str, str, str, str, int], int] = {}
    bl_groups_true: Dict[Tuple[str, str, str, str, str, int], int] = {}

    for p in proposals:
        if str(p.get("direction", "") or "").strip() != "ingress":
            continue
        if str(p.get("peer_type", "") or "").strip() != "labels":
            continue
        if str(p.get("matched_rule_category", "") or "").strip() == CAT_BOUQUET:
            continue

        anchor_app = _as_text(p.get("anchor_app") or "").strip()
        anchor_env = _as_text(p.get("anchor_env") or "").strip()
        anchor_role = _as_text(p.get("anchor_role") or "").strip()
        if not anchor_app or not anchor_env or not anchor_role:
            continue

        proto = _as_text(p.get("proto") or "").strip().lower()
        port = p.get("port", None)
        try:
            port_i = int(port)
        except Exception:
            continue
        if port_i <= 0:
            continue

        pv = p.get("peer_value", "")
        if isinstance(pv, (set, tuple, list)):
            pv_list = [str(x) for x in pv]
        else:
            pv_list = [str(pv)]

        num_flows = _to_int(p.get("num_flows"))
        num_flows_true = _to_int(p.get("num_flows_true", p.get("num_flows")))

        for pv_one in pv_list:
            src_selector = _source_selector_for_peer_labels(pv_one)
            if not src_selector:
                continue

            if str(strategy).strip().lower() == "blacklist" and _is_blacklisted(blacklist_intervals, proto, port_i):
                k_bl = (anchor_app, anchor_env, anchor_role, src_selector, proto, port_i)
                bl_groups[k_bl] = bl_groups.get(k_bl, 0) + num_flows
                bl_groups_true[k_bl] = bl_groups_true.get(k_bl, 0) + num_flows_true
                continue

            k_allow = (anchor_app, anchor_env, anchor_role, src_selector)
            g = allow_groups.setdefault(k_allow, {"services": set(), "sum": 0, "sum_true": 0})
            g["services"].add(_service_token(proto, port_i))
            g["sum"] += num_flows
            g["sum_true"] += num_flows_true

    for (anchor_app, anchor_env, anchor_role, src_selector), g in allow_groups.items():
        ruleset_name = f"{anchor_app}-{anchor_env}-RS"
        services = ";".join(sorted(g["services"], key=_service_sort_key))
        out_rows.append({
            "Direction": "ingress",
            "Strategy": strategy,
            "Source": src_selector,
            "Destination": anchor_role,
            "Services": services,
            "sum_num_flows": g["sum"],
            "sum_num_flows_true": g.get("sum_true", g["sum"]),
            "Rule Section": "Extrascope",
            "Comment": "",
            "Ruleset": ruleset_name,
        })

    for (anchor_app, anchor_env, anchor_role, src_selector, proto, port_i), ssum in bl_groups.items():
        ruleset_name = f"{anchor_app}-{anchor_env}-RS"
        ssum_true = bl_groups_true.get((anchor_app, anchor_env, anchor_role, src_selector, proto, port_i), ssum)
        out_rows.append({
            "Direction": "ingress",
            "Strategy": strategy,
            "Source": src_selector,
            "Destination": anchor_role,
            "Services": _service_token(proto, port_i),
            "sum_num_flows": ssum,
            "sum_num_flows_true": ssum_true,
            "Rule Section": "Extrascope",
            "Comment": "Blacklist Exception" if str(strategy).strip().lower() == "blacklist" else "",
            "Ruleset": ruleset_name,
        })

    return sorted(out_rows, key=lambda r: (
        str(r.get("Ruleset", "")),
        str(r.get("Rule Section", "")),
        str(r.get("Destination", "")),
        str(r.get("Source", "")),
        0 if (r.get("Comment", "") or "") == "" else 1,
        str(r.get("Services", "")),
    ))

def build_ingress_proposed_rules_v1(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    blacklist_lists_used: Optional[List[str]] = None
) -> List[Dict[str, Any]]:
    """Build V1 Proposed rules rows for ingress (adds Comment + improved Extrascope)."""
    base = build_ingress_proposed_rules(proposals, strategy, blacklist_intervals, blacklist_lists_used)
    out: List[Dict[str, Any]] = []

    for r in base:
        if str(r.get("Rule Section","")) == "Extrascope":
            continue
        rr = dict(r)
        rr["Comment"] = ""
        if str(strategy).strip().lower() == "blacklist":
            if str(rr.get("Source","")).strip() != "Any (0.0.0.0/0)":
                rr["Comment"] = "Blacklist Exception"
            else:
                rr["Comment"] = "Blacklist default rule"
        out.append(rr)

    out.extend(build_ingress_labels_extrascope_proposed_rules_v1(proposals, strategy, blacklist_intervals))
    return out
def build_ingress_proposed_rules(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    blacklist_lists_used: Optional[List[str]] = None
) -> List[Dict[str, str]]:
    """
    Build rows for the 'Proposed rules' sheet (ingress).

    Base selection from Flow-Rule Match:
      - direction == 'ingress'
      - Info != 'Bouquets Infra'
      - peer_type == 'iplist'

    IPList grouping rule (per your spec):
      - Never group IPLISTs together if they don't share the same significant prefix.

    Finegrained aggregation rule:
      - Group by (anchor_app, anchor_env, anchor_role, iplist_prefix, proto, port)
      - Aggregate all peer iplist names (Source) that target the same anchor_role on the same proto/port.

    Blacklist rule:
      - For each (anchor_role, iplist_prefix), create a DEFAULT rule that allows all ports except blacklisted.
      - Create EXCEPTION rules ONLY for blacklisted ports observed in ingress flows.
    """
    # key: (app, env, role, prefix, proto, port) -> set(iplist names)
    by_key: Dict[Tuple[str, str, str, str, str, int], Set[str]] = {}
    sum_by_key: Dict[Tuple[str, str, str, str, str, int], int] = {}
    sum_by_key_true: Dict[Tuple[str, str, str, str, str, int], int] = {}
    # sources per (app, env, role, prefix) for DEFAULT rules in blacklist
    role_sources: Dict[Tuple[str, str, str, str], Set[str]] = {}
    # roles per (app, env)
    roles_by_scope: Dict[Tuple[str, str], Set[str]] = {}

    # IMPORTANT: DEFAULT blacklist rules must list *all* roles in the scope,
    # not only those appearing in ingress+iplist flows.
    for p_any in proposals:
        if str(p_any.get("Info", "") or "").strip() == "Bouquets Infra":
            continue
        app_any = _as_text(p_any.get("anchor_app", "") or "").strip()
        env_any = _as_text(p_any.get("anchor_env", "") or "").strip()
        role_any = _as_text(p_any.get("anchor_role", "") or "").strip()
        if app_any and env_any and role_any:
            roles_by_scope.setdefault((app_any, env_any), set()).add(role_any)

    for p in proposals:
        if str(p.get("direction", "")).strip().lower() != "ingress":
            continue
        if str(p.get("Info", "")).strip() == "Bouquets Infra":
            continue
        if str(p.get("peer_type", "")).strip().lower() != "iplist":
            continue

        app = str(p.get("anchor_app", "") or "").strip()
        env = str(p.get("anchor_env", "") or "").strip()
        role = str(p.get("anchor_role", "") or "").strip()
        proto = str(p.get("proto", "") or "").strip().lower()
        try:
            port = int(str(p.get("port", "") or "0").strip() or "0")
        except Exception:
            port = 0
        peer_val = _as_text(p.get("peer_value", ""), joiner="|").strip()

        if not app or not env or not role or not proto or port <= 0 or not peer_val:
            continue

        # Peer value is expected to be an iplist name; allow "|" (or newlines) as a separator if already aggregated upstream.
        peer_val = peer_val.replace("\n", "|")
        peers = [x.strip() for x in peer_val.split("|") if x.strip()]
        if not peers:
            continue

        roles_by_scope.setdefault((app, env), set()).add(role)

        # Split peers by prefix to enforce the "never group different prefixes" contract.
        by_pref: Dict[str, List[str]] = {}
        for ipn in peers:
            pref = _iplist_prefix(ipn)
            if not pref:
                continue
            by_pref.setdefault(pref, []).append(ipn)

        for pref, ipnames in by_pref.items():
            k = (app, env, role, pref, proto, port)
            by_key.setdefault(k, set()).update(ipnames)
            sum_by_key[k] = sum_by_key.get(k, 0) + _to_int(p.get("num_flows"))
            sum_by_key_true[k] = sum_by_key_true.get(k, 0) + _to_int(p.get("num_flows_true", p.get("num_flows")))
            role_sources.setdefault((app, env, role, pref), set()).update(ipnames)

    if not by_key:
        return []

    out_rows: List[Dict[str, str]] = []
    extrascope_rows = build_ingress_labels_extrascope_proposed_rules(proposals, strategy)


    # Build deterministic scope/prefix listings
    scopes = sorted(set((app, env) for (app, env, *_rest) in by_key.keys()))
    prefixes_by_scope: Dict[Tuple[str, str], Set[str]] = {}
    for (app, env, role, pref, proto, port) in by_key.keys():
        prefixes_by_scope.setdefault((app, env), set()).add(pref)

    if strategy == "allow":
        # One broad rule per (scope, prefix) to respect the prefix constraint
        for (app, env) in scopes:
            ruleset_name = f"{app}-{env}-RS"
            roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
            for pref in sorted(prefixes_by_scope.get((app, env), set())):
                # All sources for that prefix across all roles in the scope
                src_set: Set[str] = set()
                for (a, e, r, pfx), ips in role_sources.items():
                    if (a, e) == (app, env) and pfx == pref:
                        src_set.update(ips)
                if not src_set or not roles_str:
                    continue
                # Sum num_flows for this (scope, prefix) across all roles and ports
                sum_flows = 0
                for (a2, e2, role2, pfx2, proto2, port2), _peers2 in by_key.items():
                    if (a2, e2) == (app, env) and pfx2 == pref:
                        sum_flows += sum_by_key.get((a2, e2, role2, pfx2, proto2, port2), 0)

                out_rows.append({
                    "Direction": "ingress",
                    "Strategy": strategy,
                    "Source": "|".join(sorted(src_set)),
                    "Destination": roles_str,
                    "Services": "All Services",
                    "sum_num_flows": sum_flows,
                    "Rule Section": "intra-scope",
                    "Ruleset": ruleset_name,
                })
        out_rows.extend(extrascope_rows)
        return out_rows

    if strategy == "finegrained":
        # deterministic ordering by app, env, prefix, role, proto, port
        keys_sorted = sorted(by_key.items(), key=lambda kv: (
            kv[0][0], kv[0][1], kv[0][3], kv[0][2], kv[0][4], kv[0][5]
        ))
        for (app, env, role, pref, proto, port), peers in keys_sorted:
            ruleset_name = f"{app}-{env}-RS"
            src_str = "|".join(sorted(peers))
            tok = _service_token(proto, port)
            out_rows.append({
                "Direction": "ingress",
                "Strategy": strategy,
                "Source": src_str,
                "Destination": role,
                "Services": tok,
                "sum_num_flows": sum_by_key.get((app, env, role, pref, proto, port), 0),
                "sum_num_flows_true": sum_by_key_true.get((app, env, role, pref, proto, port), 0),
                "sum_num_flows_true": sum_by_key_true.get((app, env, role, pref, proto, port), 0),
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })
        out_rows.extend(extrascope_rows)
        return out_rows

    if strategy == "blacklist":
        # If no blacklist intervals were provided, behave like allow (still marks strategy as blacklist)
        if not blacklist_intervals:
            # Defensive fallback (should not happen if CLI/config is correct):
            # still generate ONE default rule per scope.
            any_source = "Any (0.0.0.0/0)"
            for (app, env) in scopes:
                ruleset_name = f"{app}-{env}-RS"
                roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
                if not roles_str:
                    continue
                # Sum all ingress iplist flows for this scope (fallback when no blacklist intervals)
                sum_all = 0
                for (a2, e2, role2, pfx2, proto2, port2) in by_key.keys():
                    if (a2, e2) == (app, env):
                        sum_all += sum_by_key.get((a2, e2, role2, pfx2, proto2, port2), 0)

                out_rows.append({
                    "Direction": "ingress",
                    "Strategy": strategy,
                    "Source": any_source,
                    "Destination": roles_str,
                    "Services": "All Services",
                    "sum_num_flows": sum_all,
                "Comment": "Blacklist default rule",
                "Rule Section": "intra-scope",
                    "Ruleset": ruleset_name,
                })
            out_rows.extend(extrascope_rows)
            return out_rows

        allow_intervals = _complement_from_blacklist(blacklist_intervals)
        default_services = _render_intervals_as_services(allow_intervals, include_misc=True)

        # DEFAULT rule: ONE per scope (per your spec)
        # Source = Any (0.0.0.0/0)
        # Destination = all roles in the scope (pipe-separated)
        any_source = "Any (0.0.0.0/0)"
        for (app, env) in scopes:
            ruleset_name = f"{app}-{env}-RS"
            roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
            if not roles_str:
                continue
            # Sum num_flows for this scope that are allowed by the default rule (NOT blacklisted)
            allowed_sum = 0
            allowed_sum_true = 0
            allowed_sum_true = 0
            for (a2, e2, role2, pfx2, proto2, port2) in by_key.keys():
                if (a2, e2) == (app, env) and not _is_blacklisted(blacklist_intervals, proto2, port2):
                    allowed_sum += sum_by_key.get((a2, e2, role2, pfx2, proto2, port2), 0)

            out_rows.append({
                "Direction": "ingress",
                "Strategy": strategy,
                "Source": any_source,
                "Destination": roles_str,
                "Services": default_services,
                "sum_num_flows": allowed_sum,
                "sum_num_flows_true": allowed_sum_true,
                "sum_num_flows_true": allowed_sum_true,
                "Comment": "Blacklist default rule",
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })

        # EXCEPTIONS: only blacklisted ports actually observed, still grouped by (role, prefix, proto, port)
        exc_items = []
        for (app, env, role, pref, proto, port), peers in by_key.items():
            if _is_blacklisted(blacklist_intervals, proto, port):
                exc_items.append(((app, env, role, pref, proto, port), peers))
        exc_items = sorted(exc_items, key=lambda kv: (kv[0][0], kv[0][1], kv[0][3], kv[0][2], kv[0][4], kv[0][5]))

        for (app, env, role, pref, proto, port), peers in exc_items:
            ruleset_name = f"{app}-{env}-RS"
            out_rows.append({
                "Direction": "ingress",
                "Strategy": strategy,
                "Source": "|".join(sorted(peers)),
                "Destination": role,
                "Services": _service_token(proto, port),
                "sum_num_flows": sum_by_key.get((app, env, role, pref, proto, port), 0),
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })

        out_rows.extend(extrascope_rows)
        return out_rows

    # Unknown strategy -> nothing
    return []




def build_egress_proposed_rules(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    blacklist_lists_used: Optional[List[str]] = None
) -> List[Dict[str, str]]:
    """
    Build rows for the 'Proposed rules' sheet (egress).

    Base selection from Flow-Rule Match:
      - direction == 'egress'
      - Info != 'Bouquets Infra'
      - peer_type == 'iplist'

    Finegrained rule:
      - Group by (anchor_app, anchor_env, anchor_role, iplist_prefix, proto, port)
      - Destination: all peer iplist names (pipe-separated) for the group
      - Source: anchor_role

    Blacklist rule (analogous to ingress):
      - DEFAULT rule: ONE per scope (app, env)
          Source = all roles in the scope (pipe-separated)
          Destination = Any (0.0.0.0/0)
          Services = complement of blacklisted ports (rendered as intervals)
      - EXCEPTION rules: only blacklisted ports observed, still grouped by (anchor_role, iplist_prefix, proto, port)
    """
    strategy = (strategy or "").strip().lower()
    if strategy not in ("allow", "finegrained", "blacklist"):
        return []

    by_key: Dict[Tuple[str, str, str, str, str, int], Set[str]] = {}
    sum_by_key: Dict[Tuple[str, str, str, str, str, int], int] = {}
    roles_by_scope: Dict[Tuple[str, str], Set[str]] = {}
    # Keep a separate accumulator for the *true* number of flows (sum of Flow-in/out [Num Flows])
    # when available. This prevents regressions when the sheet uses num_flows_true.
    sum_true_by_key: Dict[Tuple[str, str, str, str, str, int], int] = {}

    # IMPORTANT: DEFAULT blacklist rules must list *all* roles in the scope,
    # not only those appearing in egress+iplist flows.
    for p_any in proposals:
        if str(p_any.get("Info", "") or "").strip() == "Bouquets Infra":
            continue
        app_any = _as_text(p_any.get("anchor_app", "") or "").strip()
        env_any = _as_text(p_any.get("anchor_env", "") or "").strip()
        role_any = _as_text(p_any.get("anchor_role", "") or "").strip()
        if app_any and env_any and role_any:
            roles_by_scope.setdefault((app_any, env_any), set()).add(role_any)

    for p in proposals:
        if str(p.get("direction", "")).strip().lower() != "egress":
            continue
        if str(p.get("Info", "")).strip() == "Bouquets Infra":
            continue
        if str(p.get("peer_type", "")).strip() != "iplist":
            continue

        app = _as_text(p.get("anchor_app", "") or "").strip()
        env = _as_text(p.get("anchor_env", "") or "").strip()
        role = _as_text(p.get("anchor_role", "") or "").strip()

        proto = _as_text(p.get("proto", "") or "").strip().lower()
        try:
            port = int(str(p.get("port", "") or "0").strip() or "0")
        except Exception:
            port = 0

        peer_val = _as_text(p.get("peer_value", ""), joiner="|").strip()
        if not app or not env or not role or not proto or port <= 0 or not peer_val:
            continue

        # Peer value is expected to be an iplist name; allow "|" (or newlines) as a separator if already aggregated upstream.
        peer_val = peer_val.replace("\n", "|")
        peers = [x.strip() for x in peer_val.split("|") if x.strip()]
        if not peers:
            continue

        roles_by_scope.setdefault((app, env), set()).add(role)

        # Split peers by prefix to enforce the "never group different prefixes" contract.
        by_pref: Dict[str, List[str]] = {}
        for ipn in peers:
            pref = _iplist_prefix(ipn)
            if not pref:
                continue
            by_pref.setdefault(pref, []).append(ipn)

        for pref, ipnames in by_pref.items():
            k = (app, env, role, pref, proto, port)
            by_key.setdefault(k, set()).update(set(ipnames))
            try:
                nf = int(str(p.get("num_flows", "0") or "0").strip() or "0")
            except Exception:
                nf = 0
            sum_by_key[k] = sum_by_key.get(k, 0) + nf
            sum_true_by_key[k] = sum_true_by_key.get(k, 0) + _to_int_local(p.get("num_flows_true", p.get("num_flows")))

    if not by_key:
        return []

    out_rows: List[Dict[str, str]] = []
    scopes = sorted(set((app, env) for (app, env, *_rest) in by_key.keys()))

    if strategy == "allow":
        # One broad rule per scope
        for (app, env) in scopes:
            ruleset_name = f"{app}-{env}-RS"
            roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
            if not roles_str:
                continue
            # Sum all selected egress iplist flows for this scope
            sum_all = 0
            for (a2, e2, role2, pref2, proto2, port2) in by_key.keys():
                if (a2, e2) == (app, env):
                    sum_all += sum_by_key.get((a2, e2, role2, pref2, proto2, port2), 0)
            out_rows.append({
                "Direction": "egress",
                "Strategy": strategy,
                "Source": roles_str,
                "Destination": "Any (0.0.0.0/0)",
                "Services": "All Services",
                "sum_num_flows": sum_all,
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })
        return out_rows

    if strategy == "finegrained":
        items = sorted(by_key.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][3], kv[0][2], kv[0][4], kv[0][5]))
        for (app, env, role, pref, proto, port), peers in items:
            ruleset_name = f"{app}-{env}-RS"
            out_rows.append({
                "Direction": "egress",
                "Strategy": strategy,
                "Source": role,
                "Destination": "|".join(sorted(peers)),
                "Services": _service_token(proto, port),
                "sum_num_flows": sum_by_key.get((app, env, role, pref, proto, port), 0),
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })
        return out_rows

    # blacklist
    if not blacklist_intervals:
        # Defensive fallback: still generate ONE default rule per scope.
        for (app, env) in scopes:
            ruleset_name = f"{app}-{env}-RS"
            roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
            if not roles_str:
                continue
            sum_all = 0
            for (a2, e2, role2, pref2, proto2, port2) in by_key.keys():
                if (a2, e2) == (app, env):
                    sum_all += sum_by_key.get((a2, e2, role2, pref2, proto2, port2), 0)
            out_rows.append({
                "Direction": "egress",
                "Strategy": strategy,
                "Source": roles_str,
                "Destination": "Any (0.0.0.0/0)",
                "Services": "All Services",
                "sum_num_flows": sum_all,
                "Rule Section": "intra-scope",
                "Ruleset": ruleset_name,
            })
        return out_rows

    allow_intervals = _complement_from_blacklist(blacklist_intervals)
    default_services = _render_intervals_as_services(allow_intervals, include_misc=True)

    # DEFAULT rule: ONE per scope
    any_dst = "Any (0.0.0.0/0)"
    for (app, env) in scopes:
        ruleset_name = f"{app}-{env}-RS"
        roles_str = "|".join(sorted(roles_by_scope.get((app, env), set())))
        if not roles_str:
            continue
        allowed_sum = 0
        for (a2, e2, role2, pref2, proto2, port2) in by_key.keys():
            if (a2, e2) == (app, env) and not _is_blacklisted(blacklist_intervals, proto2, port2):
                allowed_sum += sum_by_key.get((a2, e2, role2, pref2, proto2, port2), 0)
        out_rows.append({
            "Direction": "egress",
            "Strategy": strategy,
            "Source": roles_str,
            "Destination": any_dst,
            "Services": default_services,
            "sum_num_flows": allowed_sum,
            "Rule Section": "intra-scope",
            "Ruleset": ruleset_name,
        })

    # EXCEPTIONS: only blacklisted ports observed
    exc_items = []
    for (app, env, role, pref, proto, port), peers in by_key.items():
        if _is_blacklisted(blacklist_intervals, proto, port):
            exc_items.append(((app, env, role, pref, proto, port), peers))
    exc_items = sorted(exc_items, key=lambda kv: (kv[0][0], kv[0][1], kv[0][3], kv[0][2], kv[0][4], kv[0][5]))

    for (app, env, role, pref, proto, port), peers in exc_items:
        ruleset_name = f"{app}-{env}-RS"
        out_rows.append({
            "Direction": "egress",
            "Strategy": strategy,
            "Source": role,
            "Destination": "|".join(sorted(peers)),
            "Services": _service_token(proto, port),
            "sum_num_flows": sum_by_key.get((app, env, role, pref, proto, port), 0),
            "Rule Section": "intra-scope",
            "Ruleset": ruleset_name,
        })

    return out_rows


def _anchor_selector_for_unscoped(anchor_app: str, anchor_env: str, anchor_role: str) -> str:
    """Selector for the anchor side when we need explicit labels (unscoped or other-scope rules)."""
    a = (anchor_app or "").strip()
    e = (anchor_env or "").strip()
    r = (anchor_role or "").strip()
    # keep deterministic order and omit empty parts
    parts = []
    if a:
        parts.append(f"app={a}")
    if e:
        parts.append(f"env={e}")
    if r:
        parts.append(f"role={r}")
    return "|".join(parts)


def _peer_labels_components(peer_value: str) -> Tuple[str, str, str, str, bool]:
    """Parse peer labels selector and return (peer_app, peer_env, peer_role, prefix, is_managed)."""
    try:
        kv = parse_kv_tokens(_as_text(peer_value or ""))
    except Exception:
        kv = {}
    peer_app = _as_text(kv.get("app") or "").strip()
    peer_env = _as_text(kv.get("env") or "").strip()
    peer_role = _as_text(kv.get("role") or "").strip()
    pref = _app_prefix(peer_app)
    is_managed = bool(pref and pref.endswith("M"))
    return peer_app, peer_env, peer_role, pref, is_managed


def build_egress_labels_proposed_rules_v1(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    finegrained_single_ports: Optional[Dict[str, List[Tuple[int, int]]]] = None,
) -> List[Dict[str, Any]]:
    """Proposed rules V1 for egress flows where peer_type='labels' and matched_rule_category != Bouquets.

    Egress placement rules (per spec):
      - If destination label is Managed (app prefix ends with 'M'):
          Ruleset = <peer_app>-<peer_env>-RS
          Rule Section = 'Extrascope in other scope'
          Destination = 'app=<peer_app>|env=<peer_env>|role=<peer_role_or_All Roles>'
      - If destination label is Unmanaged:
          Ruleset = <anchor_app>-<anchor_env>-OUTBOUND2<PREFIX>-RS
          Rule Section = 'Unscopped in OUTBOUND2<PREFIX>'
          Destination = 'app=<peer_app>|env=<peer_env>'

    Proposed rules1 readability improvements (THIS FUNCTION ONLY affects Proposed rules1):
      - Destination (Managed) includes app/env/role (not only role).
      - Aggregation is *destination-centric*:
          * finegrained: for the same (anchor_app, anchor_env, ruleset, section, destination),
            aggregate ALL anchor roles into Source (single app/env + multi-role) and aggregate all ports into Services (';' separated).
            Ports listed in carto.conf PORTS_ADMIN/PORTS_TO_ERADICATE/PORTS_TO_CONTROL are kept one per rule.
          * blacklist: same for NON-blacklisted ports, but keep each blacklisted port as its own row (one port per row),
            with Comment='Blacklist Exception'. Roles are still aggregated per blacklisted port.
    """
    strategy = (strategy or "").strip().lower()
    if strategy not in ("allow", "finegrained", "blacklist"):
        return []

    def _render_selector(app: str, env: str, roles: Optional[Set[str]] = None) -> str:
        a = (app or "").strip()
        e = (env or "").strip()
        parts: List[str] = []
        if a:
            parts.append(f"app={a}")
        if e:
            parts.append(f"env={e}")
        if roles:
            rr = sorted(set([x.strip() for x in roles if str(x).strip()]))
            if rr:
                parts.append(f"role={';'.join(rr)}")
        return "|".join(parts)

    def _render_dst_managed(peer_app: str, peer_env: str, peer_role: str) -> str:
        # Always include app/env. Keep role deterministic; if missing, keep a readable placeholder.
        role_val = (peer_role or "").strip() or "All Roles"
        return _render_selector(peer_app, peer_env, roles={role_val})

    out_rows: List[Dict[str, Any]] = []

    # Non-blacklisted (and finegrained) aggregation:
    # key = (ruleset_name, rule_section, anchor_app, anchor_env, destination [, service_token])
    allow_groups: Dict[Tuple[Any, ...], Dict[str, Any]] = {}

    # Blacklisted exceptions aggregation:
    # key = (ruleset_name, rule_section, anchor_app, anchor_env, destination, proto, port)
    bl_exc_groups: Dict[Tuple[str, str, str, str, str, str, int], Dict[str, Any]] = {}

    for p in proposals:
        if str(p.get("direction", "") or "").strip() != "egress":
            continue
        if str(p.get("peer_type", "") or "").strip() != "labels":
            continue
        if str(p.get("matched_rule_category", "") or "").strip() == CAT_BOUQUET:
            continue


        # network-zone North/South special case:
        # keep NS rows only when the flow has NO existing rule match (to avoid bloating proposals)
        is_ns = bool(p.get("nz_ns"))
        if is_ns:
            if str(p.get("Info", "") or "").strip() != INFO_NOMATCH:
                continue

        anchor_app = _as_text(p.get("anchor_app") or "").strip()
        anchor_env = _as_text(p.get("anchor_env") or "").strip()
        anchor_role = _as_text(p.get("anchor_role") or "").strip()
        if not anchor_app or not anchor_env or not anchor_role:
            continue

        proto = _as_text(p.get("proto") or "").strip().lower()
        try:
            port_i = int(str(p.get("port") or "0").strip() or "0")
        except Exception:
            port_i = 0
        if not proto or port_i <= 0:
            # proto-only traffic (ICMP/IGMP) not handled here (keeps behavior consistent with ingress extrascope)
            continue

        try:
            ssum = int(str(p.get("num_flows", "0") or "0").strip() or "0")
            ssum_true = int(str(p.get("num_flows_true", p.get("num_flows", "0")) or "0").strip() or "0")
        except Exception:
            ssum = 0

        pv = p.get("peer_value", "")
        pv_list: List[str]
        if isinstance(pv, (set, list, tuple)):
            pv_list = [str(x) for x in pv if str(x).strip()]
        else:
            pv_list = [str(pv or "")]

        for pv_raw in pv_list:
            if not str(pv_raw).strip():
                continue
            # support multiple selectors only if upstream aggregated values inserted newlines;
            # do NOT split on '|' (it's the internal key/value separator inside one selector)
            for pv_item in [x.strip() for x in str(pv_raw).split("\n") if x.strip()]:
                peer_app, peer_env, peer_role, pref, is_managed = _peer_labels_components(pv_item)
                if not peer_app or not peer_env or not pref:
                    continue

                if is_managed:
                    ruleset_name = f"{peer_app}-{peer_env}-RS"
                    rule_section = "Extrascope in other scope"
                    dst = _render_dst_managed(peer_app, peer_env, peer_role)
                else:
                    ruleset_name = f"{anchor_app}-{anchor_env}-OUTBOUND2{pref}-RS"
                    rule_section = f"Unscopped in OUTBOUND2{pref}"
                    dst = _render_selector(peer_app, peer_env, roles=None)

                tok = _service_token(proto, port_i)

                if strategy == "allow":
                    gk = (ruleset_name, rule_section, anchor_app, anchor_env, dst)
                    gg = allow_groups.get(gk)
                    if gg is None:
                        allow_groups[gk] = {
                            "Direction": "egress",
                            "Strategy": strategy,
                            "_roles": {anchor_role},
                            "Destination": dst,
                            "Services": "All Services",
                            "sum_num_flows": ssum,
                            "sum_num_flows_true": ssum_true,
                            "Rule Section": rule_section,
                            "Comment": "",
                            "Ruleset": ruleset_name,
                            "_anchor_app": anchor_app,
                            "_anchor_env": anchor_env,
                            "_east_west": "N" if is_ns else "Y",
                        }
                    else:
                        gg["_roles"].add(anchor_role)
                        gg["sum_num_flows"] = int(gg.get("sum_num_flows", 0) or 0) + ssum
                        gg["sum_num_flows_true"] = int(gg.get("sum_num_flows_true", 0) or 0) + ssum_true
                        if is_ns:
                            gg["_east_west"] = "N"
                    continue

                if strategy == "blacklist" and _is_blacklisted(blacklist_intervals, proto, port_i):
                    bk = (ruleset_name, rule_section, anchor_app, anchor_env, dst, proto, port_i)
                    gg = bl_exc_groups.get(bk)
                    if gg is None:
                        bl_exc_groups[bk] = {
                            "_roles": {anchor_role},
                            "sum_num_flows": ssum,
                            "sum_num_flows_true": ssum_true,
                            "_east_west": "N" if is_ns else "Y",
                        }
                    else:
                        gg["_roles"].add(anchor_role)
                        gg["sum_num_flows"] = int(gg.get("sum_num_flows", 0) or 0) + ssum
                        gg["sum_num_flows_true"] = int(gg.get("sum_num_flows_true", 0) or 0) + ssum_true
                        if is_ns:
                            gg["_east_west"] = "N"
                    continue

                # finegrained OR non-blacklisted in blacklist mode -> group by destination (not per-role)
                if strategy == "finegrained" and _is_port_in_intervals(finegrained_single_ports, proto, port_i):
                    gk = (ruleset_name, rule_section, anchor_app, anchor_env, dst, tok)
                else:
                    gk = (ruleset_name, rule_section, anchor_app, anchor_env, dst)
                gg = allow_groups.get(gk)
                if gg is None:
                    allow_groups[gk] = {
                        "Direction": "egress",
                        "Strategy": strategy,
                        "_roles": {anchor_role},
                        "Destination": dst,
                        "_svc_set": {tok},
                        "_svc_all": False,
                        "sum_num_flows": ssum,
                        "sum_num_flows_true": ssum_true,
                        "Rule Section": rule_section,
                        "Comment": "",
                        "Ruleset": ruleset_name,
                        "_anchor_app": anchor_app,
                        "_anchor_env": anchor_env,
                        "_east_west": "N" if is_ns else "Y",
                    }
                else:
                    gg["_roles"].add(anchor_role)
                    gg["_svc_set"].add(tok)
                    gg["sum_num_flows"] = int(gg.get("sum_num_flows", 0) or 0) + ssum
                    if is_ns:
                        gg["_east_west"] = "N"

    # Materialize grouped rows
    def _allow_group_sort_key(k: Tuple[Any, ...]) -> Tuple[Any, ...]:
        svc = k[5] if len(k) > 5 else ""
        return (k[0], k[1], k[4], k[2], k[3], svc)

    for gk in sorted(allow_groups.keys(), key=_allow_group_sort_key):
        r = allow_groups[gk]
        anchor_app = str(r.get("_anchor_app") or "")
        anchor_env = str(r.get("_anchor_env") or "")
        roles = set(r.get("_roles") or set())
        src_selector = _render_selector(anchor_app, anchor_env, roles=roles)

        rr: Dict[str, Any] = {
            "Direction": "egress",
            "Strategy": r.get("Strategy", strategy),
            "Source": src_selector,
            "Destination": r.get("Destination", ""),
            "Services": r.get("Services", ""),
            "sum_num_flows": r.get("sum_num_flows", 0),
            "sum_num_flows_true": r.get("sum_num_flows_true", r.get("sum_num_flows", 0)),
            "Rule Section": r.get("Rule Section", ""),
            "Comment": r.get("Comment", ""),
            "Ruleset": r.get("Ruleset", ""),
            "East-West (Y/N)": r.get("_east_west", "Y"),
        }

        if rr.get("Services") != "All Services":
            toks = sorted(list(r.get("_svc_set") or []), key=_service_sort_key)
            rr["Services"] = ";".join(toks)

        out_rows.append(rr)

    # Add blacklisted exception rows (each kept as its own line; roles aggregated)
    for bk in sorted(bl_exc_groups.keys(), key=lambda k: (k[0], k[1], k[4], k[2], k[3], k[5], k[6])):
        ruleset_name, rule_section, anchor_app, anchor_env, dst, proto, port_i = bk
        g = bl_exc_groups[bk]
        roles = set(g.get("_roles") or set())
        src_selector = _render_selector(anchor_app, anchor_env, roles=roles)
        out_rows.append({
            "Direction": "egress",
            "Strategy": strategy,
            "Source": src_selector,
            "Destination": dst,
            "Services": _service_token(proto, port_i),
            "sum_num_flows": int(g.get("sum_num_flows", 0) or 0),
            "sum_num_flows_true": int(g.get("sum_num_flows_true", g.get("sum_num_flows", 0)) or 0),
            "Rule Section": rule_section,
            "Comment": "Blacklist Exception",
            "Ruleset": ruleset_name,
            "East-West (Y/N)": g.get("_east_west", "Y"),
        })

    return out_rows


def build_egress_proposed_rules_v1(
    proposals: List[Dict[str, Any]],
    strategy: str,
    blacklist_intervals: Optional[Dict[str, List[Tuple[int, int]]]] = None,
    blacklist_lists_used: Optional[List[str]] = None,
    finegrained_single_ports: Optional[Dict[str, List[Tuple[int, int]]]] = None,
) -> List[Dict[str, Any]]:
    """Build V1 Proposed rules rows for egress (adds Comment + label handling + readability grouping for Proposed rules1).

    IMPORTANT:
      - This function ONLY impacts the additive sheet "Proposed rules1".
      - The stable "Proposed rules" sheet contract remains unchanged.

    Egress + peer_type=iplist (Proposed rules1):
      - Never put multiple IPLISTs in the same Destination cell.
        If peer_value contains multiple IPLIST names separated by '|' or newlines, we *explode* them (one IPList per row).
      - Finegrained: for the same (anchor_app, anchor_env, Ruleset, Rule Section, Destination-IPList),
        aggregate ALL anchor roles into Source and ALL observed service tokens into Services (';' separated).
        Ports listed in carto.conf PORTS_ADMIN/PORTS_TO_ERADICATE/PORTS_TO_CONTROL are kept one per rule.
      - Blacklist: same aggregation for NON-blacklisted ports, but keep each blacklisted port as its own row
        (one port per row) with Comment='Blacklist Exception'. Roles are aggregated per (Destination-IPList, port).

    Label-based egress handling is appended via build_egress_labels_proposed_rules_v1().
    """
    strategy_norm = str(strategy or "").strip().lower()
    if strategy_norm not in ("allow", "finegrained", "blacklist"):
        return []

    def _to_int_local(v: Any) -> int:
        try:
            return int(str(v or "0").strip() or "0")
        except Exception:
            return 0

    # Keep the historical behavior for "allow" (broad rules) and only enhance Proposed rules1 in finegrained/blacklist.
    if strategy_norm == "allow":
        out: List[Dict[str, Any]] = []
        out.extend(build_egress_proposed_rules(proposals, strategy_norm, blacklist_intervals, blacklist_lists_used))
        out.extend(build_egress_labels_proposed_rules_v1(proposals, strategy, blacklist_intervals, finegrained_single_ports))
        return out

    default_rows: List[Dict[str, Any]] = []
    if strategy_norm == "blacklist":
        # DEFAULT rule (egress/blacklist):
        #   Source      = all roles in the scope
        #   Destination = Any (0.0.0.0/0)
        #   Services    = All services except blacklisted ports
        roles_by_scope_default: Dict[Tuple[str, str], Set[str]] = {}
        scopes_default: Set[Tuple[str, str]] = set()

        for p_any in proposals:
            if str(p_any.get("Info", "") or "").strip() == "Bouquets Infra":
                continue
            app_any = _as_text(p_any.get("anchor_app", "") or "").strip()
            env_any = _as_text(p_any.get("anchor_env", "") or "").strip()
            role_any = _as_text(p_any.get("anchor_role", "") or "").strip()
            if app_any and env_any and role_any:
                roles_by_scope_default.setdefault((app_any, env_any), set()).add(role_any)
            if str(p_any.get("direction", "") or "").strip().lower() == "egress" and app_any and env_any:
                scopes_default.add((app_any, env_any))

        if blacklist_intervals:
            allow_intervals = _complement_from_blacklist(blacklist_intervals)
            default_services = _render_intervals_as_services(allow_intervals, include_misc=True)
        else:
            default_services = "All Services"

        any_dst = "Any (0.0.0.0/0)"
        for (app, env) in sorted(scopes_default):
            roles = sorted(list(roles_by_scope_default.get((app, env), set())))
            if not roles:
                continue

            # Sum flows that are allowed by the default rule (NOT blacklisted)
            allowed_sum = 0
            allowed_sum_true = 0
            for p2 in proposals:
                if str(p2.get("direction", "") or "").strip().lower() != "egress":
                    continue
                if str(p2.get("Info", "") or "").strip() == "Bouquets Infra":
                    continue
                if _as_text(p2.get("anchor_app", "") or "").strip() != app:
                    continue
                if _as_text(p2.get("anchor_env", "") or "").strip() != env:
                    continue

                proto2 = _as_text(p2.get("proto", "") or "").strip().lower()
                port2 = _to_int_local(p2.get("port"))
                if port2 <= 0:
                    continue
                if blacklist_intervals and _is_blacklisted(blacklist_intervals, proto2, port2):
                    continue
                allowed_sum += _to_int_local(p2.get("num_flows"))
                allowed_sum_true += _to_int_local(p2.get("num_flows_true", p2.get("num_flows")))

            default_rows.append({
                "Direction": "egress",
                "Strategy": strategy_norm,
                "Source": ";".join(roles),
                "Destination": any_dst,
                "Services": default_services,
                "sum_num_flows": allowed_sum,
                "sum_num_flows_true": allowed_sum_true,
                "Rule Section": "intra-scope",
                "Comment": "Blacklist default rule",
                "Ruleset": f"{app}-{env}-RS",
            })

    # ------------------------------ IPList-based egress (intra-scope) ------------------------------
    # Aggregation is destination-centric (one IPList per line).
    allow_aggs: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    exc_aggs: Dict[Tuple[str, str, str, str, str, str, int], Dict[str, Any]] = {}

    for p in proposals:
        if str(p.get("direction", "") or "").strip().lower() != "egress":
            continue
        if str(p.get("Info", "")).strip() == "Bouquets Infra":
            continue
        if str(p.get("peer_type", "")).strip() != "iplist":
            continue

        app = _as_text(p.get("anchor_app", "") or "").strip()
        env = _as_text(p.get("anchor_env", "") or "").strip()
        role = _as_text(p.get("anchor_role", "") or "").strip()
        if not app or not env or not role:
            continue

        proto = _as_text(p.get("proto", "") or "").strip().lower()
        try:
            port = int(str(p.get("port", "") or "0").strip() or "0")
        except Exception:
            port = 0
        if not proto or port <= 0:
            continue

        # Use per-flow num_flows when available; fall back to sum_num_flows.
        nf = _to_int_local(p.get("num_flows"))
        if nf <= 0:
            nf = _to_int_local(p.get("sum_num_flows"))

        peer_val = _as_text(p.get("peer_value", "") or "", joiner="|").strip()
        if not peer_val:
            continue

        # Ensure "one IPList per row": explode any multi-valued peer_value.
        peer_val = peer_val.replace("\n", "|")
        iplists = [x.strip() for x in peer_val.split("|") if x.strip()]
        if not iplists:
            continue

        ruleset_name = f"{app}-{env}-RS"
        rule_section = "intra-scope"
        svc_tok = _service_token(proto, port)

        for ipn in iplists:
            # In egress/blacklist, intra-scope DEFAULT rule covers all NON-blacklisted ports,
            # so we ONLY generate rules for blacklisted ports (exceptions).
            if strategy_norm == "blacklist":
                if _is_blacklisted(blacklist_intervals, proto, port):
                    k_exc = (app, env, ruleset_name, rule_section, ipn, proto, port)
                    a = exc_aggs.get(k_exc)
                    if a is None:
                        a = {
                            "Direction": "egress",
                            "Strategy": strategy_norm,
                            "Source": "",
                            "Destination": ipn,
                            "Services": svc_tok,
                            "sum_num_flows": 0,
                            "sum_num_flows_true": 0,
                            "Rule Section": rule_section,
                            "Comment": "Blacklist Exception",
                            "Ruleset": ruleset_name,
                            "_roles": set(),
                        }
                        exc_aggs[k_exc] = a
                    a["_roles"].add(role)
                    a["sum_num_flows"] = _to_int_local(a.get("sum_num_flows")) + nf
                    a["sum_num_flows_true"] = _to_int_local(a.get("sum_num_flows_true")) + _to_int_local(p.get("num_flows_true", p.get("num_flows")))
                continue

            # Finegrained: aggregate all observed proto/port tokens per IPList.
            if _is_port_in_intervals(finegrained_single_ports, proto, port):
                k = (app, env, ruleset_name, rule_section, ipn, svc_tok)
            else:
                k = (app, env, ruleset_name, rule_section, ipn)
            a = allow_aggs.get(k)
            if a is None:
                a = {
                    "Direction": "egress",
                    "Strategy": strategy_norm,
                    "Source": "",
                    "Destination": ipn,
                    "Services": "",
                    "sum_num_flows": 0,
                    "Rule Section": rule_section,
                    "Comment": "",
                    "Ruleset": ruleset_name,
                    "_roles": set(),
                    "_svc_all": False,
                    "_svc_set": set(),
                }
                allow_aggs[k] = a
            a["_roles"].add(role)
            a["sum_num_flows"] = _to_int_local(a.get("sum_num_flows")) + nf
            # iplist egress services are explicit proto/port tokens
            if svc_tok == "All Services":
                a["_svc_all"] = True
            else:
                a["_svc_set"].add(svc_tok)

    out_rows: List[Dict[str, Any]] = []
    out_rows.extend(default_rows)

    # Deterministic ordering
    def _allow_aggs_sort_key(k: Tuple[Any, ...]) -> Tuple[Any, ...]:
        svc = k[5] if len(k) > 5 else ""
        return (k[0], k[1], k[4], svc)

    for k in sorted(allow_aggs.keys(), key=_allow_aggs_sort_key):  # app, env, dest
        r = allow_aggs[k]
        roles = sorted(list(r.get("_roles") or []))
        r["Source"] = ";".join(roles)
        if r.get("_svc_all"):
            r["Services"] = "All Services"
        else:
            toks = sorted(list(r.get("_svc_set") or []), key=_service_sort_key)
            r["Services"] = ";".join(toks)
        r.pop("_roles", None)
        r.pop("_svc_set", None)
        r.pop("_svc_all", None)
        out_rows.append(r)

    for k in sorted(exc_aggs.keys(), key=lambda x: (x[0], x[1], x[4], x[5], x[6])):  # app, env, dest, proto, port
        r = exc_aggs[k]
        roles = sorted(list(r.get("_roles") or []))
        r["Source"] = ";".join(roles)
        r.pop("_roles", None)
        out_rows.append(r)

    # ------------------------------ Add label-based egress rules (extrascope / outbound2*) ------------------------------
    out_rows.extend(build_egress_labels_proposed_rules_v1(proposals, strategy, blacklist_intervals, finegrained_single_ports))
    return out_rows


def main() -> int:
    ap = argparse.ArgumentParser("propose_rule_for_scope_fix06")
    ap.add_argument("--input-raw", required=True, help="RUN/raw directory")
    ap.add_argument("--derived-dir", required=True, help="RUN/derived directory")
    ap.add_argument("--excel", help="RUN/excel/export_*.xlsx to update")
    ap.add_argument("--start", required=True)
    ap.add_argument("--end", required=True)
    ap.add_argument("--prefer-raw", action="store_true", help="ignore zone-filtered flows if present")
    ap.add_argument("--network-zone", default="", help="Exact IPList name defining the East-West zone; when set, flows are filtered to stay strictly inside this zone (both endpoints in-zone) and North/South default rules are added using ZNOT_<network-zone>.")
    ap.add_argument("--min-flows", type=int, default=1)
    ap.add_argument("--group-by", choices=["labels", "hostnames"], default="labels")
    ap.add_argument("--conf", default="carto.conf", help="Path to carto.conf for IPList patterns")
    ap.add_argument("--enable-avoid-label-pairs", action="store_true", default=False,
                    help="If enabled: when a peer app is listed in carto.conf AVOID_LABEL_PAIRS, replace peer labels selector by the best-matching KUB_* IPList (from raw/export_iplists.csv include). Fallback to raw peer IP when no match.")
    ap.add_argument("--sheet-name", default="Flow-Rule Match", help="Main analysis sheet name")
    ap.add_argument("--log-level", default="INFO")
    ap.add_argument("--debug", action="store_true")
    ap.add_argument("--debug-proposed-rules", action="store_true", help="Emit full stack traces + dump non-string fields when Proposed rules generation fails")
    ap.add_argument("--strategy-intra-app", choices=["allow", "finegrained", "blacklist", "none"], default="none",
                    help="Intra-app rule proposal strategy. When set, a new sheet \"Proposed rules\" is generated.")
    ap.add_argument("--strategy-egress", choices=["allow", "finegrained", "blacklist", "none"], default="none",
                    help="Egress rule proposal strategy (generates Proposed rules rows). (Planned)")
    ap.add_argument("--strategy-ingress", choices=["allow", "finegrained", "blacklist", "none"], default="none",
                    help="Ingress rule proposal strategy (generates Proposed rules rows).")
    # Legacy (backward compatible) blacklist sources. Prefer per-direction flags when multiple blacklist directions are used.
    ap.add_argument("--ports-to-blacklist", default="",
                    help="Legacy: carto.conf port-list names (comma/semicolon/space separated), e.g. PORTS_TO_CONTROL,PORTS_ADMIN. If multiple directions use blacklist, use the per-direction flags.")
    ap.add_argument("--ports-to-blacklist-intra-app", default="",
                    help="carto.conf port-list names for intra-app blacklist strategy (e.g. PORTS_TO_CONTROL,PORTS_ADMIN)")
    ap.add_argument("--ports-to-blacklist-egress", default="",
                    help="carto.conf port-list names for egress blacklist strategy (future)")
    ap.add_argument("--ports-to-blacklist-ingress", default="",
                    help="carto.conf port-list names for ingress blacklist strategy (future)")
    # Default behavior: enabled. Use --no-mark-potential-core-service to disable.
    ap.add_argument("--no-mark-potential-core-service", dest="mark_potential_core_service", action="store_false", default=True,
                    help="Disable Potential Core Service detection in Proposed rules1 (orange highlight + Comment).")


    args = ap.parse_args()

    setup_logging(args.log_level)
    info(f"CWD: {Path.cwd()}")
    raw_dir = Path(args.input_raw).expanduser()
    derived_dir = Path(args.derived_dir).expanduser()
    excel_path = Path(args.excel).expanduser() if args.excel else None

    raw_dir, derived_dir, excel_path = cohere_paths(raw_dir, derived_dir, excel_path, debug=args.debug)
    if not raw_dir.exists():
        raise FileNotFoundError(f"raw_dir not found: {raw_dir}")
    if not derived_dir.exists():
        raise FileNotFoundError(f"derived_dir not found: {derived_dir}")

    conf_path = find_conf_path(args.conf, debug=args.debug)
    if conf_path:
        info(f"conf (resolved): {conf_path}")
    allowed, prio = load_conf(conf_path, debug=args.debug)

    # Optional: replace some peer app labels by KUB_* IPLISTS (used to avoid proposing rules based on those labels).
    avoid_apps: Set[str] = set()
    kub_resolver: Optional[Callable[[str], str]] = None
    if args.enable_avoid_label_pairs:
        avoid_apps = load_avoid_label_pairs(conf_path, debug=args.debug)
        if not avoid_apps:
            warn("AVOID_LABEL_PAIRS is empty; --enable-avoid-label-pairs has no effect.")
        kub_index = build_kub_iplist_network_index(raw_dir, debug=args.debug)
        if not kub_index:
            warn("KUB_* IPList index is empty (raw/export_iplists.csv missing or no KUB_ entries); fallback to raw peer IP.")
            kub_resolver = None
        else:
            kub_resolver = make_kub_iplist_resolver(kub_index)

    finegrained_single_ports, _fg_lists = _load_default_finegrained_single_ports(conf_path)
    if _fg_lists:
        dbg(args.debug, f"finegrained single-port lists: {_fg_lists}")
    port_list_intervals = _load_port_list_intervals(conf_path)

    
    # Optional: --network-zone (East-West only inside one IPList)
    network_zone_name: str = ""
    network_zone_znot: str = ""
    network_zone_nets: Optional[List[ipaddress._BaseNetwork]] = None
    if getattr(args, "network_zone", ""):
        network_zone_name, network_zone_znot, network_zone_nets = _resolve_network_zone_iplists(
            raw_dir, str(args.network_zone), debug=args.debug
        )
        info(
            f"network-zone enabled: in-zone only (iplist='{network_zone_name}'), "
            f"north/south complement iplist='{network_zone_znot}'"
        )

# compute applicable rules (frozen)
    t0 = time.perf_counter()
    rules_applicables, _unmatched, _eff = get_applicable_rules(raw_dir, derived_dir)  # frozen logic
    info(f"applicable rules computed: {len(rules_applicables)} in {time.perf_counter() - t0:.2f}s")

    # load labelgroups for matching (same loader as frozen module)
    lg_path = raw_dir / "export_labelgroup.csv"
    if lg_path.exists():
        lg_map = load_labelgroups(lg_path)
        info(f"labelgroups loaded: {len(lg_map)} entries from {lg_path.name}")
    else:
        lg_map = {}
        warn(f"export_labelgroup.csv not found in {raw_dir} — labelgroup matching will be degraded.")

    rows, action_rows = propose(
        raw_dir=raw_dir,
        derived_dir=derived_dir,
        start=args.start,
        end=args.end,
        prefer_raw=args.prefer_raw,
        min_flows=args.min_flows,
        group_by=args.group_by,
        allowed_pats=allowed,
        prio_pats=prio,
        rules_applicables=rules_applicables,
        lg_map=lg_map,
        debug=args.debug,
        enable_avoid_label_pairs=args.enable_avoid_label_pairs,
        avoid_label_pairs=avoid_apps,
        kub_iplist_resolver=kub_resolver,
        network_zone_nets=network_zone_nets,
        network_zone_name=network_zone_name,
    )

    rr = run_root_from(raw_dir) or run_root_from(derived_dir) or (excel_path and run_root_from(excel_path))
    if rr:
        out_csv = rr / "derived" / f"flow_rule_match_{args.start}_{args.end}.csv"
    else:
        out_csv = derived_dir / f"flow_rule_match_{args.start}_{args.end}.csv"
    rows_flow_match = [r for r in rows if not bool(r.get("nz_ns"))]
    write_csv(out_csv, rows_flow_match)

    if excel_path:
        # Main analysis sheet (new name)
        rows_sheet = list(rows_flow_match)
        append_excel(excel_path, args.sheet_name, rows_sheet)

        # Backward compatibility: keep writing "Proposed Rules" too (same content)
        if args.sheet_name != "Proposed Rules":
            append_excel(excel_path, "Proposed Rules", rows_sheet)
        # NEW: Proposed rules (intra-app + ingress strategies)
        try:
            pr_rows: List[Dict[str, str]] = []
            intra_bl_intervals = None
            intra_bl_lists = None

            # Intra-app
            if args.strategy_intra_app != "none":
                if args.strategy_intra_app == "blacklist":
                    if not conf_path:
                        raise FileNotFoundError("carto.conf not found: required for --strategy-intra-app=blacklist")
                    bl_specs = _resolve_blacklist_sources_for_direction(args)
                    bl_spec_intra = bl_specs.get("intra-app", "")
                    if not str(bl_spec_intra).strip():
                        raise ValueError("--strategy-intra-app=blacklist requires a blacklist source list. Use --ports-to-blacklist-intra-app=... (or legacy --ports-to-blacklist=... when only one blacklist direction is used).")
                    intra_bl_intervals, intra_bl_lists = _expand_blacklist_sources(conf_path, bl_spec_intra)
                    pr_rows.extend(build_intra_app_proposed_rules(
                        rows,
                        raw_dir,
                        args.strategy_intra_app,
                        intra_bl_intervals,
                        intra_bl_lists,
                        finegrained_single_ports,
                    ))
                else:
                    pr_rows.extend(build_intra_app_proposed_rules(
                        rows,
                        raw_dir,
                        args.strategy_intra_app,
                        None,
                        None,
                        finegrained_single_ports,
                    ))

            # Ingress
            ingress_bl_intervals = None
            ingress_bl_lists = None
            if getattr(args, "strategy_ingress", "none") != "none":
                if args.strategy_ingress == "blacklist":
                    if not conf_path:
                        raise FileNotFoundError("carto.conf not found: required for --strategy-ingress=blacklist")
                    bl_specs = _resolve_blacklist_sources_for_direction(args)
                    bl_spec_ing = bl_specs.get("ingress", "")
                    if not str(bl_spec_ing).strip():
                        raise ValueError("--strategy-ingre...-to-blacklist=... when only one blacklist direction is used).")
                    ingress_bl_intervals, ingress_bl_lists = _expand_blacklist_sources(conf_path, bl_spec_ing)
                pr_rows.extend(build_ingress_proposed_rules(rows, args.strategy_ingress, ingress_bl_intervals, ingress_bl_lists))

            # Egress
            egress_bl_intervals = None
            egress_bl_lists = None
            if getattr(args, "strategy_egress", "none") not in ("none", "", None):
                if args.strategy_egress == "blacklist":
                    if not conf_path:
                        raise FileNotFoundError("carto.conf not found: required for --strategy-egress=blacklist")
                    bl_specs = _resolve_blacklist_sources_for_direction(args)
                    bl_spec_eg = bl_specs.get("egress", "")
                    if not str(bl_spec_eg).strip():
                        raise ValueError("--strategy-egress=blacklist requires a blacklist source list. Use --ports-to-blacklist-egress=KEY1,KEY2 (or legacy --ports-to-blacklist=... only when exactly one direction uses blacklist).")
                    egress_bl_intervals, egress_bl_lists = _expand_blacklist_sources(conf_path, bl_spec_eg)
                    pr_rows.extend(build_egress_proposed_rules(rows, args.strategy_egress, egress_bl_intervals, egress_bl_lists))
                else:
                    pr_rows.extend(build_egress_proposed_rules(rows, args.strategy_egress, None, None))


            # ------------------------------ North/South default rules (network-zone) ------------------------------
            # When --network-zone is enabled, we filter the analysis to East-West flows strictly inside the zone.
            # To keep connectivity with everything OUTSIDE the zone, we add 2 default rules:
            #   - ingress:  Source = ZNOT_<zone>  -> Destination = all roles of the scope
            #   - egress :  Source = all roles of the scope -> Destination = ZNOT_<zone>
            ns_pr1_rows: List[Dict[str, Any]] = []
            if network_zone_nets and network_zone_znot:
                scopes: Set[Tuple[str, str]] = set()
                for r0 in rows:
                    a0 = _as_text(r0.get("anchor_app") or "").strip()
                    e0 = _as_text(r0.get("anchor_env") or "").strip()
                    if a0 and e0:
                        scopes.add((a0, e0))

                if not scopes:
                    warn("network-zone enabled but could not infer (app, env) scopes from Flow-Rule Match rows; skipping North/South default rules.")
                else:
                    wkld_m_path = raw_dir / "export_wkld.m.csv"
                    for (app0, env0) in sorted(scopes):
                        roles0 = _gather_roles_from_wkld_m(wkld_m_path, app0, env0)
                        if not roles0:
                            for r0 in rows:
                                if _as_text(r0.get("anchor_app") or "").strip() == app0 and _as_text(r0.get("anchor_env") or "").strip() == env0:
                                    rr0 = _as_text(r0.get("anchor_role") or "").strip()
                                    if rr0:
                                        roles0.add(rr0)
                        if not roles0:
                            roles0 = {"All Roles"}
                        roles_list = sorted(list(roles0))

                        # Ingress NS default rule
                        if getattr(args, "strategy_ingress", "none") != "none":
                            pr_rows.append({
                                "Direction": "ingress",
                                "Strategy": str(getattr(args, "strategy_ingress", "") or "").strip().lower(),
                                "Source": network_zone_znot,
                                "Destination": ";".join(roles_list),
                                "Services": "All Services",
                                "sum_num_flows": 0,
                                "Rule Section": "North South default rule",
                                "Ruleset": f"{app0}-{env0}-RS",
                            })
                            ns_pr1_rows.append({
                                "Direction": "ingress",
                                "Strategy": str(getattr(args, "strategy_ingress", "") or "").strip().lower(),
                                "Source": network_zone_znot,
                                "Destination": ";".join(roles_list),
                                "Services": "All Services",
                                "sum_num_flows": 1,
                                "sum_num_flows_true": 0,
                                "East-West (Y/N)": "N",
                                "Rule Section": "North South default rule",
                                "Comment": "North South default rule",
                                "Ruleset": f"{app0}-{env0}-RS",
                            })

                        # Egress NS default rule
                        if getattr(args, "strategy_egress", "none") != "none":
                            pr_rows.append({
                                "Direction": "egress",
                                "Strategy": str(getattr(args, "strategy_egress", "") or "").strip().lower(),
                                "Source": ";".join(roles_list),
                                "Destination": network_zone_znot,
                                "Services": "All Services",
                                "sum_num_flows": 0,
                                "Rule Section": "North South default rule",
                                "Ruleset": f"{app0}-{env0}-RS",
                            })
                            ns_pr1_rows.append({
                                "Direction": "egress",
                                "Strategy": str(getattr(args, "strategy_egress", "") or "").strip().lower(),
                                "Source": ";".join(roles_list),
                                "Destination": network_zone_znot,
                                "Services": "All Services",
                                "sum_num_flows": 1,
                                "sum_num_flows_true": 0,
                                "East-West (Y/N)": "N",
                                "Rule Section": "North South default rule",
                                "Comment": "North South default rule",
                                "Ruleset": f"{app0}-{env0}-RS",
                            })

            if pr_rows:
                dir_order = {"intra-app": 0, "ingress": 1, "egress": 2}
                pr_rows = sorted(pr_rows, key=lambda r: (
                    dir_order.get(str(r.get("Direction","")).strip(), 9),
                    str(r.get("Ruleset","")),
                    str(r.get("Rule Section","")),
                    str(r.get("Destination","")),
                    str(r.get("Source","")),
                    str(r.get("Services","")),
                ))
                append_excel_simple_table(excel_path, PROPOSED_RULES_SHEET, PROPOSED_RULES_HEADER, pr_rows, wrap_cols={"Services"})
                info(f"Proposed rules generated: {len(pr_rows)} rows (intra-app={args.strategy_intra_app}, ingress={getattr(args, 'strategy_ingress', 'none')})")
            # Proposed rules1 (additive sheet for iterative formatting improvements)
            pr_rows1: List[Dict[str, Any]] = []
            pr_rows1.extend(ns_pr1_rows)
            if getattr(args, "strategy_intra_app", "none") != "none":
                intra_rows_pr1 = build_intra_app_proposed_rules(
                    rows,
                    raw_dir,
                    args.strategy_intra_app,
                    intra_bl_intervals,
                    intra_bl_lists,
                    finegrained_single_ports,
                )
                for rr in intra_rows_pr1:
                    r2 = dict(rr)
                    comment = ""
                    # In Proposed rules1: keep "Rule Section" as the physical location, and use Comment for exceptions
                    if str(args.strategy_intra_app).strip().lower() == "blacklist":
                        if "exception" in str(r2.get("Rule Section", "")).lower():
                            r2["Rule Section"] = "Intrascope"
                            comment = _merge_comment(comment, "Blacklist Exception")
                        else:
                            comment = _merge_comment(comment, "Blacklist default rule")
                    r2["Comment"] = comment
                    pr_rows1.append(r2)
            if getattr(args, "strategy_ingress", "none") != "none":
                pr_rows1.extend(build_ingress_proposed_rules_v1(
                    rows,
                    args.strategy_ingress,
                    ingress_bl_intervals,
                    ingress_bl_lists,
                ))
            # Egress
            if getattr(args, "strategy_egress", "none") != "none":
                egress_bl_intervals = None
                egress_bl_lists = None
                if args.strategy_egress == "blacklist":
                    if not conf_path:
                        raise FileNotFoundError("carto.conf not found: required for --strategy-egress=blacklist")
                    bl_specs = _resolve_blacklist_sources_for_direction(args)
                    bl_spec_eg = bl_specs.get("egress", "")
                    if not str(bl_spec_eg).strip():
                        raise ValueError("--strategy-egress=blacklist requires a blacklist source list. Use --ports-to-blacklist-egress=KEY1,KEY2 (or legacy --ports-to-blacklist=... only when exactly one direction uses blacklist).")
                    egress_bl_intervals, egress_bl_lists = _expand_blacklist_sources(conf_path, bl_spec_eg)
                pr_rows1.extend(build_egress_proposed_rules_v1(
                    rows,
                    args.strategy_egress,
                    egress_bl_intervals,
                    egress_bl_lists,
                    finegrained_single_ports,
                ))
            
            if pr_rows1:
                pr_rows1 = _group_pr1_ingress_finegrained_by_src_dst(pr_rows1, finegrained_single_ports)
                if port_list_intervals:
                    for rr in pr_rows1:
                        comment_add = _port_list_comment_for_services(rr.get("Services", ""), port_list_intervals)
                        if comment_add:
                            rr["Comment"] = _merge_comment(rr.get("Comment", ""), comment_add)
                if getattr(args, "mark_potential_core_service", False):
                    rr_root = run_root_from(raw_dir)
                    enabled_rules_csv = rr_root / "raw" / "export_rules.enabled.csv"
                    bouquets_apps, bouquets_iplists = _load_bouquets_refs(enabled_rules_csv)
                    for rr in pr_rows1:
                        if _row_matches_bouquets_refs(rr, bouquets_apps, bouquets_iplists):
                            rr["Comment"] = _merge_comment(rr.get("Comment", ""), "Remote (App label/iplist) used in Bouquets")
                
                def _pr1_sort_key(r: Dict[str, Any]) -> Tuple[Any, ...]:
                    direction = str(r.get("Direction", "") or "").strip().lower()
                    comment = str(r.get("Comment", "") or "")
                    is_remote = "Remote (App label/iplist) used in Bouquets" in comment
                    ruleset = str(r.get("Ruleset", "") or "")
                    rule_section = str(r.get("Rule Section", "") or "")
                    destination = str(r.get("Destination", "") or "")
                    source = str(r.get("Source", "") or "")
                    services = str(r.get("Services", "") or "")

                    if is_remote:
                        if direction == "egress":
                            group = 0
                            primary = destination
                        elif direction == "ingress":
                            group = 1
                            primary = source
                        else:
                            group = 2
                            primary = ""
                    else:
                        if direction == "egress":
                            group = 3
                            primary = destination
                        elif direction == "ingress":
                            group = 4
                            primary = source
                        else:
                            group = 5
                            primary = ""

                    return (
                        group,
                        primary,
                        ruleset,
                        rule_section,
                        destination,
                        source,
                        services,
                        comment,
                    )

                pr_rows1 = sorted(pr_rows1, key=_pr1_sort_key)
                
                def _pr1_row_fill(r: Dict[str, Any]) -> Optional[str]:
                    return "FFFFC000" if "Remote (App label/iplist) used in Bouquets" in str(r.get("Comment","") or "") else None
                
                def _pr1_row_bold(r: Dict[str, Any]) -> bool:
                    return "Blacklist default rule" in str(r.get("Comment","") or "")
                
                # Normalize PR1 flow counters:
                #   - num_aggregated_rows = previous sum_num_flows (row count of aggregated lines)
                #   - sum_num_flows      = real sum of flows (from raw [Num Flows], col AG)
                pr_rows1_norm: List[Dict[str, Any]] = []
                for _r in pr_rows1:
                    _rr = dict(_r)
                    _n_rows = _to_int(_rr.pop("sum_num_flows", 0))
                    _n_true = _to_int(_rr.pop("sum_num_flows_true", _n_rows))
                    _rr["num_aggregated_rows"] = _n_rows
                    _rr["sum_num_flows"] = _n_true
                    pr_rows1_norm.append(_rr)

                pr1_header = list(PROPOSED_RULES1_HEADER)
                if network_zone_nets:
                    pr1_header.insert(pr1_header.index("Rule Section"), "East-West (Y/N)")
                    for _rr in pr_rows1_norm:
                        if not _rr.get("East-West (Y/N)"):
                            _rr["East-West (Y/N)"] = "Y"

                append_excel_simple_table(
                    excel_path,
                    PROPOSED_RULES1_SHEET,
                    pr1_header,
                    pr_rows1_norm,
                    wrap_cols={"Services"},
                    row_fill_fn=(_pr1_row_fill if getattr(args, "mark_potential_core_service", False) else None),
                    row_bold_fn=_pr1_row_bold,
                )
                logger.info("Wrote %d Proposed rules1 rows", len(pr_rows1_norm))

                try:
                    add_flow_rule_match_legend(excel_path, sheet_name=getattr(args, "sheet_name", "Flow-Rule Match"))
                except Exception as e:
                    warn(f"Could not update Flow-Rule Match legend: {e}")
            
        except Exception as e:
            warn(f"Could not generate Proposed rules sheet: {e}")
            import traceback
            warn("---- Proposed rules stack trace (most recent call last) ----")
            warn(traceback.format_exc())
            if getattr(args, "debug_proposed_rules", False) or getattr(args, "debug", False):
                _debug_dump_non_string_fields(rows, limit=60)
            warn("---- End Proposed rules stack trace ----")



        # Action plan sheet
        append_excel(excel_path, "Action Plan", action_rows)

        # Post-formatting pass (presentation only)
        try:
            post_format_excel_presentation(excel_path)
        except Exception as e:
            warn(f"Post-formatting pass failed: {e}")


        info(f"Excel updated: {excel_path} (sheets='{args.sheet_name}', 'Proposed Rules', 'Action Plan')")
    else:
        info("No --excel provided: CSV only.")

    return 0

if __name__ == "__main__":
    sys.exit(main())
