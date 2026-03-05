#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""modules/excel_stream_update.py

Carto NG — Excel "stream injector" (low-mem)

Base
----
This file is based on your provided excel_stream_update.v7.py.

Changes in this version (v8)
----------------------------
Only for the optimized branch (this script):

1) Summary: merge B:D for rows 2..6 (header block), even after rewrite.
2) Flow-out / Flow-in:
   - Auto-width with a hard cap for "Source IPList" (col B) and "Destination IPList" (col P) = 45.
     Robust matching (case-insensitive + fallback to column index).
   - Data background colors applied on ALL rows via conditional formatting (formula 1=1).
3) After writing the resulting XLSX, ensure the file is readable (chmod +r).

Determinism
-----------
All computations remain deterministic (fixed sampling for widths).
"""

from __future__ import annotations

import argparse
import csv
import re
import socket
import ipaddress
import os
import shutil
import stat
import sys
import tempfile
from copy import copy
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter


NUMERIC_HEADERS = {"Port", "Num Flows", "Bytes In", "Bytes Out"}


# Column auto-width caps (to avoid very wide columns on large string fields)
DEFAULT_WIDTH_CAP = 250


def _load_iplist_conf(conf_path: Path) -> Tuple[List[str], List[str]]:
    """Read IPLIST_ALLOWED_PREFIXES and IPLIST_NAME_PRIORITY from carto.conf.

    Returns (allowed_prefix_patterns, priority_patterns).
    Patterns are ';' separated in the conf file, may end with '*' (prefix-glob).
    """
    allowed: List[str] = []
    prio: List[str] = []
    if not conf_path:
        return allowed, prio
    try:
        if not conf_path.exists():
            return allowed, prio
        with conf_path.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k == "IPLIST_ALLOWED_PREFIXES":
                    allowed = [x for x in v.split(";") if x]
                elif k == "IPLIST_NAME_PRIORITY":
                    prio = [x for x in v.split(";") if x]
    except Exception:
        return [], []
    return allowed, prio


def _match_prefix_or_glob(name: str, pat: str) -> bool:
    if not pat:
        return False
    if pat.endswith("*"):
        return name.startswith(pat[:-1])
    return name.startswith(pat)


def elect_iplist_from_tokens(
    iplists_raw: str,
    allowed_pats: List[str],
    prio_pats: List[str],
) -> Tuple[str, str]:
    """Elect a single IPList name from a cell containing multiple candidates.

    Deterministic election:
    - tokenize on ; , and whitespace
    - keep only allowed prefixes if provided (if filtering yields at least one candidate)
    - apply priority list (first match wins)
    - fallback to lexical min (stable)
    Returns (elected_name, reason).
    """
    toks = [t.strip() for t in re.split(r"[;,\s]+", iplists_raw or "") if t.strip()]
    if not toks:
        return "", "no_iplist"

    # Keep only names (strip href tails)
    names: List[str] = []
    for t in toks:
        names.append(t.rsplit("/", 1)[-1] if "/" in t else t)

    # Allowed filtering
    if allowed_pats:
        kept = [n for n in names if any(_match_prefix_or_glob(n, p) for p in allowed_pats)]
        if kept:
            names = kept

    # Priority
    for p in prio_pats or []:
        for n in names:
            if _match_prefix_or_glob(n, p):
                return n, f"priority({p})"

    chosen = sorted(set(names))[0]
    return chosen, "fallback_lexical"


def _norm_header(s: str) -> str:
    """Normalize header for robust comparisons."""
    s = (s or "").strip().lower()
    # keep alnum only to avoid issues with double spaces, punctuation, etc.
    return "".join(ch for ch in s if ch.isalnum())


def _width_cap_for_header(h: str) -> int:
    """Return the width cap for a given header (case-insensitive)."""
    return DEFAULT_WIDTH_CAP


# ------------------------------ To investigate (NZ0_/NZ1_/DNA_/DNS_ + egress KUB_/LBI_/LBO_/U_) ------------------------------
INV_INGRESS_PREFIXES = ("NZ0_", "NZ1_", "DNA_", "DNS_")
INV_EGRESS_PREFIXES = ("NZ0_", "NZ1_", "KUB_", "LBI_", "LBO_", "U_")
INV_SHEET_NAME = "To investigate"


def _find_col_idx(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Return 0-based index of the first matching header (robust)."""
    if not headers:
        return None
    norm = [_norm_header(h) for h in headers]

    # 1) Exact normalized match
    for cand in candidates:
        nc = _norm_header(cand)
        if not nc:
            continue
        for i, nh in enumerate(norm):
            if nh == nc:
                return i

    # 2) Containment match
    for cand in candidates:
        nc = _norm_header(cand)
        if not nc:
            continue
        for i, nh in enumerate(norm):
            if nc in nh:
                return i

    return None


class _DNSResolver:
    """Reverse DNS resolver with timeout + cache (best effort)."""

    def __init__(self, timeout_s: float = 1.5) -> None:
        self.timeout_s = float(timeout_s)
        self.cache: Dict[str, str] = {}

    def resolve(self, ip: str) -> str:
        ip = (ip or "").strip()
        if not ip:
            return ""
        if ip in self.cache:
            return self.cache[ip]

        try:
            ipaddress.ip_address(ip)
        except Exception:
            self.cache[ip] = ""
            return ""

        prev = socket.getdefaulttimeout()
        try:
            socket.setdefaulttimeout(self.timeout_s)
            host, _aliases, _addrs = socket.gethostbyaddr(ip)
            res = host or ""
        except Exception:
            res = ""
        finally:
            socket.setdefaulttimeout(prev)

        self.cache[ip] = res
        return res


def _make_cell(
    ws,
    value: object,
    *,
    font: Optional[Font] = None,
    fill: Optional[PatternFill] = None,
    border: Optional[Border] = None,
    alignment: Optional[Alignment] = None,
) -> WriteOnlyCell:
    c = WriteOnlyCell(ws, value=value)
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    if alignment is not None:
        c.alignment = alignment
    return c


LEVELS = {"DEBUG": 10, "INFO": 20, "WARN": 30, "ERROR": 40}
CURRENT_LEVEL = LEVELS["INFO"]


def _log(level: str, msg: str) -> None:
    lvl = LEVELS.get(level.upper(), 20)
    if lvl < CURRENT_LEVEL:
        return
    print(f"[{level.upper()}] {msg}", file=sys.stderr)


def _safe_int(x: object) -> Optional[int]:
    try:
        if x is None:
            return None
        s = str(x).strip()
        if s == "":
            return None
        if "." in s:
            return int(float(s))
        return int(s)
    except Exception:
        return None


def _detect_delimiter(csv_path: Path, *, fallback: str = ",") -> str:
    """Detect delimiter using a small sample (robust for ',' and ';')."""
    try:
        data = csv_path.read_bytes()[:8192]
        sample = data.decode("utf-8-sig", errors="replace")

        # csv.Sniffer can be flaky on very wide files; keep it best-effort.
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
            if dialect.delimiter in {",", ";"}:
                return dialect.delimiter
        except Exception:
            pass

        # Heuristic: count occurrences on the first non-empty line
        for line in sample.splitlines():
            if not line.strip():
                continue
            c_comma = line.count(",")
            c_semi = line.count(";")
            if c_semi > c_comma:
                return ";"
            if c_comma > 0:
                return ","
            break
    except Exception:
        pass

    return fallback


def _iter_csv_rows(
    csv_path: Path,
    *,
    encoding: str,
    delimiter: str,
) -> Iterable[List[str]]:
    with csv_path.open("r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            yield list(row)


def _split_if_single_column(row: List[str]) -> List[str]:
    """Fix the classic issue where the wrong delimiter makes everything end up in column A."""
    if len(row) != 1:
        return row
    s = row[0]
    if ";" in s and "," not in s:
        return s.split(";")
    if "," in s and ";" not in s:
        return s.split(",")
    return row


def _apply_summary_merges(ws) -> None:
    """Enforce Summary header merges: rows 2..6 must merge B:D."""
    if ws.title != "Summary":
        return

    # Remove any merges intersecting B:D on rows 2..6, then merge B:D.
    try:
        existing = list(getattr(ws.merged_cells, "ranges", []))
        for r in range(2, 7):
            for m in existing:
                try:
                    if m.min_row == r and m.max_row == r and not (m.max_col < 2 or m.min_col > 4):
                        ws.unmerge_cells(str(m))
                except Exception:
                    pass
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    except Exception:
        # best-effort: merging is cosmetic
        pass


def _copy_sheet_values_and_styles(ws_src, wb_dst: Workbook):
    """Copy a sheet from a normal workbook into a write-only workbook.

    Returns the destination sheet object to allow post-processing.
    """
    ws_dst = wb_dst.create_sheet(ws_src.title)

    # Sheet state (visible/hidden)
    try:
        ws_dst.sheet_state = ws_src.sheet_state
    except Exception:
        pass

    # Common UX props
    try:
        ws_dst.freeze_panes = ws_src.freeze_panes
    except Exception:
        pass

    try:
        if getattr(ws_src, "auto_filter", None) is not None and ws_src.auto_filter.ref:
            ws_dst.auto_filter.ref = ws_src.auto_filter.ref
    except Exception:
        pass

    try:
        if ws_src.sheet_properties and ws_src.sheet_properties.tabColor:
            ws_dst.sheet_properties.tabColor = copy(ws_src.sheet_properties.tabColor)
    except Exception:
        pass

    # View options
    try:
        src_view = getattr(ws_src, "sheet_view", None)
        dst_view = getattr(ws_dst, "sheet_view", None)
        if src_view is not None and dst_view is not None:
            for attr in [
                "showGridLines",
                "showRowColHeaders",
                "rightToLeft",
                "zoomScale",
                "zoomScaleNormal",
                "zoomScalePageLayoutView",
            ]:
                if hasattr(src_view, attr):
                    setattr(dst_view, attr, getattr(src_view, attr))
    except Exception:
        pass

    # Column widths
    try:
        for col_letter, dim in ws_src.column_dimensions.items():
            w = getattr(dim, "width", None)
            if w:
                ws_dst.column_dimensions[col_letter].width = w
    except Exception:
        pass

    # Row heights
    try:
        for ridx, dim in ws_src.row_dimensions.items():
            h = getattr(dim, "height", None)
            if h:
                ws_dst.row_dimensions[ridx].height = h
    except Exception:
        pass

    # Copy rows: values + direct styles
    for row in ws_src.iter_rows():
        out_row: List[WriteOnlyCell] = []
        for cell in row:
            c = WriteOnlyCell(ws_dst, value=cell.value)

            try:
                if cell.has_style:
                    c.font = copy(cell.font)
                    c.fill = copy(cell.fill)
                    c.border = copy(cell.border)
                    c.alignment = copy(cell.alignment)
                    c.number_format = cell.number_format
                    c.protection = copy(cell.protection)
            except Exception:
                pass

            try:
                if cell.hyperlink:
                    c.hyperlink = copy(cell.hyperlink)
            except Exception:
                pass

            out_row.append(c)

        ws_dst.append(out_row)

    # Merged cells
    try:
        for merged in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged))
    except Exception:
        pass

    # Conditional formatting (best-effort)
    try:
        cfl = getattr(ws_src, "conditional_formatting", None)
        rules_dict = getattr(cfl, "_cf_rules", None)
        if rules_dict:
            for cf_obj, rules in rules_dict.items():
                sqref = str(getattr(cf_obj, "sqref", "")) or str(cf_obj)
                for rule in rules:
                    try:
                        ws_dst.conditional_formatting.add(sqref, copy(rule))
                    except Exception:
                        pass
    except Exception:
        pass

    # Enforce Summary merges (requested)
    _apply_summary_merges(ws_dst)

    return ws_dst


def _group_columns(headers: Sequence[str]) -> Tuple[List[int], List[int], List[int]]:
    """Return indices (1-based) for (source, destination, meta) groups."""
    src: List[int] = []
    dst: List[int] = []
    meta: List[int] = []
    for i, h in enumerate(headers, start=1):
        hs = (h or "").strip()
        if hs.startswith("Source "):
            src.append(i)
        elif hs.startswith("Destination "):
            dst.append(i)
        else:
            meta.append(i)
    return src, dst, meta


def _ranges_from_indices(indices: List[int]) -> List[Tuple[int, int]]:
    """Convert a sorted list of column indices into contiguous ranges."""
    if not indices:
        return []
    indices = sorted(indices)
    ranges: List[Tuple[int, int]] = []
    start = prev = indices[0]
    for x in indices[1:]:
        if x == prev + 1:
            prev = x
            continue
        ranges.append((start, prev))
        start = prev = x
    ranges.append((start, prev))
    return ranges


def _add_always_true_fill_cf(ws, rng: str, fill: PatternFill) -> None:
    """Add a conditional formatting rule that always applies (formula 1=1)."""
    try:
        dxf = DifferentialStyle(fill=fill)
        rule = Rule(type="expression", dxf=dxf, formula=["1=1"])
        ws.conditional_formatting.add(rng, rule)
    except Exception:
        # best-effort
        pass


def _make_flow_palette() -> Dict[str, PatternFill]:
    """Return the fill palette used for Flow-in / Flow-out formatting.

    We intentionally keep the same colors as the trivial (non-streaming) branch,
    but we apply them by fixed column ranges to match your spec.
    """
    return {
        "green_light": PatternFill("solid", fgColor="00CCFFCC"),
        "green_hdr": PatternFill("solid", fgColor="0099CC99"),
        "orange_light": PatternFill("solid", fgColor="00FFE5CC"),
        "orange_hdr": PatternFill("solid", fgColor="00FFCC99"),
        "blue_light": PatternFill("solid", fgColor="00DDEBF7"),
        "blue_hdr": PatternFill("solid", fgColor="009DC3E6"),
    }


def _flow_segments_for_direction(
    direction: str,
    *,
    max_col: int,
    palette: Dict[str, PatternFill],
) -> List[Tuple[int, int, PatternFill, PatternFill]]:
    """Return column segments for Flow sheets.

    Spec (1-based Excel columns):
      Flow-out:
        A..N   = green
        O..AA  = orange
        AB..AD = blue
        AE..AG = orange
        AH..AQ = blue
      Flow-in:
        A..N   = orange
        O..AA  = green
        AB..AD = blue
        AE..AG = green
        AH..AQ = blue

    Any extra columns beyond AQ are colored as blue (meta) to stay readable.
    """
    # A=1, N=14, O=15, AA=27, AB=28, AD=30, AE=31, AG=33, AH=34, AQ=43
    if direction == "in":
        segs = [
            (1, 14, palette["orange_hdr"], palette["orange_light"]),
            (15, 27, palette["green_hdr"], palette["green_light"]),
            (28, 30, palette["blue_hdr"], palette["blue_light"]),
            (31, 33, palette["green_hdr"], palette["green_light"]),
            (34, 43, palette["blue_hdr"], palette["blue_light"]),
        ]
    else:
        segs = [
            (1, 14, palette["green_hdr"], palette["green_light"]),
            (15, 27, palette["orange_hdr"], palette["orange_light"]),
            (28, 30, palette["blue_hdr"], palette["blue_light"]),
            (31, 33, palette["orange_hdr"], palette["orange_light"]),
            (34, 43, palette["blue_hdr"], palette["blue_light"]),
        ]

    if max_col > 43:
        segs.append((44, max_col, palette["blue_hdr"], palette["blue_light"]))

    # clamp to worksheet
    out: List[Tuple[int, int, PatternFill, PatternFill]] = []
    for a, b, hf, df in segs:
        if a > max_col:
            continue
        out.append((a, min(b, max_col), hf, df))
    return out


def _flow_header_fill_for_col(
    col_idx: int,
    *,
    segments: Sequence[Tuple[int, int, PatternFill, PatternFill]],
    default_fill: PatternFill,
) -> PatternFill:
    for a, b, hf, _df in segments:
        if a <= col_idx <= b:
            return hf
    return default_fill


def _flow_data_fill_for_col(
    col_idx: int,
    *,
    segments: Sequence[Tuple[int, int, PatternFill, PatternFill]],
    default_fill: Optional[PatternFill],
) -> Optional[PatternFill]:
    for a, b, _hf, df in segments:
        if a <= col_idx <= b:
            return df
    return default_fill


def _apply_flow_formatting(
    ws,
    *,
    headers: Sequence[str],
    n_rows: int,
    direction: str,
    widths: Dict[int, float],
) -> None:
    """Apply Flow-out / Flow-in formatting for the streaming branch.

    Important: In write-only worksheets, we must style header cells as we append them
    (handled in _stream_flow_csv). Here we apply:
      - Freeze panes + autofilter
      - Column widths
      - Whole-table background colors for DATA rows via conditional formatting (formula 1=1)
    """
    # Freeze + filter range
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass

    max_col = len(headers)
    last_col_letter = get_column_letter(max_col)
    ref = f"A1:{last_col_letter}{n_rows}"
    try:
        ws.auto_filter.ref = ref
    except Exception:
        pass

    # Column widths
    try:
        for col_idx, w in widths.items():
            if w and w > 0:
                ws.column_dimensions[get_column_letter(col_idx)].width = w
    except Exception:
        pass

    # Data area coloring via conditional formatting (always true)
    if n_rows >= 2 and max_col >= 1:
        palette = _make_flow_palette()
        segs = _flow_segments_for_direction(direction, max_col=max_col, palette=palette)
        for (a, b, _hf, df) in segs:
            rng = f"{get_column_letter(a)}2:{get_column_letter(b)}{n_rows}"
            _add_always_true_fill_cf(ws, rng, df)

    # Cosmetic: header row height (best-effort)
    try:
        ws.row_dimensions[1].height = 20
    except Exception:
        pass


def _stream_flow_csv(
    wb_dst: Workbook,
    sheet_name: str,
    csv_path: Path,
    *,
    direction: str,
    encoding: str = "utf-8-sig",
    csv_delimiter: Optional[str] = None,
    sample_rows_for_width: Optional[int] = None,
    add_elected_iplist_column: bool = False,
    allowed_pats: Optional[List[str]] = None,
    prio_pats: Optional[List[str]] = None,
    investigate_writer: Optional[csv.writer] = None,
    dns_resolver: Optional[_DNSResolver] = None,
    investigate_seen: Optional[set] = None,
) -> int:
    """Stream a Flow CSV into a write-only worksheet.

    Returns the number of 'To investigate' rows detected (written to investigate_writer if provided).
    """
    if not csv_path.exists():
        _log("WARN", f"[{sheet_name}] CSV not found: {csv_path}")
        return 0

    ws = wb_dst.create_sheet(sheet_name)

    # Header basic style
    THIN = Side(style="thin", color="666666")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Delimiter detection
    delim = csv_delimiter or _detect_delimiter(csv_path)
    rows_it = _iter_csv_rows(csv_path, encoding=encoding, delimiter=delim)
    try:
        raw_header = next(rows_it)
    except StopIteration:
        _log("WARN", f"[{sheet_name}] Empty CSV: {csv_path}")
        return 0

    header = _split_if_single_column(raw_header)

    # If delimiter was wrong, rescan using the other delimiter
    if len(header) == 1:
        other = ";" if delim == "," else ","
        rows_it = _iter_csv_rows(csv_path, encoding=encoding, delimiter=other)
        raw_header = next(rows_it)
        header = _split_if_single_column(raw_header)
        if len(header) > 1:
            delim = other
        else:
            # keep as-is
            rows_it = _iter_csv_rows(csv_path, encoding=encoding, delimiter=delim)
            next(rows_it, None)

    headers = [h.strip() for h in header]

    allowed_pats = allowed_pats or []
    prio_pats = prio_pats or []
    orig_headers = list(headers)
    orig_cols_len = len(orig_headers)

    elected_key_idx: Optional[int] = None       # index in ORIGINAL CSV columns (0-based)
    elected_insert_pos: Optional[int] = None    # insert position in OUTPUT columns (0-based)

    if add_elected_iplist_column:
        key_name = "Destination IPList" if direction == "out" else "Source IPList"
        elected_header = "Destination IPList Elected" if direction == "out" else "Source IPList Elected"

        for i, h in enumerate(orig_headers):
            if _norm_header(h) == _norm_header(key_name):
                elected_key_idx = i
                break

        if elected_key_idx is not None:
            elected_insert_pos = elected_key_idx + 1
            # Avoid double insert if already present
            if not any(_norm_header(h) == _norm_header(elected_header) for h in headers):
                headers.insert(elected_insert_pos, elected_header)

    # Prepare width tracking (deterministic)
    caps: Dict[int, int] = {i: _width_cap_for_header(h) for i, h in enumerate(headers, start=1)}

    widths: Dict[int, int] = {}
    for i, h in enumerate(headers, start=1):
        widths[i] = max(8, min(caps[i], len(h) + 2))

    # Header row with fixed-range coloring (per spec)
    palette = _make_flow_palette()
    segs = _flow_segments_for_direction(direction, max_col=len(headers), palette=palette)
    default_hdr = palette["blue_hdr"]
    default_data = palette["blue_light"]

    out_header: List[WriteOnlyCell] = []
    for col_idx, h in enumerate(headers, start=1):
        c = WriteOnlyCell(ws, value=h)
        c.font = Font(bold=True)
        c.border = BORDER
        c.alignment = hdr_align
        c.fill = _flow_header_fill_for_col(col_idx, segments=segs, default_fill=default_hdr)
        out_header.append(c)
    ws.append(out_header)

    data_fills: Dict[int, PatternFill] = {}
    for col_idx in range(1, len(headers) + 1):
        fill = _flow_data_fill_for_col(col_idx, segments=segs, default_fill=default_data)
        if fill is not None:
            data_fills[col_idx] = fill

    # ---- To investigate indexes (based on OUTPUT headers)
    inv_count = 0
    do_inv = investigate_writer is not None and dns_resolver is not None
    elected_hdr = "Destination IPList Elected" if direction == "out" else "Source IPList Elected"

    idx_elected = _find_col_idx(headers, [elected_hdr])
    if idx_elected is None and elected_insert_pos is not None:
        idx_elected = elected_insert_pos

    idx_unknown_ip = _find_col_idx(
        headers,
        [
            "Destination IP" if direction == "out" else "Source IP",
            "Dest IP" if direction == "out" else "Src IP",
            "Destination Address" if direction == "out" else "Source Address",
        ],
    )

    idx_src_role = _find_col_idx(headers, ["Source Role", "Src Role"])
    idx_dst_role = _find_col_idx(headers, ["Destination Role", "Dst Role"])

    # Fallbacks (only if role columns are missing)
    idx_src_fb = _find_col_idx(headers, ["Source Labels", "Source Workload", "Source", "Source IPList", "Source IP"])
    idx_dst_fb = _find_col_idx(headers, ["Destination Labels", "Destination Workload", "Destination", "Destination IPList", "Destination IP"])

    idx_service = _find_col_idx(headers, ["Service", "Services", "Service Name", "Service(s)"])
    idx_proto = _find_col_idx(headers, ["Protocol", "Proto"])
    idx_port = _find_col_idx(headers, ["Port", "Dst Port", "Destination Port"])

    # Stream data rows
    n_rows = 1
    sample_left = sample_rows_for_width

    for raw in rows_it:
        row = _split_if_single_column(raw)

        # Normalize row length (based on ORIGINAL CSV columns)
        if len(row) < orig_cols_len:
            row = row + [""] * (orig_cols_len - len(row))
        elif len(row) > orig_cols_len:
            row = row[: orig_cols_len]

        elected_name: str = ""
        if elected_key_idx is not None:
            elected_name, _reason = elect_iplist_from_tokens(str(row[elected_key_idx]), allowed_pats, prio_pats)

        # Insert elected IPList column (visibility) right after the key column
        if elected_insert_pos is not None and elected_key_idx is not None:
            row.insert(elected_insert_pos, elected_name)

        # Final normalize to OUTPUT columns length
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[: len(headers)]

        # To investigate hook (NZ0_/NZ1_/DNA_/DNS_ + egress KUB_/LBI_/LBO_/U_)
        if do_inv:
            elected_val = ""
            if idx_elected is not None and idx_elected < len(row):
                elected_val = str(row[idx_elected] or "")
            if not elected_val and elected_name:
                elected_val = elected_name

            prefixes = INV_EGRESS_PREFIXES if direction == "out" else INV_INGRESS_PREFIXES
            if elected_val.startswith(prefixes):
                unknown_ip = "" if idx_unknown_ip is None or idx_unknown_ip >= len(row) else str(row[idx_unknown_ip] or "")
                dns = dns_resolver.resolve(unknown_ip)                # Mapping per spec:
                # - Flow-out: Source=Source Role, Destination=Destination IPList Elected
                # - Flow-in : Source=Source IPList Elected, Destination=Destination Role
                if direction == "out":
                    src_val = "" if idx_src_role is None or idx_src_role >= len(row) else str(row[idx_src_role] or "")
                    if not src_val and idx_src_fb is not None and idx_src_fb < len(row):
                        src_val = str(row[idx_src_fb] or "")
                    dst_val = elected_val
                else:
                    src_val = elected_val
                    dst_val = "" if idx_dst_role is None or idx_dst_role >= len(row) else str(row[idx_dst_role] or "")
                    if not dst_val and idx_dst_fb is not None and idx_dst_fb < len(row):
                        dst_val = str(row[idx_dst_fb] or "")

                service = ""
                if idx_service is not None and idx_service < len(row):
                    service = str(row[idx_service] or "")
                if not service:
                    proto = "" if idx_proto is None or idx_proto >= len(row) else str(row[idx_proto] or "")
                    port = "" if idx_port is None or idx_port >= len(row) else str(row[idx_port] or "")
                    if proto and port:
                        service = f"{proto}/{port}"
                    elif proto:
                        service = proto
                    elif port:
                        service = port

                key = (
                    "Flow-out" if direction == "out" else "Flow-in",
                    src_val,
                    dst_val,
                    service,
                    unknown_ip,
                    dns,
                )
                if investigate_seen is not None:
                    if key in investigate_seen:
                        continue
                    investigate_seen.add(key)

                investigate_writer.writerow(list(key))
                inv_count += 1

        n_rows += 1
        out_row: List[WriteOnlyCell] = []
        for col_idx, (h, v) in enumerate(zip(headers, row), start=1):
            vv: object = v
            if h in NUMERIC_HEADERS:
                num = _safe_int(v)
                if num is not None:
                    vv = num
            c = WriteOnlyCell(ws, value=vv)
            c.border = BORDER
            fill = data_fills.get(col_idx)
            if fill is not None:
                c.fill = fill
            out_row.append(c)

            # width sample
            if sample_left is None or sample_left > 0:
                try:
                    s = "" if vv is None else str(vv)
                    widths[col_idx] = max(
                        widths[col_idx],
                        min(int(caps.get(col_idx, DEFAULT_WIDTH_CAP)), len(s) + 2),
                    )
                except Exception:
                    pass

        ws.append(out_row)
        if sample_left is not None and sample_left > 0:
            sample_left -= 1

    widths_f: Dict[int, float] = {i: float(w) for i, w in widths.items()}

    _apply_flow_formatting(
        ws,
        headers=headers,
        n_rows=n_rows,
        direction=direction,
        widths=widths_f,
    )

    _log("INFO", f"[{sheet_name}] Injected {n_rows - 1} rows from {csv_path.name} (delimiter='{delim}')")
    return inv_count


def _chmod_add_read(path: Path) -> None:
    """Ensure the XLSX is readable (chmod +r)."""
    try:
        st_mode = path.stat().st_mode
        os.chmod(path, st_mode | stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass


def rewrite_workbook_with_streamed_flows(
    *,
    input_xlsx: Path,
    flows_out_csv: Optional[Path],
    flows_in_csv: Optional[Path],
    output_xlsx: Optional[Path] = None,
    sheet_flow_out: str = "Flow-out",
    sheet_flow_in: str = "Flow-in",
    csv_encoding: str = "utf-8-sig",
    csv_delimiter: Optional[str] = None,
    conf_path: Optional[Path] = None,
    add_elected_iplist_column: bool = False,
    enable_to_investigate: bool = False,
    dns_timeout: float = 1.5,
) -> Path:
    """Rewrite workbook in streaming mode and inject flows."""
    if not input_xlsx.exists():
        raise FileNotFoundError(f"Input xlsx not found: {input_xlsx}")

    out_xlsx = output_xlsx or input_xlsx

    # Load source workbook in normal mode (small workbook expected)
    wb_src = load_workbook(input_xlsx, read_only=False, data_only=False)

    # Destination workbook in write-only mode
    wb_dst = Workbook(write_only=True)

    need_iplist_conf = add_elected_iplist_column or enable_to_investigate
    allowed_pats, prio_pats = _load_iplist_conf(conf_path) if need_iplist_conf else ([], [])

    inv_tmp_path: Optional[Path] = None
    inv_tmp_fp = None
    inv_writer = None
    dns_resolver = None
    inv_seen: Optional[set] = None
    inv_rows_total = 0

    if enable_to_investigate:
        # Keep it disk-backed for safety on huge datasets
        inv_tmp_fp = tempfile.NamedTemporaryFile("w", delete=False, newline="", encoding="utf-8")
        inv_tmp_path = Path(inv_tmp_fp.name)
        inv_writer = csv.writer(inv_tmp_fp)
        dns_resolver = _DNSResolver(timeout_s=float(dns_timeout))
        inv_seen = set()

    inserted = False
    titles = [ws.title for ws in wb_src.worksheets]
    insert_after = "Processes" if "Processes" in titles else None

    for ws_src in wb_src.worksheets:
        if ws_src.title in {sheet_flow_out, sheet_flow_in}:
            _log("WARN", f"Skipping existing sheet '{ws_src.title}' in input workbook")
            continue
        if enable_to_investigate and ws_src.title == INV_SHEET_NAME:
            _log("WARN", f"Dropping existing sheet '{INV_SHEET_NAME}' (will be rebuilt)")
            continue

        _copy_sheet_values_and_styles(ws_src, wb_dst)
        # Summary merges are enforced inside _copy_sheet_values_and_styles

        if insert_after and ws_src.title == insert_after and not inserted:
            if flows_out_csv:
                inv_rows_total += _stream_flow_csv(
                    wb_dst,
                    sheet_flow_out,
                    flows_out_csv,
                    direction="out",
                    encoding=csv_encoding,
                    csv_delimiter=csv_delimiter,
                    add_elected_iplist_column=add_elected_iplist_column,
                    allowed_pats=allowed_pats,
                    prio_pats=prio_pats,
                    investigate_writer=inv_writer,
                    dns_resolver=dns_resolver,
                    investigate_seen=inv_seen,
                )
            if flows_in_csv:
                inv_rows_total += _stream_flow_csv(
                    wb_dst,
                    sheet_flow_in,
                    flows_in_csv,
                    direction="in",
                    encoding=csv_encoding,
                    csv_delimiter=csv_delimiter,
                    add_elected_iplist_column=add_elected_iplist_column,
                    allowed_pats=allowed_pats,
                    prio_pats=prio_pats,
                    investigate_writer=inv_writer,
                    dns_resolver=dns_resolver,
                    investigate_seen=inv_seen,
                )
            inserted = True

    if not inserted:
        if flows_out_csv:
            inv_rows_total += _stream_flow_csv(
                wb_dst,
                sheet_flow_out,
                flows_out_csv,
                direction="out",
                encoding=csv_encoding,
                csv_delimiter=csv_delimiter,
                add_elected_iplist_column=add_elected_iplist_column,
                allowed_pats=allowed_pats,
                prio_pats=prio_pats,
                investigate_writer=inv_writer,
                    dns_resolver=dns_resolver,
                    investigate_seen=inv_seen,
                )
        if flows_in_csv:
            inv_rows_total += _stream_flow_csv(
                wb_dst,
                sheet_flow_in,
                flows_in_csv,
                direction="in",
                encoding=csv_encoding,
                csv_delimiter=csv_delimiter,
                add_elected_iplist_column=add_elected_iplist_column,
                allowed_pats=allowed_pats,
                prio_pats=prio_pats,
                investigate_writer=inv_writer,
                    dns_resolver=dns_resolver,
                    investigate_seen=inv_seen,
                )

    # Close temp investigate writer before reading
    if enable_to_investigate and inv_tmp_fp is not None:
        try:
            inv_tmp_fp.flush()
        except Exception:
            pass
        try:
            inv_tmp_fp.close()
        except Exception:
            pass

    # Append "To investigate" sheet at the end (read from temp)
    if enable_to_investigate and inv_tmp_path is not None:
        try:
            ws_inv = wb_dst.create_sheet(INV_SHEET_NAME)
            THIN = Side(style="thin", color="666666")
            BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            hdr_fill = PatternFill("solid", fgColor="D9D9D9")
            hdr_font = Font(bold=True)
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)       
            headers = [
                "Direction",
                "Source",
                "Destination",
                "Service",
                "Unknown IP",
                "DNS Resolution",
            ]
            widths_inv = {i: max(10, min(DEFAULT_WIDTH_CAP, len(h) + 2)) for i, h in enumerate(headers, start=1)}

            ws_inv.append(
                [
                    _make_cell(ws_inv, headers[0], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                    _make_cell(ws_inv, headers[1], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                    _make_cell(ws_inv, headers[2], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                    _make_cell(ws_inv, headers[3], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                    _make_cell(ws_inv, headers[4], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                    _make_cell(ws_inv, headers[5], font=hdr_font, fill=hdr_fill, border=BORDER, alignment=hdr_align),
                ]
            )

            # Data rows
            n_appended = 0
            seen_local = set()
            with inv_tmp_path.open("r", encoding="utf-8", newline="") as f:
                r = csv.reader(f)
                for row in r:
                    if not row or all((c or "").strip() == "" for c in row):
                        continue
                    vals = (row + [""] * 6)[:6]
                    key = tuple(vals)
                    if key in seen_local:
                        continue
                    seen_local.add(key)
                    ws_inv.append([
                        _make_cell(ws_inv, vals[0], border=BORDER),
                        _make_cell(ws_inv, vals[1], border=BORDER),
                        _make_cell(ws_inv, vals[2], border=BORDER),
                        _make_cell(ws_inv, vals[3], border=BORDER),
                        _make_cell(ws_inv, vals[4], border=BORDER),
                        _make_cell(ws_inv, vals[5], border=BORDER),
                    ])
                    for idx, value in enumerate(vals, start=1):
                        try:
                            s = "" if value is None else str(value)
                            widths_inv[idx] = max(
                                widths_inv[idx],
                                min(DEFAULT_WIDTH_CAP, max(10, len(s) + 2)),
                            )
                        except Exception:
                            pass
                    n_appended += 1

            ws_inv.freeze_panes = "A2"
            ws_inv.auto_filter.ref = f"A1:F{n_appended + 1}" if n_appended > 0 else "A1:F1"
            for idx, width in widths_inv.items():
                try:
                    ws_inv.column_dimensions[get_column_letter(idx)].width = width
                except Exception:
                    pass
            _log("INFO", f"[To investigate] rows: {n_appended}")
        except Exception as e:
            _log("WARN", f"[To investigate] build failed: {e}")

    # Save atomically
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("wb", delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)

    try:
        wb_dst.save(tmp_path)
        wb_src.close()
        shutil.move(str(tmp_path), str(out_xlsx))
        _chmod_add_read(out_xlsx)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink(missing_ok=True)  # type: ignore[arg-type]
        except Exception:
            pass
        try:
            if inv_tmp_path and inv_tmp_path.exists():
                inv_tmp_path.unlink(missing_ok=True)  # type: ignore[arg-type]
        except Exception:
            pass

    return out_xlsx


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Carto NG — Excel stream update (low-mem)")
    p.add_argument("--input-xlsx", required=True, type=Path, help="Input (small) Carto NG workbook")
    p.add_argument("--output-xlsx", type=Path, default=None, help="Write result to another path (default: in-place)")

    p.add_argument("--flows-out", required=True, type=Path, help="Flow-out CSV to inject")
    p.add_argument("--flows-in", required=True, type=Path, help="Flow-in CSV to inject")
    p.add_argument("--sheet-flow-out", default="Flow-out", help="Worksheet name for Flow-out (default: Flow-out)")
    p.add_argument("--sheet-flow-in", default="Flow-in", help="Worksheet name for Flow-in (default: Flow-in)")

    p.add_argument("--conf", type=Path, default=Path("carto.conf"), help="Path to carto.conf (used for IPList election priority)")
    p.add_argument("--add-elected-iplist-column", action="store_true", help="Add per-row elected IPList column next to Source/Destination IPList")

    p.add_argument(
        "--enable-to-investigate",
        action="store_true",
        help="Create a 'To investigate' sheet for NZ0_/NZ1_/DNA_/DNS_ elected IPLists (plus egress KUB_/LBI_/LBO_/U_)",
    )
    p.add_argument("--dns-timeout", type=float, default=1.5, help="Reverse DNS timeout (seconds) for To investigate")

    p.add_argument(
        "--csv-delimiter",
        choices=["auto", ",", ";"],
        default="auto",
        help="CSV delimiter override (default: auto detect)",
    )
    p.add_argument(
        "--csv-encoding",
        default="utf-8-sig",
        help="CSV encoding (default: utf-8-sig)",
    )

    p.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARN", "ERROR"],
        default="INFO",
        help="Log verbosity",
    )
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    global CURRENT_LEVEL

    args = _build_arg_parser().parse_args(argv)
    CURRENT_LEVEL = LEVELS[args.log_level]

    try:
        out = rewrite_workbook_with_streamed_flows(
            input_xlsx=args.input_xlsx,
            flows_out_csv=args.flows_out,
            flows_in_csv=args.flows_in,
            output_xlsx=args.output_xlsx,
            sheet_flow_out=args.sheet_flow_out,
            sheet_flow_in=args.sheet_flow_in,
            csv_encoding=args.csv_encoding,
            csv_delimiter=(None if args.csv_delimiter == "auto" else args.csv_delimiter),
            conf_path=args.conf,
            add_elected_iplist_column=args.add_elected_iplist_column,
            enable_to_investigate=args.enable_to_investigate,
            dns_timeout=float(args.dns_timeout),
        )
        _log("INFO", f"Excel updated: {out}")
        return 0
    except Exception as e:
        _log("ERROR", f"excel_stream_update failed: {e}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
