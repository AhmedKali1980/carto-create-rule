
# -*- coding: utf-8 -*-
"""
Module: export_services (v1.3)
- Génère derived/services_catalog.csv à partir des flows (in/out)
- Catégorise via carto.conf: PORTS_TO_CONTROL, PORTS_TO_ERADICATE, PORTS_ADMIN, ALLOWED_PORTS
- Append Excel (openpyxl uniquement) avec feuille "Service Catalogue"
- Mise en forme :
    * Ajuste la largeur des colonnes aux contenus
    * Freeze header + AutoFilter + Bold sur l'en-tête
    * Bordure fine (thin) sur toutes les cellules
    * Couleur de fond légère sur la colonne A selon 'category'
"""

import argparse
import csv
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------- Helpers ----------
def load_conf(path: Path) -> Dict[str, str]:
    conf: Dict[str, str] = {}
    if not path.exists():
        return conf
    with path.open(encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if "=" in s:
                k, v = s.split("=", 1)
                conf[k.strip()] = v.strip()
    return conf

def load_csv(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        rows = [{k: (v or "").strip() for k, v in row.items()} for row in r]
        return rows, (r.fieldnames or [])

def pick(cols: List[str], *cands: str) -> str:
    low = {c.lower(): c for c in (cols or [])}
    for c in cands:
        if c.lower() in low:
            return low[c.lower()]
    return ""

# ---------- Parse port lists from conf ----------
class PortSpec:
    __slots__ = ("proto", "start", "end")
    def __init__(self, proto: str, start: int, end: int):
        self.proto = proto.upper()
        self.start = int(start)
        self.end = int(end)
    def match(self, proto: str, port: int) -> bool:
        return self.proto == proto.upper() and self.start <= port <= self.end

def parse_port_specs(spec_str: str) -> List[PortSpec]:
    out: List[PortSpec] = []
    if not spec_str:
        return out
    items = [x.strip() for x in spec_str.split(";") if x.strip()]
    for it in items:
        if "/" not in it:
            # ICMP or IGMP
            proto = it.upper()
            out.append(PortSpec(proto, 0, 65535))
            continue
        proto, p = it.split("/", 1)
        proto = proto.strip().upper()
        p = p.strip()
        if "-" in p:
            a, b = [x.strip() for x in p.split("-", 1)]
            try:
                out.append(PortSpec(proto, int(a), int(b)))
            except Exception:
                pass
        else:
            try:
                out.append(PortSpec(proto, int(p), int(p)))
            except Exception:
                pass
    return out

# ---------- Core ----------
def normalize_flow_row(row: Dict[str, str], fns: List[str]) -> Optional[Dict[str, str]]:
    c_port  = pick(fns, "Port")
    c_proto = pick(fns, "Protocol", "Proto")
    c_tx    = pick(fns, "Transmission", "Direction")
    c_nf    = pick(fns, "Num Flows", "Hits")
    c_first = pick(fns, "First Detected", "First Seen")
    c_last  = pick(fns, "Last Detected", "Last Seen")
    if not c_port or not c_proto:
        return None
    try:
        port = int((row.get(c_port) or "0").strip() or 0)
    except Exception:
        port = 0
    proto = (row.get(c_proto) or "").strip().upper() or "TCP"
    tx    = (row.get(c_tx) or "").strip()
    nf    = (row.get(c_nf) or "").strip()
    first = (row.get(c_first) or "").strip()
    last  = (row.get(c_last) or "").strip()
    try:
        nf = int(nf) if nf else 0
    except Exception:
        nf = 0
    return {
        "proto": proto,
        "port": port,
        "tx": tx,
        "num_flows": nf,
        "first": first,
        "last": last,
    }

def categorize(proto: str, port: int, specs: Dict[str, List[PortSpec]]) -> str:
    # priorité: ERADICATE > CONTROL > ADMIN > ALLOWED > UNKNOWN
    for cat in ("ERADICATE", "CONTROL", "ADMIN", "ALLOWED"):
        for sp in specs.get(cat, []):
            if sp.match(proto, port):
                return cat
    return "UNKNOWN"

def to_rows_catalog(flows_norm: List[Dict[str, str]], specs: Dict[str, List[PortSpec]]) -> List[Dict[str, str]]:
    agg: Dict[tuple, Dict[str, str]] = {}
    for f in flows_norm:
        key = (f["proto"], f["port"])
        cur = agg.get(key)
        if not cur:
            agg[key] = {
                "category": categorize(f["proto"], f["port"], specs),
                "proto": f["proto"],
                "port": f["port"],
                "flows_total": 0,
                "flows_in": 0,
                "flows_out": 0,
                "first_seen": f["first"] or "",
                "last_seen": f["last"] or "",
            }
        cur = agg[key]
        cur["flows_total"] += f["num_flows"] or 1
        if (f["tx"] or "").lower().startswith("incoming"):
            cur["flows_in"] += f["num_flows"] or 1
        elif (f["tx"] or "").lower().startswith("outgoing"):
            cur["flows_out"] += f["num_flows"] or 1
        # update first/last
        if f["first"]:
            if not cur["first_seen"] or f["first"] < cur["first_seen"]:
                cur["first_seen"] = f["first"]
        if f["last"]:
            if not cur["last_seen"] or f["last"] > cur["last_seen"]:
                cur["last_seen"] = f["last"]
    out = list(agg.values())
    out.sort(key=lambda r: (r["category"], r["proto"], r["port"]))
    return out

def write_catalog_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    cols = ["category","proto","port","flows_total","flows_in","flows_out","first_seen","last_seen"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in cols})

# ---------- Excel append (openpyxl) ----------
def append_to_excel_openpyxl(excel_path: Path, rows: List[Dict[str, str]], sheet_name: str = "Service Catalogue") -> bool:
    """
    Remplace la feuille si elle existe; écrit les données; applique la mise en forme.
    """
    try:
        # Charger ou créer le workbook
        if excel_path.exists():
            wb = load_workbook(filename=str(excel_path))
        else:
            wb = Workbook()
            # Supprimer la feuille par défaut si on crée un nouveau fichier
            if wb.active and wb.active.title == "Sheet":
                wb.remove(wb.active)

        # Supprimer feuille existante si même nom
        if sheet_name in wb.sheetnames:
            ws_old = wb[sheet_name]
            wb.remove(ws_old)

        # Créer feuille
        ws = wb.create_sheet(title=sheet_name)

        # Colonnes (ordre)
        cols = ["category","proto","port","flows_total","flows_in","flows_out","first_seen","last_seen"]

        # Écrire l'en-tête
        ws.append([c for c in cols])

        # Écrire les données
        for r in rows:
            ws.append([r.get(c, "") for c in cols])

        # Mise en forme
        # a) freeze header + filtre + header bold
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # b) Bordure fine partout + alignement des nombres à droite
        thin = Side(style="thin", color="666666")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Colonnes numériques
        num_cols = {"port", "flows_total", "flows_in", "flows_out"}
        header_index = {c: idx+1 for idx, c in enumerate(cols)}  # 1-based

        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(1, max_row+1):
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_thin
                # nombre -> format '0' + alignement à droite
                if r >= 2 and ws.cell(row=1, column=c).value in num_cols:
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right")

        # c) Couleur de fond légère sur la colonne A selon 'category'
        fill_map = {
            "ERADICATE": PatternFill("solid", fgColor="00F4CCCC"),  # light red
            "CONTROL":   PatternFill("solid", fgColor="00FCE4D6"),  # light orange/peach
            "ADMIN":     PatternFill("solid", fgColor="00DDEBF7"),  # light blue
            "ALLOWED":   PatternFill("solid", fgColor="00C6EFCE"),  # light green
            "UNKNOWN":   PatternFill("solid", fgColor="00E7E6E6"),  # light gray
        }
        for r in range(2, max_row+1):
            cat = ws.cell(row=r, column=1).value or ""
            fill = fill_map.get(str(cat).upper(), fill_map["UNKNOWN"])
            ws.cell(row=r, column=1).fill = fill

        # d) Ajuster la largeur des colonnes aux contenus
        #    -> largeur approximative: len(max texte)*1.2, avec bornes min/max
        MIN_W, MAX_W = 10, 60
        for c in range(1, max_col+1):
            max_len = 0
            for r in range(1, max_row+1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                s = str(val)
                if len(s) > max_len:
                    max_len = len(s)
            width = max(MIN_W, min(MAX_W, int(max_len * 1.2)))
            ws.column_dimensions[get_column_letter(c)].width = width

        # Enregistrer
        wb.save(str(excel_path))
        return True

    except Exception as e:
        print(f"[WARN] Excel append failed: {e}")
        return False

# ---------- Main ----------
def main() -> int:
    ap = argparse.ArgumentParser("export_services")
    ap.add_argument("--input-raw", required=True)
    ap.add_argument("--derived-dir", required=True)
    ap.add_argument("--conf", default="carto.conf")
    ap.add_argument("--start", required=True)
    ap.add_argument("--end", required=True)
    ap.add_argument("--excel")
    args = ap.parse_args()

    raw = Path(args.input_raw)
    der = Path(args.derived_dir)
    conf = load_conf(Path(args.conf))

    fin = raw / f"flows_in_{args.start}_{args.end}.csv"
    fout = raw / f"flows_out_{args.start}_{args.end}.csv"
    rows_in, fns_in = load_csv(fin)
    rows_out, fns_out = load_csv(fout)
    if not rows_in and not rows_out:
        print(f"[WARN] no flows found: {fin} / {fout}")
        return 0

    specs = {
        "CONTROL":   parse_port_specs(conf.get("PORTS_TO_CONTROL", "")),
        "ERADICATE": parse_port_specs(conf.get("PORTS_TO_ERADICATE", "")),
        "ADMIN":     parse_port_specs(conf.get("PORTS_ADMIN", "")),
        "ALLOWED":   parse_port_specs(conf.get("ALLOWED_PORTS", "")),
    }

    flows_norm: List[Dict[str, str]] = []
    for r in rows_in:
        n = normalize_flow_row(r, fns_in)
        if n: flows_norm.append(n)
    for r in rows_out:
        n = normalize_flow_row(r, fns_out)
        if n: flows_norm.append(n)

    catalog_rows = to_rows_catalog(flows_norm, specs)
    out_csv = der / "services_catalog.csv"
    write_catalog_csv(out_csv, catalog_rows)
    print(f"[INFO] services_catalog written: {out_csv} rows={len(catalog_rows)}")

    if args.excel:
        # Nom de feuille souhaité (convention): "Service Catalogue"
        sheet_name = "Service Catalogue"
        ok = append_to_excel_openpyxl(Path(args.excel), catalog_rows, sheet_name=sheet_name)
        if ok:
            print(f"[INFO] Excel updated: {args.excel} (sheet='{sheet_name}')")
        else:
            print(f"[INFO] Excel not updated due to previous error")

    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

