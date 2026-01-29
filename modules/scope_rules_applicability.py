
# -*- coding: utf-8 -*-
"""
scope_rules_applicability.py — Carto NG (v3.1-min+src_labels+category+bouquet)

STRICT MINIMUM + extension + colonne A "Rule Category" + Bouquets Infra rule :
- Lit uniquement:
  (1) les labels entrés par l’utilisateur (derived/scope.params.txt),
  (2) raw/export_rules.enabled.csv (colonnes ruleset_scope, src/dst_labels, groups & exclusions),
  (3) raw/export_wkld.m.csv (pour lister les labels distincts du scope via app/env),
  (4) raw/export_labelgroup.csv (pour les inclusions de label groups).

Sélection (enabled) si:
  A) Business strict   : ruleset_scope == {app:<val>, env:<val>} (ensemble strict)
  B) Business élargi   : src_labels contient toutes les clés du scope (app & env) avec valeurs égales
  C) Bouquets Infra    : ruleset_scope vide ET (Bloc SOURCE OU Bloc DESTINATION) valide entièrement:
       - Bloc NON-TRIVIAL : contient (au moins) 1 label ET 1 label group
       - MATCH labels              (OR par type, AND entre types présents)
       - NOT MATCH labels_exclusions
       - MATCH label_groups        (AND strict par groupe listé : chaque groupe doit toucher le scope pour son type)
       - NOT MATCH label_groups_exclusions (AND strict : si un groupe d’exclusion touche le scope => bloc invalide)

Feuille "Scope Applicable Rules" (colonne A "Rule Category"):
  - "Bouquets Infra rule"        pour les règles infra (C)
  - "Business Rule in Scope"     si ruleset_scope vide OU contient toutes les clés du scope utilisateur
  - "Business Rule in other Scope" si ruleset_scope non vide ET ne contient pas toutes les clés du scope
"""
import csv
from pathlib import Path
from typing import Dict, List, Tuple, Set

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ----------------------------- CSV utils -------------------------------------
def pick(cols: List[str], *cands: str) -> str:
    low = {c.lower(): c for c in (cols or [])}
    for c in cands:
        if c.lower() in low:
            return low[c.lower()]
    return ''

def iter_csv_rows(path: Path):
    if (not path.exists()) or path.stat().st_size == 0:
        class _Empty:
            fieldnames_list: List[str] = []
            def __iter__(self):
                if False:
                    yield {}
        return _Empty()
    fh = path.open(newline='', encoding='utf-8-sig')
    sample = fh.read(4096)
    fh.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(fh, dialect=dialect)
    class _Wrap:
        fieldnames_list = reader.fieldnames or []
        def __iter__(self):
            try:
                for row in reader:
                    yield {k: (v or '').strip() for k, v in row.items()}
            finally:
                try: fh.close()
                except Exception: pass
    return _Wrap()

# ----------------------- selected scope (derived) -----------------------------
def _load_selected_scope(derived_dir: Path) -> Dict[str, Set[str]]:
    out: Dict[str, Set[str]] = {}
    p = derived_dir / "scope.params.txt"
    if not p.exists() or p.stat().st_size == 0:
        return out
    try:
        with p.open(encoding='utf-8') as f:
            for line in f:
                s = (line or '').strip()
                if not s or s.startswith('#') or '=' not in s:
                    continue
                k, v = s.split('=', 1)
                k = (k or '').strip().lower()
                v = (v or '').strip()
                if k == 'os': k = 'OS'
                if k and v:
                    out.setdefault(k, set()).add(v)
    except Exception:
        pass
    return out

# ----------------------- scope/labels parsers ---------------------------------
def parse_kv_tokens(val: str) -> Dict[str, Set[str]]:
    out: Dict[str, Set[str]] = {}
    if not val:
        return out
    tokens = [t.strip() for t in str(val).replace(',', ';').split(';') if t.strip()]
    for t in tokens:
        if '=' in t:
            k, v = t.split('=', 1)
        elif ':' in t:
            k, v = t.split(':', 1)
        else:
            continue
        k = (k or '').strip().lower()
        v = (v or '').strip()
        if not k or not v:
            continue
        if k == 'os': k = 'OS'
        out.setdefault(k, set()).add(v)
    return out

# ----------------------- Label Groups -----------------------------------------
def load_labelgroups(path: Path) -> Dict[Tuple[str, str], Set[str]]:
    gen = iter_csv_rows(path)
    fns = getattr(gen, 'fieldnames_list', [])
    if not fns:
        return {}
    c_name = pick(fns, 'name')
    c_key  = pick(fns, 'key')
    c_members = pick(fns, 'member_labels', 'members', 'labels')
    c_expanded = pick(fns, 'fully_expanded_members', 'expanded_members')
    out: Dict[Tuple[str, str], Set[str]] = {}
    for r in gen:
        name = (r.get(c_name) or '').strip()
        key  = (r.get(c_key) or '').strip().lower()
        vals: List[str] = []
        for fld in [c_members, c_expanded]:
            if fld and r.get(fld):
                vals.extend([x.strip() for x in str(r.get(fld)).split(';') if x.strip()])
        out[(('OS' if key == 'os' else key), name)] = set(vals)
    return out

# ----------------------- Scope workloads label values -------------------------
def gather_scope_label_values_from_wkld_m(raw_dir: Path, selected: Dict[str, Set[str]]) -> Dict[str, Set[str]]:
    path = raw_dir / "export_wkld.m.csv"
    gen = iter_csv_rows(path)
    fns = getattr(gen, 'fieldnames_list', [])
    if not fns:
        return {k: set(v) for k, v in selected.items()}
    c_role = pick(fns, 'role')
    c_app  = pick(fns, 'app')
    c_env  = pick(fns, 'env')
    c_loc  = pick(fns, 'loc')
    c_os   = pick(fns, 'OS', 'os')

    def eq(a: str, vals: Set[str]) -> bool:
        return (a or '').strip() in set(x.strip() for x in (vals or set()))

    out: Dict[str, Set[str]] = {k: set() for k in ['role','app','env','loc','OS']}
    for r in gen:
        ok = True
        for key, col in [('app', c_app), ('env', c_env), ('role', c_role), ('loc', c_loc), ('OS', c_os)]:
            if selected.get(key):
                if not eq((r.get(col) or ''), selected[key]):
                    ok = False
                    break
        if not ok:
            continue
        for k, col in [('role', c_role), ('app', c_app), ('env', c_env), ('loc', c_loc), ('OS', c_os)]:
            v = (r.get(col) or '').strip()
            if v:
                out[k].add(v)
    if (not out['app']) and selected.get('app'):
        out['app'] |= set(selected['app'])
    if (not out['env']) and selected.get('env'):
        out['env'] |= set(selected['env'])
    return out

# ----------------------- load enabled rules -----------------------------------
def load_rules_enabled(raw_dir: Path) -> List[Dict[str, str]]:
    path = raw_dir / 'export_rules.enabled.csv'
    gen = iter_csv_rows(path)
    fns = getattr(gen, 'fieldnames_list', [])
    if not fns:
        return []

    c = lambda *x: pick(fns, *x)
    cols = {
        'rule_enabled': c('rule_enabled'),
        'rule_href': c('rule_href'),
        'rule_description': c('rule_description'),
        'ruleset_name': c('ruleset_name'),
        'ruleset_scope': c('ruleset_scope','scope'),
        'ruleset_enabled': c('ruleset_enabled'),
        'rule_type': c('rule_type'),
        'services': c('services'),
        'src_all_workloads': c('src_all_workloads', 'consumer_all_workloads'),
        'dst_all_workloads': c('dst_all_workloads', 'provider_all_workloads'),
        'src_labels': c('src_labels', 'consumer_labels'),
        'dst_labels': c('dst_labels', 'provider_labels'),
        'src_label_groups': c('src_label_groups', 'consumer_label_groups'),
        'dst_label_groups': c('dst_label_groups', 'provider_label_groups'),
        'src_label_groups_exclusions': c('src_label_groups_exclusions', 'consumer_label_groups_exclusions'),
        'dst_label_groups_exclusions': c('dst_label_groups_exclusions', 'provider_label_groups_exclusions'),
        'src_labels_exclusions': c('src_labels_exclusions', 'consumer_labels_exclusions'),
        'dst_labels_exclusions': c('dst_labels_exclusions', 'provider_labels_exclusions'),
        'src_iplists': c('src_iplists', 'consumer_iplists'),
        'dst_iplists': c('dst_iplists', 'provider_iplists'),
    }

    out: List[Dict[str, str]] = []
    for r in gen:
        truth = (r.get(cols['rule_enabled']) or '').strip().lower() in ('true', '1', 'yes', 'y')
        if not truth:
            continue
        out.append({
            'href': (r.get(cols['rule_href']) or '').strip(),
            'display': (r.get(cols['rule_description']) or '').strip(),
            'ruleset_name': (r.get(cols['ruleset_name']) or '').strip(),
            'ruleset_scope': (r.get(cols['ruleset_scope']) or '').strip(),
            'ruleset_enabled': (r.get(cols['ruleset_enabled']) or '').strip(),
            'rule_type': (r.get(cols['rule_type']) or '').strip(),
            'services_raw': (r.get(cols['services']) or '').strip(),
            'src_all_raw': (r.get(cols['src_all_workloads']) or '').strip(),
            'dst_all_raw': (r.get(cols['dst_all_workloads']) or '').strip(),
            'src_labels_raw': (r.get(cols['src_labels']) or '').strip(),
            'dst_labels_raw': (r.get(cols['dst_labels']) or '').strip(),
            'src_groups_raw': (r.get(cols['src_label_groups']) or '').strip(),
            'dst_groups_raw': (r.get(cols['dst_label_groups']) or '').strip(),
            'src_groups_excl_raw': (r.get(cols['src_label_groups_exclusions']) or '').strip(),
            'dst_groups_excl_raw': (r.get(cols['dst_label_groups_exclusions']) or '').strip(),
            'src_labels_excl_raw': (r.get(cols['src_labels_exclusions']) or '').strip(),
            'dst_labels_excl_raw': (r.get(cols['dst_labels_exclusions']) or '').strip(),
            'src_iplists_raw': (r.get(cols['src_iplists']) or '').strip(),
            'dst_iplists_raw': (r.get(cols['dst_iplists']) or '').strip(),
        })
    return out

# ----------------------- matching helpers -------------------------------------
def _rs_strict_match(rs_map: Dict[str, Set[str]], target_keys: Set[str], target_map: Dict[str, Set[str]]) -> bool:
    return (set(rs_map.keys()) == target_keys) and all((rs_map.get(k) == target_map.get(k)) for k in target_keys)

def _rs_contains_all_user_labels(rs_map: Dict[str, Set[str]], target_keys: Set[str], target_map: Dict[str, Set[str]]) -> bool:
    if not target_keys:
        return False
    for k in target_keys:
        have = rs_map.get(k) or set()
        want = next(iter(target_map.get(k) or []), None)
        if want not in have:
            return False
    return True

def _src_labels_contains_scope(src_map: Dict[str, Set[str]], target_keys: Set[str], target_map: Dict[str, Set[str]]) -> bool:
    if not target_keys:
        return False
    for k in target_keys:
        have = src_map.get(k)
        want = next(iter(target_map.get(k) or []), None)
        if not have or (want not in have):
            return False
    return True

_KEYS = ['role','app','env','loc','OS']

def _labels_match(block_map: Dict[str, Set[str]], scope_vals: Dict[str, Set[str]]) -> bool:
    for key in _KEYS:
        vals = block_map.get(key)
        if vals:
            if not (scope_vals.get(key) and (scope_vals[key] & vals)):
                return False
    return True

def _labels_excl_not_match(excl_map: Dict[str, Set[str]], scope_vals: Dict[str, Set[str]]) -> bool:
    for key in _KEYS:
        vals = excl_map.get(key)
        if vals and scope_vals.get(key) and (scope_vals[key] & vals):
            return False
    return True

def _group_touches_scope(group_name: str,
                         scope_vals: Dict[str, Set[str]],
                         lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    """
    Un groupe "touche" le scope si, pour au moins un type (normalement 1),
    l'intersection (members(group), scope_vals[type]) est non vide.
    """
    touched = False
    for key in _KEYS:
        members = lg_map.get((key, group_name)) or set()
        if members:
            if scope_vals.get(key) and (scope_vals[key] & members):
                touched = True
            else:
                # groupe résolu pour ce type mais ne touche pas => OK seulement si un autre type touche
                pass
    return touched

def _groups_match_AND_per_group(names_str: str,
                                scope_vals: Dict[str, Set[str]],
                                lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    names = [x.strip() for x in (names_str or '').split(';') if x.strip()]
    if not names:
        return False  # NON-TRIVIAL : au moins un groupe requis
    for nm in names:
        # Doit toucher le scope pour son/ses type(s) ; sinon échec
        if not _group_touches_scope(nm, scope_vals, lg_map):
            return False
    return True

def _groups_excl_not_match_AND_per_group(names_str: str,
                                         scope_vals: Dict[str, Set[str]],
                                         lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    names = [x.strip() for x in (names_str or '').split(';') if x.strip()]
    if not names:
        return True
    for nm in names:
        # SI un groupe d'exclusion touche le scope (quelque soit le type) => bloc invalide
        for key in _KEYS:
            members = lg_map.get((key, nm)) or set()
            if members and scope_vals.get(key) and (scope_vals[key] & members):
                return False
    return True

def _bouquet_block_valid(labels_raw: str,
                         labels_excl_raw: str,
                         groups_raw: str,
                         groups_excl_raw: str,
                         scope_vals: Dict[str, Set[str]],
                         lg_map: Dict[Tuple[str, str], Set[str]]) -> bool:
    labels_map      = parse_kv_tokens(labels_raw)
    labels_excl_map = parse_kv_tokens(labels_excl_raw)

    # NON-TRIVIALITÉ : le bloc doit déclarer au moins un label ET au moins un groupe
    has_labels_constraints = any(labels_map.get(k) for k in _KEYS)
    has_groups_constraints = bool([x for x in (groups_raw or '').split(';') if x.strip()])
    if not (has_labels_constraints and has_groups_constraints):
        return False

    # Labels + exclusions
    if not _labels_match(labels_map, scope_vals):
        return False
    if not _labels_excl_not_match(labels_excl_map, scope_vals):
        return False

    # Groups + exclusions (AND strict par groupe listé)
    if not _groups_match_AND_per_group(groups_raw, scope_vals, lg_map):
        return False
    if not _groups_excl_not_match_AND_per_group(groups_excl_raw, scope_vals, lg_map):
        return False

    return True

# ----------------------- API: applicable rules (Business + Bouquets) ----------
def get_applicable_rules(raw_dir: Path, derived_dir: Path) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    selected = _load_selected_scope(derived_dir)

    app_vals = list(selected.get('app', []))
    env_vals = list(selected.get('env', []))
    app = app_vals[0] if app_vals else ''
    env = env_vals[0] if env_vals else ''
    target_keys: Set[str] = set()
    target_map: Dict[str, Set[str]] = {}
    if app:
        target_keys.add('app');  target_map['app'] = {app}
    if env:
        target_keys.add('env');  target_map['env'] = {env}

    scope_vals = gather_scope_label_values_from_wkld_m(raw_dir, selected)
    lg_map = load_labelgroups(raw_dir / 'export_labelgroup.csv')

    rules = load_rules_enabled(raw_dir)

    applicable: List[Dict[str, str]] = []
    by_ruleset: Dict[str, Dict[str, str]] = {}
    unmatched_rows: List[Dict[str, str]] = []

    for rr in rules:
        rs_map  = parse_kv_tokens(rr.get('ruleset_scope', ''))
        src_map = parse_kv_tokens(rr.get('src_labels_raw', ''))

        rule_category = None

        # Bouquets Infra rule (prioritaire)
        if len(rs_map) == 0:
            src_ok = _bouquet_block_valid(
                rr.get('src_labels_raw',''),
                rr.get('src_labels_excl_raw',''),
                rr.get('src_groups_raw',''),
                rr.get('src_groups_excl_raw',''),
                scope_vals, lg_map
            )
            dst_ok = _bouquet_block_valid(
                rr.get('dst_labels_raw',''),
                rr.get('dst_labels_excl_raw',''),
                rr.get('dst_groups_raw',''),
                rr.get('dst_groups_excl_raw',''),
                scope_vals, lg_map
            )
            if src_ok or dst_ok:
                rule_category = 'Bouquets Infra rule'

        # Business rules (no regression)
        if rule_category is None:
            is_applicable_business = _rs_strict_match(rs_map, target_keys, target_map) or _src_labels_contains_scope(src_map, target_keys, target_map)
            if is_applicable_business:
                if len(rs_map) == 0 or _rs_contains_all_user_labels(rs_map, target_keys, target_map):
                    rule_category = 'Business Rule in Scope'
                else:
                    rule_category = 'Business Rule in other Scope'

        if rule_category is not None:
            row = dict(rr)
            row['rule_category'] = rule_category
            applicable.append(row)

            rsn = rr.get('ruleset_name', '')
            e = by_ruleset.setdefault(rsn, {
                'ruleset_name': rsn,
                'ruleset_scope': rr.get('ruleset_scope', ''),
                'ruleset_enabled': rr.get('ruleset_enabled', ''),
                'applies_to_scope': 'Y',
                'covered_flows_count': 0,
                'rules_enabled_count': 0,
                'rules_applicable_count': 0,
                'rules_with_hits_count': 0,
                'top_services': '',
            })
            e['rules_enabled_count'] = int(e.get('rules_enabled_count') or 0) + 1
            e['rules_applicable_count'] = int(e.get('rules_applicable_count') or 0) + 1
        else:
            unmatched_rows.append({
                'ruleset_name': rr.get('ruleset_name',''),
                'rule_href': rr.get('href',''),
                'rule_type': rr.get('rule_type',''),
                'services': rr.get('services_raw',''),
                'unmatched_reason': 'no_match',
            })

    eff_rows = list(by_ruleset.values())
    eff_rows.sort(key=lambda r: (r.get('ruleset_name') or ''))

    return applicable, unmatched_rows, eff_rows

# ----------------------- Excel: Scope Applicable Rules ------------------------
def _auto_width(ws, max_w=120, min_w=10):
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column; max_row = ws.max_row
    for c in range(1, max_col+1):
        m = 0
        for r in range(1, max_row+1):
            v = ws.cell(row=r, column=c).value
            s = '' if v is None else str(v)
            if len(s) > m: m = len(s)
        ws.column_dimensions[get_column_letter(c)].width = max(min_w, min(max_w, int(m*1.1)))

def append_scope_rules_sheet(excel_path: Path,
                             rules_applicables: List[Dict[str, str]],
                             eff_rows: List[Dict[str, str]]) -> bool:
    try:
        if excel_path.exists():
            wb = load_workbook(filename=str(excel_path))
        else:
            wb = Workbook()
        if wb.active and wb.active.title == 'Sheet':
            wb.remove(wb.active)

        if 'Scope Applicable Rules' in wb.sheetnames:
            wb.remove(wb['Scope Applicable Rules'])
        ws = wb.create_sheet(title='Scope Applicable Rules')
        cols = [
            'Rule Category',
            'ruleset_name','ruleset_scope','ruleset_enabled','rule_type','services',
            'src_all_workloads','src_labels','src_label_groups','src_iplists',
            'dst_all_workloads','dst_labels','dst_label_groups','dst_iplists'
        ]
        ws.append(cols)
        for r in rules_applicables:
            ws.append([
                r.get('rule_category',''),
                r.get('ruleset_name',''), r.get('ruleset_scope',''), r.get('ruleset_enabled',''),
                r.get('rule_type',''), r.get('services_raw',''),
                r.get('src_all_raw',''), r.get('src_labels_raw',''), r.get('src_groups_raw',''), r.get('src_iplists_raw',''),
                r.get('dst_all_raw',''), r.get('dst_labels_raw',''), r.get('dst_groups_raw',''), r.get('dst_iplists_raw',''),
            ])
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions
        F_BOLD = Font(bold=True); THIN = Side(style='thin', color='666666')
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        FILL_HDR = PatternFill('solid', fgColor='D9E1F2')
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=1, column=c)
            cell.font = F_BOLD; cell.fill = FILL_HDR; cell.border = BORDER
        _auto_width(ws)

        if 'Ruleset Effectiveness' in wb.sheetnames:
            wb.remove(wb['Ruleset Effectiveness'])
        ws2 = wb.create_sheet(title='Ruleset Effectiveness')
        cols2 = ['ruleset_name','ruleset_scope','ruleset_enabled','applies_to_scope',
                 'covered_flows_count','rules_enabled_count','rules_applicable_count','rules_with_hits_count','top_services']
        ws2.append(cols2)
        for e in eff_rows:
            ws2.append([e.get(c, '') for c in cols2])
        ws2.freeze_panes = 'A2'; ws2.auto_filter.ref = ws.dimensions
        for c in range(1, ws2.max_column+1):
            cell = ws2.cell(row=1, column=c)
            cell.font = F_BOLD; cell.fill = FILL_HDR; cell.border = BORDER
        _auto_width(ws2, max_w=80)

        wb.save(str(excel_path))
        return True
    except Exception:
        return False
