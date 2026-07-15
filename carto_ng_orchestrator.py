
# -*- coding: utf-8 -*-
"""
Carto NG Orchestrator — SCOPE LabelGroup v3 (DEBUG Instrumented + ENRICHED + FIX + BULK PATCH)
- Robust dupecheck schema detection (raw/exploded/final-like)
- HREF normalization (/orgs/<id>/workloads/<uuid>)
- Scope fallback (when wkld.m scope empty and dupecheck is final-like)
- DEBUG flags: --debug-echo-dupecheck, --debug-no-scope-filter
- LEGACY ENRICHMENT: add managed-only rows (Not Active in CMDB) and unmanaged-only rows (Missing Agent)
  by scanning export_wkld.m (scope) and export_wkld.u (quadruplets role/app/env/loc)
- Summary: Section 3 "Workloads status"
- Processes: group coloring by hostname (two blues)
- Flows: column-family coloring & numeric formatting for Port & Num Flows
PATCHES (2025-12-11):
- FIX (critical): process-export-bulk now receives <href_file> and <output_csv>.
- NEW: Generate derived/list_href.managed.csv from filtered managed workloads.
- FALLBACK: If bulk fails, attempt process-export-single via workloader_process.sh; else keep empty header (no regression).
- HARDEN: --dev-stub-dupecheck-filtered copy is now protected (try/except) — no crash if source stub missing.
- NEW: --network-zone hook: filter flows to intra-zone (East-West) based on raw/export_iplists.csv
- NEW: --CreateRules / --RecertifyRules options (modules called after Excel is produced to append sheets)

PATCH v2.2 (Flow→Rule Hits integration):
- NEW: --FlowRuleHits (bool) pour déclencher modules/flows_to_rules.py v2.2.0
- NEW: Flags de contrôle et filtrage transmis à flows_to_rules:
  --frh-filter-direction, --frh-filter-proto, --frh-filter-port,
  --frh-ruleset-name-contains, --frh-exclude-all-workloads-rules,
  --frh-prefer-raw, --frh-limit-flows,
  --frh-debug, --frh-debug-matches-only, --frh-debug-max-rows, --frh-debug-sample-rate, --frh-log-level
- Par défaut si --frh-ruleset-name-contains est absent, on utilise "app;env;role" du scope courant.
"""

import os
import sys
import csv
import argparse
import subprocess
import time
import re
import socket
import ipaddress
import logging
import atexit
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Set
import pandas as pd
from modules.email_utils import (
    build_execution_summary_table,
    parse_recipients,
    send_carto_notification,
)

# ------------------------------ config/env ------------------------------
def load_conf(path: str) -> Dict[str, str]:
    conf: Dict[str, str] = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if "=" in s:
                k, v = s.split("=", 1)
                conf[k.strip()] = v.strip()
    return conf

def load_env_file(path: Path) -> Dict[str, str]:
    env: Dict[str, str] = {}
    if not path.exists():
        return env
    with path.open(encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if "=" in s:
                k, v = s.split("=", 1)
                env[k.strip()] = v.strip()
    return env

# ------------------------------ utils ------------------------------
def ensure_dir(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)

def now_stamp(fmt: str) -> str:
    return datetime.now().strftime(fmt)

def start_stdout_tee(log_path: Path) -> Tuple[int, int, subprocess.Popen]:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    tee_proc = subprocess.Popen(["tee", "-a", str(log_path)], stdin=subprocess.PIPE)
    if tee_proc.stdin is None:
        raise RuntimeError("Failed to initialize stdout tee process")
    orig_stdout_fd = os.dup(sys.stdout.fileno())
    orig_stderr_fd = os.dup(sys.stderr.fileno())
    os.dup2(tee_proc.stdin.fileno(), sys.stdout.fileno())
    os.dup2(tee_proc.stdin.fileno(), sys.stderr.fileno())
    return orig_stdout_fd, orig_stderr_fd, tee_proc

def stop_stdout_tee(orig_stdout_fd: int, orig_stderr_fd: int, tee_proc: subprocess.Popen) -> None:
    try:
        sys.stdout.flush()
        sys.stderr.flush()
    except Exception:
        pass
    os.dup2(orig_stdout_fd, sys.stdout.fileno())
    os.dup2(orig_stderr_fd, sys.stderr.fileno())
    os.close(orig_stdout_fd)
    os.close(orig_stderr_fd)
    if tee_proc.stdin:
        tee_proc.stdin.close()
    tee_proc.wait()

def finalize_email_notification() -> None:
    tee_state = MAIL_CONTEXT.get("tee_state")
    if tee_state:
        try:
            stop_stdout_tee(*tee_state)
        except Exception as exc:
            print(f"[WARN] Failed to stop stdout tee: {exc}")
        MAIL_CONTEXT["tee_state"] = None

    if not MAIL_CONTEXT.get("enabled"):
        return

    recipients = MAIL_CONTEXT.get("recipients") or []
    conf = MAIL_CONTEXT.get("conf") or {}
    log_path = MAIL_CONTEXT.get("log_path")
    if not log_path:
        return

    summary_text, summary_html = build_execution_summary_table(DUR)
    status = MAIL_CONTEXT.get("status") or "FAIL"
    app_name = MAIL_CONTEXT.get("app") or "N/A"
    env_name = MAIL_CONTEXT.get("env") or "N/A"
    excel_path = MAIL_CONTEXT.get("excel_path") or "N/A"

    subject = f"[CARTO][{status}] app={app_name} env={env_name}"
    body_text = (
        f"Result Excel: {excel_path}\n\n"
        "Execution summary:\n"
        f"{summary_text}\n"
    )
    body_html = (
        f"<p><strong>Result Excel:</strong> {excel_path}</p>"
        "<h3>Execution summary</h3>"
        f"{summary_html}"
    )
    try:
        send_carto_notification(
            conf=conf,
            recipients=recipients,
            subject=subject,
            body_text=body_text,
            body_html=body_html,
            attachment_path=Path(log_path),
            logger=logger,
        )
    except Exception as exc:
        print(f"[WARN] Notification email failed: {exc}")

def rename_final_excel(xlsx_path: Path, app: str, envl: str) -> Path:
    ts = now_stamp('%Y%m%d-%H%M%S')
    name_parts = [sanitize_token(p) for p in (app, envl) if p]
    scope = "-".join(name_parts) if name_parts else "scope"
    target_name = f"carto_{scope}_{ts}.xlsx"
    target_path = xlsx_path.with_name(target_name)
    if target_path.exists():
        idx = 1
        while True:
            candidate = xlsx_path.with_name(f"carto_{scope}_{ts}_{idx}.xlsx")
            if not candidate.exists():
                target_path = candidate
                break
            idx += 1
    xlsx_path.rename(target_path)
    return target_path

def sanitize_token(s: str) -> str:
    allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-"
    return "".join(ch if ch in allowed else "_" for ch in (s or ""))

def compute_window_days(days: int) -> Tuple[str, str]:
    today = datetime.now().date()
    start = (today - timedelta(days=days)).strftime("%Y-%m-%d")
    end = today.strftime("%Y-%m-%d")
    return start, end

# ------------------------------ CSV ------------------------------
def load_csv(path: Path) -> Tuple[List[Dict[str, str]], List[str]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open(newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        rows = [{k: (v or "").strip() for k, v in row.items()} for row in r]
        return rows, (r.fieldnames or [])

def write_list_semicolon(path: Path, items: List[str]) -> None:
    ensure_dir(path)
    with path.open("w", encoding="utf-8", newline="") as f:
        f.write(";".join(items) + "\n")

def write_list_newline(path: Path, items: List[str]) -> None:
    ensure_dir(path)
    with path.open("w", encoding="utf-8", newline="") as f:
        for it in items:
            it = (it or "").strip()
            if it:
                f.write(it + "\n")

# ------------------------------ durations ------------------------------
DUR: Dict[str, float] = {}
logger = logging.getLogger('carto_ng_orchestrator')
MAIL_CONTEXT: Dict[str, object] = {"enabled": False}

def run_step(name: str, cmd: List[str], env: Dict[str, str], cwd: Path) -> bool:
    t0 = time.perf_counter()
    rc = subprocess.run(cmd, env=env, cwd=str(cwd))
    dt = time.perf_counter() - t0
    DUR[name] = DUR.get(name, 0.0) + dt
    print(f"{name}: rc={rc.returncode} duration={dt:.1f}s")
    return rc.returncode == 0

# ------------------------------ label/dupe helpers ------------------------------
LABEL_KEYS = ["app", "env", "loc", "role", "OS"]
HREF_RE = re.compile(r'(?:https?://[^\s;\)]+)?(/orgs/\d+/workloads/[A-Fa-f0-9\-]+)', re.IGNORECASE)

# Normalize href to '/orgs/<id>/workloads/<uuid>'
def norm_href(h: str) -> str:
    h = (h or '').strip()
    m = re.search(r'/orgs/\d+/workloads/[A-Fa-f0-9\-]+', h)
    return m.group(0) if m else h

def read_labels_csv(path: Path) -> List[Dict[str, str]]:
    rows, fn = load_csv(path)
    need = {"href", "key", "value"}
    if rows and need.issubset(set(fn or [])):
        return rows
    missing = need - set(fn or [])
    raise SystemExit(f"ERROR: {path} missing columns {missing}")

def find_label_hrefs(rows: List[Dict[str, str]], filters: Dict[str, str]) -> List[str]:
    hits: Dict[str, List[str]] = {}
    for k, v in filters.items():
        if not v:
            continue
        s: List[str] = []
        role_spec = parse_role_filter(v) if k.lower() == "role" else None
        for row in rows:
            if (row.get("key", "").strip().lower()) != k.lower():
                continue
            row_value = row.get("value", "").strip()
            if role_spec is not None:
                matched = role_value_matches(row_value, role_spec)
            else:
                matched = row_value.lower() == v.strip().lower()
            if matched:
                h = (row.get("href") or "").strip()
                if h:
                    s.append(norm_href(h))
        hits[k] = sorted(set(s))
    for k, s in hits.items():
        if not s:
            raise SystemExit(f"No label href found for {k}={filters.get(k)}")
    out = set()
    for s in hits.values():
        out |= set(s)
    return sorted(out)

def normalize_arg_value(value: object) -> str:
    if isinstance(value, list):
        return " ".join(str(v) for v in value)
    return str(value or "")

def normalize_csv_arg_value(value: object) -> str:
    if isinstance(value, list):
        return ",".join(str(v).strip().rstrip(",") for v in value if str(v).strip())
    return str(value or "")

def parse_csv_tokens(value: str) -> List[str]:
    return [part.strip() for part in (value or "").split(",") if part.strip()]

def parse_role_filter(value: str) -> Dict[str, object]:
    raw = (value or "").strip()
    if not raw:
        return {"mode": "include", "values": []}
    exclude = raw.startswith("!") or raw.lower().startswith("not:")
    token_source = raw[1:] if raw.startswith("!") else raw[4:] if raw.lower().startswith("not:") else raw
    values = parse_csv_tokens(token_source)
    if not values:
        raise SystemExit("--role cannot be empty when using exclusion. Example: --exclude-role FRONTEND")
    if any(tok.startswith("!") for tok in values):
        raise SystemExit("--role does not support mixing include and exclude tokens. Use either A,B or --exclude-role A,B.")
    return {"mode": "exclude" if exclude else "include", "values": values}

def build_role_filter_arg(role_value: object, exclude_role_value: object) -> str:
    role = normalize_arg_value(role_value)
    exclude_role = normalize_csv_arg_value(exclude_role_value)
    if role and exclude_role:
        raise SystemExit("Choose only one role exclusion syntax: --role or --exclude-role")
    if exclude_role:
        return f"!{exclude_role}"
    return role

def role_value_matches(value: str, role_spec: Dict[str, object]) -> bool:
    val = (value or "").strip().lower()
    values = {str(v).strip().lower() for v in role_spec.get("values", []) if str(v).strip()}
    if role_spec.get("mode") == "exclude":
        return bool(val) and val not in values
    return val in values

def filter_value_matches(key: str, value: str, expected: str) -> bool:
    if not expected:
        return True
    if key.lower() == "role":
        return role_value_matches(value, parse_role_filter(expected))
    return (value or "").strip().lower() == (expected or "").strip().lower()

def pick(cols: List[str], *cands: str) -> str:
    low = {c.strip().lower(): c for c in cols if c is not None}
    for c in cands:
        key = c.strip().lower()
        if key in low:
            return low[key]
    return ""

def is_truthy(v: str) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes", "y")

_SPLIT_IPLIST_RE = re.compile(r"[;\t,\|]+")
_SPLIT_INTERFACES_RE = re.compile(r"[;\n]+")

def _extract_networks_from_include(include_value: str) -> List[ipaddress._BaseNetwork]:
    s = (include_value or "").strip()
    if not s:
        return []

    nets: List[ipaddress._BaseNetwork] = []
    parts: List[str] = []
    for p in _SPLIT_IPLIST_RE.split(s):
        p = p.strip()
        if not p:
            continue
        for q in p.split():
            q = q.strip()
            if q:
                parts.append(q)

    for tok in parts:
        t = tok.strip()
        if not t:
            continue
        if t.startswith("!"):
            continue
        if "#" in t:
            t = t.split("#", 1)[0].strip()
        if t.startswith("#"):
            continue
        try:
            if "/" in t:
                nets.append(ipaddress.ip_network(t, strict=False))
            else:
                ip = ipaddress.ip_address(t)
                suffix = "/32" if ip.version == 4 else "/128"
                nets.append(ipaddress.ip_network(f"{t}{suffix}", strict=False))
        except Exception:
            continue

    uniq: Dict[Tuple[int, int, str], ipaddress._BaseNetwork] = {}
    for n in nets:
        key = (n.version, n.prefixlen, str(n))
        uniq[key] = n
    out = list(uniq.values())
    out.sort(key=lambda n: (n.version, -n.prefixlen, str(n)))
    return out

def _load_zone_networks(raw_dir: Path, zone_name: str) -> Tuple[List[ipaddress._BaseNetwork], Optional[str]]:
    ipl_csv = raw_dir / "export_iplists.csv"
    if not ipl_csv.exists():
        return [], f"{ipl_csv} not found (required for --network-zone)"

    with ipl_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        if not cols:
            return [], f"{ipl_csv} has no header (required for --network-zone)"

        c_name = pick(cols, "name", "iplist_name")
        c_inc = pick(cols, "include", "includes", "ip_ranges", "cidrs")
        if not (c_name and c_inc):
            return [], f"{ipl_csv} missing required columns for --network-zone (need name/include)"

        zone_key = (zone_name or "").strip().lower()
        zone_row: Optional[Dict[str, str]] = None

        for r in reader:
            n = (r.get(c_name, "") or "").strip().lower()
            if n == zone_key:
                zone_row = r
                break

        if zone_row is None:
            return [], f"IPList '{zone_name}' not found in {ipl_csv.name} (column '{c_name}')"

        nets = _extract_networks_from_include(zone_row.get(c_inc, "") or "")
        if not nets:
            sample = (zone_row.get(c_inc, "") or "")[:120]
            return [], (
                f"IPList '{zone_name}' has no parsable CIDRs in column '{c_inc}'. "
                f"Example include='{sample}...'"
            )

        return nets, None

def _extract_interface_ips(value: str) -> List[str]:
    if not value:
        return []
    out: List[str] = []
    for raw in _SPLIT_INTERFACES_RE.split(value):
        tok = raw.strip()
        if not tok:
            continue
        ip_part = tok.split(":", 1)[1].strip() if ":" in tok else tok
        if not ip_part:
            continue
        try:
            if "/" in ip_part:
                ip = ipaddress.ip_interface(ip_part).ip
            else:
                ip = ipaddress.ip_address(ip_part)
            out.append(str(ip))
        except Exception:
            continue
    return out

def _ip_in_zone(ip_s: str,
                nets_v4: List[ipaddress.IPv4Network],
                nets_v6: List[ipaddress.IPv6Network]) -> bool:
    if not ip_s:
        return False
    try:
        ip = ipaddress.ip_address(ip_s)
    except Exception:
        return False
    if ip.version == 4:
        return any(ip in n for n in nets_v4)
    return any(ip in n for n in nets_v6)

def explode_reason_to_matches(dup_rows: List[Dict[str, str]], dup_fn: List[str]) -> List[Dict[str, str]]:
    if "reason" not in (dup_fn or []):
        return []
    col_href_un = pick(dup_fn, "href")
    out: List[Dict[str, str]] = []
    seen = set()
    for r in dup_rows:
        reason = (r.get("reason") or "").strip()
        if not reason:
            continue
        frags = [f.strip() for f in reason.split(";") if f.strip()]
        per: Dict[str, str] = {}
        for f in frags:
            hrefs = HREF_RE.findall(f)
            if not hrefs:
                continue
            mtype = "one-interface-match" if "one-interface-match" in f.lower() else "all interfaces"
            for h in hrefs:
                h = norm_href(h)
                prev = per.get(h)
                if prev is None:
                    per[h] = mtype
                else:
                    if "one-interface-match" in (prev, mtype):
                        per[h] = "one-interface-match"
        for href_m, mtype in per.items():
            base = dict(r)
            base["href_managed"] = href_m
            base["match_type"] = mtype
            key = (base.get(col_href_un, ""), href_m, mtype)
            if key not in seen:
                out.append(base)
                seen.add(key)
    return out

def find_managed_wkld_hrefs_for_filters(wkld_m_path: Path, filters: Dict[str, str]) -> List[str]:
    rows, _ = load_csv(wkld_m_path)
    def eq(a: str, b: str) -> bool:
        return (a or "").strip().lower() == (b or "").strip().lower()
    hrefs: List[str] = []
    _, fn = load_csv(wkld_m_path)
    c_href = pick(fn, "href")
    c_app = pick(fn, "app")
    c_env = pick(fn, "env")
    c_role = pick(fn, "role")
    c_loc = pick(fn, "loc")
    c_os = pick(fn, "OS", "os")
    for r in rows:
        ok = True
        for key, col in [("app", c_app), ("env", c_env), ("role", c_role), ("loc", c_loc), ("OS", c_os)]:
            v = filters.get(key)
            if v:
                rv = (r.get(col) or r.get(col.lower()) or "").strip()
                if key == "role":
                    if not filter_value_matches(key, rv, v):
                        ok = False
                        break
                elif not eq(rv, v):
                    ok = False
                    break
        if ok:
            h = norm_href(r.get(c_href) or "")
            if h:
                hrefs.append(h)
    return sorted(set(hrefs))

# JOIN WITH MANAGED (uses normalization)
def join_with_managed(df_exp_rows: List[Dict[str, str]],
                      df_m_rows: List[Dict[str, str]],
                      df_m_fn: List[str]) -> List[Dict[str, str]]:
    col_href_m = pick(df_m_fn, "href")
    col_hn_m = pick(df_m_fn, "hostname", "host_name", "name")
    col_role_m = pick(df_m_fn, "role")
    col_app_m = pick(df_m_fn, "app")
    col_env_m = pick(df_m_fn, "env")
    col_loc_m = pick(df_m_fn, "loc")
    col_os_m = pick(df_m_fn, "OS", "os")
    col_if_m = pick(df_m_fn, "interfaces", "network_interfaces", "interfaces_ips", "interfaces_ip", "ips")
    col_man_m = pick(df_m_fn, "managed", "is_managed", "managed_state")
    col_enf_m = pick(df_m_fn, "enforcement", "enforcement_mode", "mode", "workload_mode")
    idx: Dict[str, Dict[str, str]] = {}
    for m in df_m_rows:
        key = norm_href(m.get(col_href_m) or "")
        if not key:
            continue
        idx[key] = {
            "hostname_managed": (m.get(col_hn_m) or "").strip(),
            "role_managed": (m.get(col_role_m)or "").strip(),
            "app_managed": (m.get(col_app_m) or "").strip(),
            "env_managed": (m.get(col_env_m) or "").strip(),
            "loc_managed": (m.get(col_loc_m) or "").strip(),
            "os_managed": (m.get(col_os_m) or "").strip(),
            "interfaces_managed": (m.get(col_if_m) or "").strip(),
            "managed": (m.get(col_man_m) or "").strip(),
            "enforcement": (m.get(col_enf_m) or "").strip(),
        }
    out: List[Dict[str, str]] = []
    for e in df_exp_rows:
        href_m = norm_href(e.get("href_managed") or "")
        add = idx.get(href_m, {})
        row = dict(e)
        def rn(src, dst):
            if src in row and dst not in row:
                row[dst] = row.pop(src)
        rn("href", "href_unmanaged")
        rn("hostname", "hostname_unmanaged")
        rn("name", "name_unmanaged")
        rn("interfaces", "interfaces_unmanaged")
        rn("role", "role_unmanaged")
        rn("app", "app_unmanaged")
        rn("env", "env_unmanaged")
        rn("loc", "loc_unmanaged")
        rn("OS", "os_unmanaged")
        row.update(add)
        row["href_managed"] = href_m
        out.append(row)
    return out

# ---- LEGACY ENRICHMENT (managed-only and unmanaged-only + Info column)
FINAL_ORDER = [
    'href_managed','hostname_managed','role_managed','app_managed','env_managed','loc_managed','os_managed','interfaces_managed','managed','enforcement',
    'href_unmanaged','hostname_unmanaged','name_unmanaged','interfaces_unmanaged','role_unmanaged','app_unmanaged','env_unmanaged','loc_unmanaged','os_unmanaged',
    'match_type','reason'
]

def _pick_col(cols: List[str], *cands: str) -> str:
    low = {c.lower(): c for c in cols}
    for c in cands:
        if c.lower() in low:
            return low[c.lower()]
    return ''

def enrich_workloads(base_rows: List[Dict[str,str]],
                     filters: Dict[str,str],
                     wkld_m_path: Path,
                     wkld_u_path: Path,
                     debug: bool=False) -> Tuple[List[Dict[str,str]], Dict[str,int]]:
    out_base: List[Dict[str,str]] = []
    hrefs_m_in = set()
    hrefs_u_in = set()
    for r in base_rows:
        rr = {k: r.get(k, '') for k in FINAL_ORDER}
        if not rr.get('href_unmanaged') and r.get('href'):
            rr['href_unmanaged'] = r.get('href','')
        for k in ['hostname','name','interfaces','role','app','env','loc','OS']:
            dst = k + '_unmanaged' if k != 'OS' else 'os_unmanaged'
            if not rr.get(dst) and r.get(k):
                rr[dst] = r.get(k,'')
        rr['href_managed'] = norm_href(rr.get('href_managed',''))
        rr['href_unmanaged'] = norm_href(rr.get('href_unmanaged',''))
        rr['Info'] = rr.get('Info','') or 'Duplicate found'
        out_base.append(rr)
        if rr.get('href_managed'):
            hrefs_m_in.add(rr['href_managed'])
        if rr.get('href_unmanaged'):
            hrefs_u_in.add(rr['href_unmanaged'])

    wk_m_rows, wk_m_fn = load_csv(wkld_m_path)
    c_href_m = _pick_col(wk_m_fn, 'href')
    c_hn_m = _pick_col(wk_m_fn, 'hostname','name','host_name')
    c_role_m = _pick_col(wk_m_fn, 'role')
    c_app_m = _pick_col(wk_m_fn, 'app')
    c_env_m = _pick_col(wk_m_fn, 'env')
    c_loc_m = _pick_col(wk_m_fn, 'loc')
    c_os_m = _pick_col(wk_m_fn, 'OS','os')
    c_if_m = _pick_col(wk_m_fn, 'interfaces','network_interfaces','interfaces_ips','interfaces_ip','ips')
    c_man_m = _pick_col(wk_m_fn, 'managed')
    c_enf_m = _pick_col(wk_m_fn, 'enforcement','enforcement_mode','mode','workload_mode')

    def eq(a: str, b: str) -> bool:
        return (a or '').strip().lower() == (b or '').strip().lower()

    scope_rows: List[Dict[str,str]] = []
    for m in wk_m_rows:
        ok = True
        for key, col in [('app', c_app_m), ('env', c_env_m), ('role', c_role_m), ('loc', c_loc_m), ('OS', c_os_m)]:
            v = filters.get(key)
            if v:
                mv = (m.get(col) or m.get((col or '').lower()) or '').strip()
                if key == 'role':
                    if not filter_value_matches(key, mv, v):
                        ok = False
                        break
                elif not eq(mv, v):
                    ok = False
                    break
        if ok:
            scope_rows.append(m)

    managed_only: List[Dict[str,str]] = []
    for m in scope_rows:
        href = norm_href((m.get(c_href_m) or '').strip())
        if (not href) or (href in hrefs_m_in):
            continue
        out = {k: '' for k in FINAL_ORDER}
        out['href_managed'] = href
        out['hostname_managed'] = (m.get(c_hn_m) or '').strip()
        out['role_managed'] = (m.get(c_role_m)or '').strip()
        out['app_managed'] = (m.get(c_app_m) or '').strip()
        out['env_managed'] = (m.get(c_env_m) or '').strip()
        out['loc_managed'] = (m.get(c_loc_m) or '').strip()
        out['os_managed'] = (m.get(c_os_m) or '').strip()
        out['interfaces_managed'] = (m.get(c_if_m) or '').strip()
        out['managed'] = (m.get(c_man_m) or '').strip()
        out['enforcement'] = (m.get(c_enf_m) or '').strip()
        out['Info'] = 'Not Active in CMDB'
        managed_only.append(out)

    quads: Set[Tuple[str,str,str,str]] = set()
    for r in out_base:
        quad = (
            (r.get('role_unmanaged') or '').strip().lower(),
            (r.get('app_unmanaged') or '').strip().lower(),
            (r.get('env_unmanaged') or '').strip().lower(),
            (r.get('loc_unmanaged') or '').strip().lower(),
        )
        if any(quad):
            quads.add(quad)

    wk_u_rows, wk_u_fn = load_csv(wkld_u_path)
    c_href_u = _pick_col(wk_u_fn, 'href')
    c_hn_u = _pick_col(wk_u_fn, 'hostname','name','host_name')
    c_role_u = _pick_col(wk_u_fn, 'role')
    c_app_u = _pick_col(wk_u_fn, 'app')
    c_env_u = _pick_col(wk_u_fn, 'env')
    c_loc_u = _pick_col(wk_u_fn, 'loc')
    c_os_u = _pick_col(wk_u_fn, 'OS','os')
    c_if_u = _pick_col(wk_u_fn, 'interfaces','network_interfaces','interfaces_ips','interfaces_ip','ips')

    unmanaged_only: List[Dict[str,str]] = []
    for u in wk_u_rows:
        quad_u = (
            (u.get(c_role_u) or '').strip().lower(),
            (u.get(c_app_u) or '').strip().lower(),
            (u.get(c_env_u) or '').strip().lower(),
            (u.get(c_loc_u) or '').strip().lower(),
        )
        if quad_u not in quads:
            continue
        href_u = norm_href((u.get(c_href_u) or '').strip())
        if (not href_u) or (href_u in hrefs_u_in):
            continue
        out = {k: '' for k in FINAL_ORDER}
        out['href_unmanaged'] = href_u
        out['hostname_unmanaged'] = (u.get(c_hn_u) or '').strip()
        out['name_unmanaged'] = (u.get('name') or u.get(c_hn_u) or '').strip()
        out['interfaces_unmanaged'] = (u.get(c_if_u) or '').strip()
        out['role_unmanaged'] = (u.get(c_role_u) or '').strip()
        out['app_unmanaged'] = (u.get(c_app_u) or '').strip()
        out['env_unmanaged'] = (u.get(c_env_u) or '').strip()
        out['loc_unmanaged'] = (u.get(c_loc_u) or '').strip()
        out['os_unmanaged'] = (u.get(c_os_u) or '').strip()
        out['Info'] = 'Missing Agent'
        unmanaged_only.append(out)

    all_rows = out_base + managed_only + unmanaged_only
    counts = {
        'duplicates': len(out_base),
        'managed_only': len(managed_only),
        'unmanaged_only': len(unmanaged_only),
        'managed_scope_total': len(scope_rows),
    }
    if debug:
        print(f"[DEBUG] enrich: base={len(out_base)} managed_only={len(managed_only)} unmanaged_only={len(unmanaged_only)} scope_total={len(scope_rows)}")
    return all_rows, counts

# NEW: Gather distinct scope label values from managed workloads
def gather_scope_label_values(wkld_m_path: Path, hrefs_scope: Set[str]) -> Dict[str, Set[str]]:
    rows, fn = load_csv(wkld_m_path)
    keys = ["app", "role", "env", "loc", "OS"]
    out: Dict[str, Set[str]] = {k: set() for k in keys}
    col_href = pick(fn, "href")
    colmap = {
        'app': pick(fn, 'app'),
        'role': pick(fn, 'role'),
        'env': pick(fn, 'env'),
        'loc': pick(fn, 'loc'),
        'OS': pick(fn, 'OS', 'os'),
    }
    for r in rows:
        h = norm_href(r.get(col_href) or '')
        if h and h in hrefs_scope:
            for k, col in colmap.items():
                v = (r.get(col) or '').strip()
                if v:
                    out[k].add(v)
    return out

# ------------------------------ labelgroups ------------------------------
def read_labelgroups_csv(path: Path) -> List[Dict[str, str]]:
    rows, fn = load_csv(path)
    return rows

def _split_semicol_vals(s: str) -> List[str]:
    return [x.strip() for x in (s or "").split(";") if x.strip()]

def build_labelgroup_inclusions_single(rows: List[Dict[str, str]], key: str, label_val: str) -> List[Dict[str, str]]:
    groups_by_name: Dict[str, Dict[str, str]] = {}
    for r in rows:
        name = (r.get('name') or '').strip()
        if name:
            groups_by_name[name] = r

    parents_map: Dict[str, set] = {}
    for parent_name, r in groups_by_name.items():
        children = _split_semicol_vals(r.get('member_label_groups', '') or r.get('member_label_groups'.lower(), ''))
        for child in children:
            parents_map.setdefault(child, set()).add(parent_name)

    direct_index: Dict[Tuple[str, str], set] = {}
    for name, r in groups_by_name.items():
        k = (r.get('key') or r.get('key'.lower()) or '').strip()
        labels = _split_semicol_vals(r.get('member_labels', '') or r.get('member_labels'.lower(), ''))
        for lab in labels:
            direct_index.setdefault((k, lab), set()).add(name)

    expanded_index: Dict[Tuple[str, str], set] = {}
    for name, r in groups_by_name.items():
        k = (r.get('key') or r.get('key'.lower()) or '').strip()
        expanded = _split_semicol_vals(r.get('fully_expanded_members', '') or r.get('fully_expanded_members'.lower(), ''))
        for lab in expanded:
            expanded_index.setdefault((k, lab), set()).add(name)

    seed_groups = set()
    seed_groups |= direct_index.get((key, label_val), set())
    seed_groups |= expanded_index.get((key, label_val), set())

    out_rows: List[Dict[str, str]] = []
    visited: set = set()
    for g in sorted(seed_groups):
        row = groups_by_name.get(g, {})
        out_rows.append({'name': g, 'href': (row.get('href') or '').strip(), 'level': 'Direct'})
        visited.add(g)

    frontier = [(g, 0) for g in seed_groups]
    while frontier:
        next_frontier = []
        for g, dist in frontier:
            parents = parents_map.get(g, set())
            for p in parents:
                if p in visited:
                    continue
                prow = groups_by_name.get(p, {})
                out_rows.append({'name': p, 'href': (prow.get('href') or '').strip(), 'level': f'Ancestor-{dist+1}'})
                visited.add(p)
                next_frontier.append((p, dist+1))
        frontier = next_frontier
    return out_rows

def build_labelgroup_inclusions_for_scope(rows: List[Dict[str, str]], values_by_key: Dict[str, Set[str]]) -> Dict[Tuple[str, str], List[Dict[str, str]]]:
    result: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for key, values in values_by_key.items():
        for val in sorted(values):
            if not val:
                continue
            result[(key, val)] = build_labelgroup_inclusions_single(rows, key, val)
    return result

# ------------------------------ Excel ------------------------------
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def auto_width(ws, max_width: int = 60, min_width: int = 10, special_caps: Optional[Dict[str, int]] = None) -> None:
    from openpyxl.utils import get_column_letter
    caps = special_caps or {}
    headers = [str(cell.value or "") for cell in ws[1]] if ws.max_row >= 1 else []
    for i, header in enumerate(headers, start=1):
        cap = caps.get(header.lower(), max_width)
        col_letter = get_column_letter(i)
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            v = row[0].value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        width = max(min_width, min(cap, int(max_len * 1.2)))
        ws.column_dimensions[col_letter].width = width


# ------------------------------ unmanaged app label rules sheet ------------------------------
_APP_VAL_RE = re.compile(r"(?i)\bapp\s*[:=]\s*([A-Za-z0-9_.\-]+)")

def extract_app_label_values(label_expr: str) -> Set[str]:
    """Extract app label values from Illumio label expressions (e.g. 'app=FOO;env=PRD')."""
    s = (label_expr or "").strip()
    if not s:
        return set()
    return set(_APP_VAL_RE.findall(s))

def find_rules_with_unmanaged_app_labels(
    *,
    rules_enabled_csv: Path,
    unmanaged_app_values: Set[str],
) -> pd.DataFrame:
    """Scan export_rules.enabled.csv and return rows referencing any unmanaged app label values."""
    out_cols = [
        "ruleset_name",
        "ruleset_scope",
        "rule_type",
        "rule_description",
        "rule_enabled",
        "services",
        "src_labels",
        "src_labels_exclusions",
        "dst_labels",
        "dst_labels_exclusions",
        "matched_umgd_apps",
        "matched_in_columns",
    ]
    if not unmanaged_app_values:
        return pd.DataFrame(columns=out_cols)

    if not rules_enabled_csv.exists():
        return pd.DataFrame(columns=out_cols)

    def _pick(fieldnames: List[str], *cands: str) -> str:
        low = {c.lower(): c for c in (fieldnames or [])}
        for c in cands:
            if c.lower() in low:
                return low[c.lower()]
        return ""

    rows_out: List[Dict[str, str]] = []
    with rules_enabled_csv.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except Exception:
            dialect = csv.excel
        r = csv.DictReader(f, dialect=dialect)
        fns = r.fieldnames or []

        c_ruleset_name = _pick(fns, "ruleset_name")
        c_ruleset_scope = _pick(fns, "ruleset_scope")
        c_rule_type = _pick(fns, "rule_type")
        c_rule_desc = _pick(fns, "rule_description")
        c_rule_enabled = _pick(fns, "rule_enabled")
        c_services = _pick(fns, "services")

        # Columns requested by spec (letters based on export rules enabled)
        c_src_labels = _pick(fns, "src_labels")
        c_src_excl = _pick(fns, "src_labels_exclusions")
        c_dst_labels = _pick(fns, "dst_labels")
        c_dst_excl = _pick(fns, "dst_labels_exclusions")

        to_scan: List[Tuple[str, str]] = [
            ("ruleset_scope", c_ruleset_scope),
            ("src_labels", c_src_labels),
            ("src_labels_exclusions", c_src_excl),
            ("dst_labels", c_dst_labels),
            ("dst_labels_exclusions", c_dst_excl),
        ]

        for row in r:
            def g(c: str) -> str:
                return (row.get(c) or "").strip() if c else ""

            matched: Set[str] = set()
            matched_fields: List[str] = []
            for fname, col in to_scan:
                v = g(col)
                if not v:
                    continue
                apps = extract_app_label_values(v)
                hit = apps & unmanaged_app_values
                if hit:
                    matched |= hit
                    matched_fields.append(fname)

            if not matched:
                continue

            rows_out.append({
                "ruleset_name": g(c_ruleset_name),
                "ruleset_scope": g(c_ruleset_scope),
                "rule_type": g(c_rule_type),
                "rule_description": g(c_rule_desc),
                "rule_enabled": g(c_rule_enabled),
                "services": g(c_services),
                "src_labels": g(c_src_labels),
                "src_labels_exclusions": g(c_src_excl),
                "dst_labels": g(c_dst_labels),
                "dst_labels_exclusions": g(c_dst_excl),
                "matched_umgd_apps": ",".join(sorted(matched)),
                "matched_in_columns": ",".join(matched_fields),
            })

    return pd.DataFrame(rows_out, columns=out_cols)

def format_dark_header_table(ws, header_color: str = "404040") -> None:
    """Apply a standard formatting: dark header, borders, freeze header."""
    ws.freeze_panes = "A2"
    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=header_color)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(vertical="center", wrap_text=True)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = border
            cell.alignment = hdr_align

    body_align = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = body_align

def build_final_csv(derived_dir: Path, all_rows: List[Dict[str, str]]) -> Path:
    final_csv = derived_dir / 'export_dupecheck.final.csv'
    ensure_dir(final_csv)
    final_order = ['Info',
                   'href_managed','hostname_managed','role_managed','app_managed','env_managed','loc_managed','os_managed','interfaces_managed','managed','enforcement',
                   'href_unmanaged','hostname_unmanaged','name_unmanaged','interfaces_unmanaged','role_unmanaged','app_unmanaged','env_unmanaged','loc_unmanaged','os_unmanaged',
                   'match_type','reason']
    with final_csv.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=final_order)
        w.writeheader()
        for r in all_rows:
            if 'Info' not in r:
                r['Info'] = ''
            w.writerow({k: r.get(k, '') for k in final_order})
    print(f"[DEBUG] final_csv rows written: {len(all_rows)} -> {final_csv}")
    return final_csv

def build_final_excel(
    excel_dir: Path,
    raw_dir: Path,
    all_rows: List[Dict[str, str]],
    processes_rows: List[Dict[str, str]],
    start: str,
    end: str,
    app: str,
    envl: str,
    role: str,
    lbl_groups_incl_scope: Dict[Tuple[str, str], List[Dict[str, str]]],
    counts: Dict[str, int],
    exec_start_dt: datetime,
    include_flow_sheets: bool = True,
    add_elected_iplist_column: bool = False,
    enable_umgd_app_label_rules_sheet: bool = False,
    network_zone_name: Optional[str] = None,
) -> Path:
    ts = now_stamp('%Y%m%d-%H%M%S')
    name_parts = [sanitize_token(p) for p in (app, envl, role) if p]
    base = "-".join(name_parts) if name_parts else "scope"
    final_xlsx = excel_dir / f"export_dupecheck.final_{base}_{ts}.xlsx"
    ensure_dir(final_xlsx)

    from pandas import ExcelWriter
    with ExcelWriter(str(final_xlsx), engine='openpyxl') as writer:
        wsname = 'Summary'
        pd.DataFrame({'Summary': []}).to_excel(writer, index=False, sheet_name=wsname)
        ws_sum = writer.sheets[wsname]
        TITLE = Font(bold=True, size=14)
        H2 = Font(bold=True, size=12)
        F_BOLD = Font(bold=True)
        THIN = Side(style='thin', color='666666')
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        FILL_H = PatternFill('solid', fgColor='D9E1F2')
        BLUE_HDR = PatternFill('solid', fgColor='FF9DC3E6')

        ws_sum['A1'] = 'Carto NG — Summary'
        ws_sum['A1'].font = TITLE

        # Header: scope / time window / execution timestamps (key/value layout)
        scope_parts = []
        if app:
            scope_parts.append(f"app={app}")
        if envl:
            scope_parts.append(f"env={envl}")
        if role:
            scope_parts.append(f"role={role}")
        scope_val = " ".join(scope_parts) if scope_parts else "(none)"

        exec_start_str = exec_start_dt.strftime('%Y-%m-%d %H:%M:%S')

        KV_HDR = Font(bold=True)

        def _kv(row: int, label: str, value: str) -> None:
            # label
            c1 = ws_sum.cell(row=row, column=1, value=label)
            c1.font = KV_HDR
            c1.fill = BLUE_HDR
            c1.border = BORDER
            c1.alignment = Alignment(vertical='center')

            # value (merged across B:E for readability)
            ws_sum.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            c2 = ws_sum.cell(row=row, column=2, value=value)
            c2.border = BORDER
            c2.alignment = Alignment(vertical='center', wrap_text=True)

            # apply borders to merged cells (best effort)
            for col in range(3, 6):
                ws_sum.cell(row=row, column=col).border = BORDER

        _kv(2, 'Scope', scope_val)
        _kv(3, 'Time window', f"Start: {start}    End: {end}")
        _kv(4, 'Execution start', exec_start_str)
        _kv(5, 'Execution end', '...')
        _kv(6, 'Total duration', '...')

        ws_sum['A8'] = 'Section 1: Exported files'
        ws_sum['A8'].font = H2

        headers = ['File', 'Rows', 'Size (bytes)', 'Duration']
        for i, h in enumerate(headers, start=1):
            c = ws_sum.cell(row=9, column=i, value=h)
            c.font = F_BOLD
            c.fill = FILL_H
            c.border = BORDER

        files = [
            ('export_label.csv', raw_dir/'export_label.csv', 'label-export'),
            ('export_labelgroup.csv', raw_dir/'export_labelgroup.csv', 'labelgroup-export'),
            ('export_wkld.m.csv', raw_dir/'export_wkld.m.csv', 'wkld-export-m'),
            ('export_processes.csv', raw_dir/'export_processes.csv', 'process-export-aggregate'),
            ('export_wkld.u.csv', raw_dir/'export_wkld.u.csv', 'wkld-export-u'),
            ('export_ruleset.csv', raw_dir/'export_ruleset.csv', 'ruleset-export'),
            ('export_rules.enabled.csv', raw_dir/'export_rules.enabled.csv', 'rule-export-enabled-rs'),
            ('export_dupecheck.csv', raw_dir/'export_dupecheck.csv', 'dupecheck'),
            (f'flows_out_{start}_{end}.csv', raw_dir/f"flows_out_{start}_{end}.csv", 'traffic-out'),
            (f'flows_in_{start}_{end}.csv', raw_dir/f"flows_in_{start}_{end}.csv", 'traffic-in'),
        ]
        r = 10
        for name, path, label in files:
            ws_sum.cell(row=r, column=1, value=name).border = BORDER
            if path and path.exists():
                rows, _ = load_csv(path)
                ws_sum.cell(row=r, column=2, value=len(rows)).border = BORDER
                ws_sum.cell(row=r, column=3, value=path.stat().st_size).border = BORDER
            else:
                ws_sum.cell(row=r, column=2, value=0).border = BORDER
                ws_sum.cell(row=r, column=3, value=0).border = BORDER
            dur = DUR.get(label, 0.0)
            ws_sum.cell(row=r, column=4, value=f"{dur:.1f}s").border = BORDER
            r += 1
        ws_sum.column_dimensions['A'].width = 40
        ws_sum.column_dimensions['B'].width = 16
        ws_sum.column_dimensions['C'].width = 16
        ws_sum.column_dimensions['D'].width = 14
        ws_sum.column_dimensions['E'].width = 70

        ws_sum.cell(row=r + 1, column=1, value='Section 2: Label Groups inclusion (Scope-wide)').font = H2
        r2 = r + 2
        hdr2 = ['Label Key', 'Label Value', 'Group Name']
        for i, h in enumerate(hdr2, start=1):
            c = ws_sum.cell(row=r2, column=i, value=h)
            c.font = F_BOLD
            c.fill = FILL_H
            c.border = BORDER
        r2 += 1

        ORDER = ['role', 'app', 'env', 'loc', 'OS']
        order_index = {k: i for i, k in enumerate(ORDER)}
        items = list(lbl_groups_incl_scope.items())
        items_sorted = sorted(items, key=lambda kv: (order_index.get(kv[0][0], 999), str(kv[0][1]).lower()))
        for (lk, lv), group_rows in items_sorted:
            if not group_rows:
                ws_sum.cell(row=r2, column=1, value=lk).border = BORDER
                ws_sum.cell(row=r2, column=2, value=lv).border = BORDER
                ws_sum.cell(row=r2, column=3, value='').border = BORDER
                ws_sum.cell(row=r2, column=4, value='').border = BORDER
                ws_sum.cell(row=r2, column=5, value='').border = BORDER
                r2 += 1
                continue
            for it in group_rows:
                ws_sum.cell(row=r2, column=1, value=lk).border = BORDER
                ws_sum.cell(row=r2, column=2, value=lv).border = BORDER
                ws_sum.cell(row=r2, column=3, value=it.get('name', '')).border = BORDER
                r2 += 1
        for col in ['A', 'B', 'C']:
            ws_sum.column_dimensions[col].width = max(ws_sum.column_dimensions[col].width or 10, 28)

        ws_sum.cell(row=r2 + 2, column=1, value='Section 3: Workloads status').font = H2
        s3r = r2 + 3
        titles = [
            ('Managed workloads (scope):', int(counts.get('managed_scope_total', 0))),
            ('With unmanaged duplicate:', int(counts.get('duplicates', 0))),
            ('Missing agent (unmanaged-only):', int(counts.get('unmanaged_only', 0))),
            ('Managed without duplicate (not in CMDB):', int(counts.get('managed_only', 0))),
        ]
        for text, val in titles:
            c_title = ws_sum.cell(row=s3r, column=1, value=text)
            c_title.font = F_BOLD
            c_title.fill = FILL_H
            c_title.border = BORDER
            c_val = ws_sum.cell(row=s3r, column=2, value=val)
            c_val.border = BORDER
            s3r += 1

        dupe_cols = [
            'Info',
            'href_managed','hostname_managed','role_managed','app_managed','env_managed','loc_managed','os_managed','interfaces_managed','managed','enforcement',
            'href_unmanaged','hostname_unmanaged','name_unmanaged','interfaces_unmanaged','role_unmanaged','app_unmanaged','env_unmanaged','loc_unmanaged','os_unmanaged',
            'match_type','reason'
        ]
        if network_zone_name:
            insert_idx = dupe_cols.index('interfaces_managed') + 1
            dupe_cols = dupe_cols[:insert_idx] + ['Inside Network Zone'] + dupe_cols[insert_idx:]
        df_dupe = pd.DataFrame(all_rows, columns=dupe_cols if all_rows else dupe_cols)
        if network_zone_name:
            nets, err = _load_zone_networks(raw_dir, network_zone_name)
            if err:
                msg = f"WARN: cannot compute Inside Network Zone ({err})"
                print(msg)
                try:
                    logger.warning(msg)
                except Exception:
                    pass
                df_dupe['Inside Network Zone'] = ["UNKNOWN"] * len(df_dupe.index)
            else:
                nets_v4 = [n for n in nets if isinstance(n, ipaddress.IPv4Network)]
                nets_v6 = [n for n in nets if isinstance(n, ipaddress.IPv6Network)]
                inside_vals: List[str] = []
                for r in all_rows:
                    iface_val = r.get('interfaces_managed', '') or ''
                    if not iface_val:
                        iface_val = r.get('interfaces_unmanaged', '') or ''
                    ips = _extract_interface_ips(iface_val)
                    ips = _extract_interface_ips(r.get('interfaces_managed', '') or '')
                    in_zone = any(_ip_in_zone(ip_s, nets_v4, nets_v6) for ip_s in ips)
                    inside_vals.append("Y" if in_zone else "N")
                df_dupe['Inside Network Zone'] = inside_vals
        df_dupe.to_excel(writer, index=False, sheet_name='Workloads')
        ws_w = writer.sheets['Workloads']
        ws_w.freeze_panes = 'A2'
        ws_w.auto_filter.ref = ws_w.dimensions

        for cell in ws_w[1]:
            cell.font = Font(bold=True)
        fill_info = PatternFill('solid', fgColor='D9D9D9')
        fill_man = PatternFill('solid', fgColor='C6EFCE')
        fill_unm = PatternFill('solid', fgColor='FCE4D6')
        fill_match = PatternFill('solid', fgColor='D9E1F2')
        THIN2 = Side(style='thin', color='666666')
        BORDER2 = Border(left=THIN2, right=THIN2, top=THIN2, bottom=THIN2)
        header_idx = {cell.value: idx for idx, cell in enumerate(ws_w[1], start=1)}
        nrows_w = ws_w.max_row

        def paint(col_name: str, fill: PatternFill) -> None:
            col = header_idx.get(col_name)
            if not col:
                return
            for r_i in range(1, nrows_w + 1):
                c = ws_w.cell(row=r_i, column=col)
                c.fill = fill
                c.border = BORDER2

        paint('Info', fill_info)
        managed_cols = ['href_managed','hostname_managed','role_managed','app_managed','env_managed','loc_managed','os_managed','interfaces_managed','managed','enforcement']
        if network_zone_name:
            managed_cols.insert(managed_cols.index('interfaces_managed') + 1, 'Inside Network Zone')
        for c in managed_cols:
            paint(c, fill_man)
        for c in ['href_unmanaged','hostname_unmanaged','name_unmanaged','interfaces_unmanaged','role_unmanaged','app_unmanaged','env_unmanaged','loc_unmanaged','os_unmanaged']:
            paint(c, fill_unm)
        for c in ['match_type','reason']:
            paint(c, fill_match)

        auto_width(ws_w)

        # NEW (gated): Rules with unmanaged app labels (from Workloads[app_unmanaged])
        if enable_umgd_app_label_rules_sheet:
            try:
                unmanaged_vals: Set[str] = set()
                if "app_unmanaged" in df_dupe.columns:
                    unmanaged_vals = set(
                        str(v).strip()
                        for v in df_dupe["app_unmanaged"].dropna().tolist()
                        if str(v).strip()
                    )
                rules_csv = raw_dir / "export_rules.enabled.csv"
                df_umgd = find_rules_with_unmanaged_app_labels(
                    rules_enabled_csv=rules_csv,
                    unmanaged_app_values=unmanaged_vals,
                )
                df_umgd.to_excel(writer, index=False, sheet_name="Rules with umgd app labels")
                ws_umgd = writer.sheets["Rules with umgd app labels"]
                format_dark_header_table(ws_umgd, header_color="1F4E79")
                auto_width(ws_umgd)
            except Exception as e:
                msg = f"WARN: [Rules with umgd app labels] skipped due to: {e}"
                print(msg)
                try:
                    logger.warning(msg)
                except Exception:
                    pass


        proc_cols = ['hostname','process_path','service_name','port','proto','role','app','env','loc','OS','href']
        df_proc = pd.DataFrame(processes_rows, columns=proc_cols if processes_rows else proc_cols)
        if 'port' in df_proc.columns:
            try:
                df_proc['port'] = pd.to_numeric(df_proc['port'], errors='coerce')
            except Exception:
                pass
        if not df_proc.empty:
            sort_cols = [c for c in ['hostname','role'] if c in df_proc.columns]
            if sort_cols:
                df_proc = df_proc.sort_values(by=sort_cols, kind='stable')
            df_proc.to_excel(writer, index=False, sheet_name='Processes')
            ws_p = writer.sheets['Processes']
            ws_p.freeze_panes = 'A2'
            ws_p.auto_filter.ref = ws_p.dimensions
            for cell in ws_p[1]:
                cell.font = Font(bold=True)
            blue1 = PatternFill('solid', fgColor='00DDEBF7')
            blue2 = PatternFill('solid', fgColor='00EBF3FB')
            THINP = Side(style='thin', color='666666')
            BORDERP = Border(left=THINP, right=THINP, top=THINP, bottom=THINP)
            max_row_p = ws_p.max_row
            max_col_p = ws_p.max_column
            current_host = None
            toggle = False
            for r_i in range(2, max_row_p + 1):
                host_val = ws_p.cell(row=r_i, column=1).value
                if host_val != current_host:
                    current_host = host_val
                    toggle = not toggle
                fill = blue1 if toggle else blue2
                for c_i in range(1, max_col_p + 1):
                    cell = ws_p.cell(row=r_i, column=c_i)
                    cell.fill = fill
                    cell.border = BORDERP
            for r_i in range(2, max_row_p + 1):
                c_port = ws_p.cell(row=r_i, column=4)
                c_port.number_format = '0'
                c_port.alignment = Alignment(horizontal='right')
            auto_width(ws_p)

        def style_flow_sheet(df: pd.DataFrame, sheet_name: str, col_sets: Dict[str, List[str]]) -> None:
            df = df.copy()
            for num_col in ['Port', 'Num Flows']:
                if num_col in df.columns:
                    try:
                        df[num_col] = pd.to_numeric(df[num_col], errors='coerce')
                    except Exception:
                        pass
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            wsf = writer.sheets[sheet_name]
            wsf.freeze_panes = 'A2'
            wsf.auto_filter.ref = wsf.dimensions
            GREEN_LIGHT = PatternFill('solid', fgColor='00CCFFCC')
            GREEN_HDR = PatternFill('solid', fgColor='0099CC99')
            ORANGE_LIGHT = PatternFill('solid', fgColor='00FFE5CC')
            ORANGE_HDR = PatternFill('solid', fgColor='00FFCC99')
            BLUE_LIGHT = PatternFill('solid', fgColor='00DDEBF7')
            BLUE_HDR2 = PatternFill('solid', fgColor='009DC3E6')
            THINX = Side(style='thin', color='666666')
            BORDERX = Border(left=THINX, right=THINX, top=THINX, bottom=THINX)
            headers = [str(c.value or '') for c in wsf[1]]
            header_pos = {h: i for i, h in enumerate(headers, start=1)}
            for i, h in enumerate(headers, start=1):
                cell = wsf.cell(row=1, column=i)
                cell.font = Font(bold=True)
                cell.border = BORDERX
                hl = h.strip()
                if hl in col_sets.get('green', []):
                    cell.fill = GREEN_HDR
                elif hl in col_sets.get('orange', []):
                    cell.fill = ORANGE_HDR
                elif hl in col_sets.get('blue', []):
                    cell.fill = BLUE_HDR2
                else:
                    cell.fill = PatternFill('solid', fgColor='00D9E1F2')
            max_row_f = wsf.max_row
            for r_i in range(2, max_row_f + 1):
                for h, col in header_pos.items():
                    cell = wsf.cell(row=r_i, column=col)
                    cell.border = BORDERX
                    if h in col_sets.get('green', []):
                        cell.fill = GREEN_LIGHT
                    elif h in col_sets.get('orange', []):
                        cell.fill = ORANGE_LIGHT
                    elif h in col_sets.get('blue', []):
                        cell.fill = BLUE_LIGHT
                    if h.lower() in ['port', 'num flows']:
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='right')
            special_caps: Dict[str, int] = {}
            for h in headers:
                if 'iplists' in h.lower():
                    special_caps[h.lower()] = 200
            auto_width(wsf, special_caps=special_caps)

        if include_flow_sheets:
            flow_out_path = raw_dir / f"flows_out_{start}_{end}.csv"
            flow_in_path = raw_dir / f"flows_in_{start}_{end}.csv"
            df_flow_out = pd.DataFrame(load_csv(flow_out_path)[0]) if flow_out_path.exists() else None
            df_flow_in = pd.DataFrame(load_csv(flow_in_path)[0]) if flow_in_path.exists() else None
            # Optional (gated): add an explicit per-row elected IPList column (visibility/debug)
            elect_allowed_pats: List[str] = []
            elect_prio_pats: List[str] = []

            def _match_prefix_or_glob(name: str, pat: str) -> bool:
                if not pat:
                    return False
                if pat.endswith('*'):
                    return name.startswith(pat[:-1])
                return name.startswith(pat)

            def _elect_iplist_cell(cell: str) -> str:
                toks = [t.strip() for t in re.split(r"[;,\s]+", (cell or "")) if t.strip()]
                if not toks:
                    return ""
                names: List[str] = []
                for t in toks:
                    names.append(t.rsplit('/', 1)[-1] if '/' in t else t)

                if elect_allowed_pats:
                    kept = [n for n in names if any(_match_prefix_or_glob(n, p) for p in elect_allowed_pats)]
                    if kept:
                        names = kept

                for p in elect_prio_pats:
                    for n in names:
                        if _match_prefix_or_glob(n, p):
                            return n

                return sorted(set(names))[0]

            if add_elected_iplist_column:
                try:
                    conf_local = load_conf('carto.conf')
                    elect_allowed_pats = [x for x in (conf_local.get('IPLIST_ALLOWED_PREFIXES', '') or '').split(';') if x]
                    elect_prio_pats = [x for x in (conf_local.get('IPLIST_NAME_PRIORITY', '') or '').split(';') if x]
                except Exception:
                    elect_allowed_pats = []
                    elect_prio_pats = []

            if df_flow_out is not None and not df_flow_out.empty:
                cols_green_out = ['Source IP','Source IPList','Source Name','Source Hostname','Source Managed','Source Enforcement Mode','Source Application','Source Environment','Source Location','Source OS','Source Role','Source Process','Source Service','Source Username']
                cols_orange_out = ['Destination IP','Destination IPList','Destination Name','Destination Hostname','Destination Managed','Destination Enforcement Mode','Destination Application','Destination Environment','Destination Location','Destination OS','Destination Role','Destination FQDN','Destination Process','Destination Service','Destination Username']
                cols_blue_out = ['Transmission','Port','Protocol','Num Flows','Connection State','Reported Policy Decision','Reported Enforcement Boundary','Reported by','First Detected','Last Detected','Network','Bytes In','Bytes Out']
                if add_elected_iplist_column:
                    key_col = 'Destination IPList'
                    new_col = 'Destination IPList Elected'
                    try:
                        if key_col in df_flow_out.columns and new_col not in df_flow_out.columns:
                            elected = df_flow_out[key_col].astype(str).apply(_elect_iplist_cell)
                            df_flow_out.insert(df_flow_out.columns.get_loc(key_col) + 1, new_col, elected)
                    except Exception:
                        pass
                    try:
                        if new_col not in cols_orange_out:
                            if key_col in cols_orange_out:
                                cols_orange_out.insert(cols_orange_out.index(key_col) + 1, new_col)
                            else:
                                cols_orange_out.append(new_col)
                    except Exception:
                        pass

                style_flow_sheet(df_flow_out, 'Flow-out', {'green': cols_green_out, 'orange': cols_orange_out, 'blue': cols_blue_out})
            if df_flow_in is not None and not df_flow_in.empty:
                cols_orange_src = ['Source IP','Source IPList','Source Name','Source Hostname','Source Managed','Source Enforcement Mode','Source Application','Source Environment','Source Location','Source OS','Source Role','Source Process','Source Service','Source Username']
                cols_green_dst = ['Destination IP','Destination IPList','Destination Name','Destination Hostname','Destination Managed','Destination Enforcement Mode','Destination Application','Destination Environment','Destination Location','Destination OS','Destination Role','Destination FQDN','Destination Process','Destination Service','Destination Username']
                cols_blue_in = ['Transmission','Port','Protocol','Num Flows','Connection State','Reported Policy Decision','Reported Enforcement Boundary','Reported by','First Detected','Last Detected','Network','Bytes In','Bytes Out']
                if add_elected_iplist_column:
                    key_col = 'Source IPList'
                    new_col = 'Source IPList Elected'
                    try:
                        if key_col in df_flow_in.columns and new_col not in df_flow_in.columns:
                            elected = df_flow_in[key_col].astype(str).apply(_elect_iplist_cell)
                            df_flow_in.insert(df_flow_in.columns.get_loc(key_col) + 1, new_col, elected)
                    except Exception:
                        pass
                    try:
                        if new_col not in cols_orange_src:
                            if key_col in cols_orange_src:
                                cols_orange_src.insert(cols_orange_src.index(key_col) + 1, new_col)
                            else:
                                cols_orange_src.append(new_col)
                    except Exception:
                        pass

                style_flow_sheet(df_flow_in, 'Flow-in', {'green': cols_green_dst, 'orange': cols_orange_src, 'blue': cols_blue_in})
        else:
            # Skip Flow-in/out sheets (large) to keep the workbook small.
            # They can be injected at the end via modules/excel_stream_update.py when --excel-stream-update is enabled.
            pass


        # Update execution end/duration now that the report is fully generated
        exec_end_dt = datetime.now()
        exec_end_str = exec_end_dt.strftime('%Y-%m-%d %H:%M:%S')
        duration_td = exec_end_dt - exec_start_dt
        duration_str = str(duration_td).split('.')[0]
        ws_sum['B5'] = exec_end_str
        ws_sum['B6'] = duration_str
        ws_sum['B5'].border = BORDER
        ws_sum['B6'].border = BORDER
        ws_sum['B5'].alignment = Alignment(vertical='center')
        ws_sum['B6'].alignment = Alignment(vertical='center')


    return final_xlsx

# ------------------------------ Main ------------------------------


# ------------------------------ To investigate (NZ0_/NZ1_/DNA_/DNS_ + egress KUB_/LBI_/LBO_/USR_/U_) ------------------------------
INV_INGRESS_PREFIXES = ("NZ0_", "NZ1_", "DNA_", "DNS_")
INV_EGRESS_PREFIXES = ("NZ0_", "NZ1_", "KUB_", "LBI_", "LBO_", "USR_", "U_")

def _toinvest_find_col(headers, candidates):
    if not headers:
        return None
    norm = [str(h or '').strip().lower().replace(' ', '').replace('-', '').replace('_', '') for h in headers]
    # exact match
    for cand in candidates:
        nc = str(cand).strip().lower().replace(' ', '').replace('-', '').replace('_', '')
        if not nc:
            continue
        for i, nh in enumerate(norm):
            if nh == nc:
                return i
    # containment
    for cand in candidates:
        nc = str(cand).strip().lower().replace(' ', '').replace('-', '').replace('_', '')
        if not nc:
            continue
        for i, nh in enumerate(norm):
            if nc in nh:
                return i
    return None


class _ToInvestDNS:
    def __init__(self, timeout_s: float = 1.5):
        self.timeout_s = float(timeout_s)
        self.cache = {}

    def resolve(self, ip: str) -> str:
        ip = (ip or '').strip()
        if not ip:
            return ''
        if ip in self.cache:
            return self.cache[ip]
        try:
            ipaddress.ip_address(ip)
        except Exception:
            self.cache[ip] = ''
            return ''
        prev = socket.getdefaulttimeout()
        try:
            socket.setdefaulttimeout(self.timeout_s)
            host, _aliases, _addrs = socket.gethostbyaddr(ip)
            res = host or ''
        except Exception:
            res = ''
        finally:
            socket.setdefaulttimeout(prev)
        self.cache[ip] = res
        return res


def build_to_investigate_sheet(xlsx_path: Path, *, dns_timeout: float = 1.5) -> int:
    """Create/replace a sheet 'To investigate' from Flow-in/Flow-out elected NZ0_/NZ1_/DNA_/DNS_.

    Includes egress (Flow-out) rows with elected KUB_/LBI_/LBO_/USR_/U_ prefixes.

    Trivial branch only (small/normal workbook). Best-effort reverse DNS.

    Output columns:
      - Direction
      - Source
      - Destination
      - Service
      - Unknown IP
      - DNS Resolution

    Mapping rules:
      - Direction=Flow-out:  Source=Source Role, Destination=Destination IPList Elected
      - Direction=Flow-in :  Source=Source IPList Elected, Destination=Destination Role

    Duplicates are removed (stable: keep first occurrence).
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = load_workbook(xlsx_path)

    # Remove existing sheet to avoid stale data
    if 'To investigate' in wb.sheetnames:
        del wb['To investigate']

    ws_out = wb['Flow-out'] if 'Flow-out' in wb.sheetnames else None
    ws_in = wb['Flow-in'] if 'Flow-in' in wb.sheetnames else None

    rows = []
    seen = set()
    dns = _ToInvestDNS(timeout_s=dns_timeout)

    def scan(ws, direction: str):
        if ws is None:
            return
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return
        headers = list(header)

        elected_hdr = 'Destination IPList Elected' if direction == 'out' else 'Source IPList Elected'
        idx_elected = _toinvest_find_col(headers, [elected_hdr])

        unknown_hdr = 'Destination IP' if direction == 'out' else 'Source IP'
        idx_unknown = _toinvest_find_col(headers, [unknown_hdr, 'Dest IP' if direction == 'out' else 'Src IP'])

        idx_src_role = _toinvest_find_col(headers, ['Source Role', 'Src Role'])
        idx_dst_role = _toinvest_find_col(headers, ['Destination Role', 'Dst Role'])

        # Fallbacks (only if role columns are missing)
        idx_src_fb = _toinvest_find_col(headers, ['Source Labels', 'Source Workload', 'Source', 'Source IPList', 'Source IP'])
        idx_dst_fb = _toinvest_find_col(headers, ['Destination Labels', 'Destination Workload', 'Destination', 'Destination IPList', 'Destination IP'])

        idx_service = _toinvest_find_col(headers, ['Service', 'Services', 'Service Name'])
        idx_proto = _toinvest_find_col(headers, ['Protocol', 'Proto'])
        idx_port = _toinvest_find_col(headers, ['Port', 'Dst Port', 'Destination Port'])

        for r in it:
            if not r:
                continue
            r = list(r)

            elected_val = ''
            if idx_elected is not None and idx_elected < len(r):
                elected_val = str(r[idx_elected] or '')
            prefixes = INV_EGRESS_PREFIXES if direction == 'out' else INV_INGRESS_PREFIXES
            if not elected_val.startswith(prefixes):
                continue

            unknown_ip = ''
            if idx_unknown is not None and idx_unknown < len(r):
                unknown_ip = str(r[idx_unknown] or '')

            # Service
            service = ''
            if idx_service is not None and idx_service < len(r):
                service = str(r[idx_service] or '')
            if not service:
                proto = str(r[idx_proto] or '') if idx_proto is not None and idx_proto < len(r) else ''
                port = str(r[idx_port] or '') if idx_port is not None and idx_port < len(r) else ''
                if proto and port:
                    service = f"{proto}/{port}"
                else:
                    service = proto or port

            if direction == 'out':
                # Source Role / Destination IPList Elected
                src_val = ''
                if idx_src_role is not None and idx_src_role < len(r):
                    src_val = str(r[idx_src_role] or '')
                if not src_val and idx_src_fb is not None and idx_src_fb < len(r):
                    src_val = str(r[idx_src_fb] or '')
                dst_val = elected_val
            else:
                # Source IPList Elected / Destination Role
                src_val = elected_val
                dst_val = ''
                if idx_dst_role is not None and idx_dst_role < len(r):
                    dst_val = str(r[idx_dst_role] or '')
                if not dst_val and idx_dst_fb is not None and idx_dst_fb < len(r):
                    dst_val = str(r[idx_dst_fb] or '')

            dns_val = dns.resolve(unknown_ip)

            key = (
                'Flow-out' if direction == 'out' else 'Flow-in',
                src_val,
                dst_val,
                service,
                unknown_ip,
                dns_val,
            )
            if key in seen:
                continue
            seen.add(key)

            rows.append(list(key))

    scan(ws_out, 'out')
    scan(ws_in, 'in')

    ws = wb.create_sheet('To investigate')

    # Style header
    THIN = Side(style='thin', color='666666')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    hdr_fill = PatternFill('solid', fgColor='D9D9D9')
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers_out = ['Direction', 'Source', 'Destination', 'Service', 'Unknown IP', 'DNS Resolution']
    ws.append(headers_out)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = BORDER
        cell.alignment = hdr_align

    for row in rows:
        ws.append(row)

    # Border for full table
    for rr in range(2, len(rows) + 2):
        for cc in range(1, 7):
            ws.cell(row=rr, column=cc).border = BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 45

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:F{len(rows)+1}" if rows else 'A1:F1'

    wb.save(xlsx_path)
    return len(rows)


def build_to_investigate_ip_sheets(xlsx_path: Path) -> Tuple[int, int]:
    """Create/replace IP summary sheets derived from 'To investigate'."""
    import ipaddress
    import re
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = load_workbook(xlsx_path)
    if "To investigate" not in wb.sheetnames:
        return 0, 0

    ws_src = wb["To investigate"]
    it = ws_src.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return 0, 0

    headers = list(header)
    idx_dir = _toinvest_find_col(headers, ["Direction"])
    idx_src = _toinvest_find_col(headers, ["Source"])
    idx_dst = _toinvest_find_col(headers, ["Destination"])
    idx_ip = _toinvest_find_col(headers, ["Unknown IP"])
    idx_dns = _toinvest_find_col(headers, ["DNS Resolution"])
    idx_fqdn = _toinvest_find_col(headers, ["FQDN found in PCE"])

    if idx_dir is None or idx_ip is None:
        return 0, 0

    ip_pattern = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")
    egress_rows: Dict[str, Dict[str, str]] = {}
    ingress_rows: Dict[str, Dict[str, str]] = {}

    for row in it:
        if not row:
            continue
        direction = str(row[idx_dir] or "").strip().lower()
        ip_cell = str(row[idx_ip] or "").strip()
        if not ip_cell:
            continue
        candidates = ip_pattern.findall(ip_cell)
        if not candidates:
            continue
        ip_val = candidates[0]

        if direction == "flow-out":
            if ip_val in egress_rows:
                continue
            egress_rows[ip_val] = {
                "Direction": "Flow-out",
                "Destination": str(row[idx_dst] or "") if idx_dst is not None else "",
                "Unknown IP": ip_val,
                "DNS Resolution": str(row[idx_dns] or "") if idx_dns is not None else "",
                "FQDN found in PCE": str(row[idx_fqdn] or "") if idx_fqdn is not None else "",
            }
        elif direction == "flow-in":
            if ip_val in ingress_rows:
                continue
            subnet = ""
            try:
                subnet = str(ipaddress.ip_network(f"{ip_val}/21", strict=False))
            except Exception:
                subnet = ""
            ingress_rows[ip_val] = {
                "Direction": "Flow-in",
                "Source": str(row[idx_src] or "") if idx_src is not None else "",
                "Unknown IP": ip_val,
                "Subnet /21": subnet,
            }

    # Remove existing sheets
    for name in ("Egress IP to investigate", "Ingress IP to investigate"):
        if name in wb.sheetnames:
            del wb[name]

    # Shared styling
    THIN = Side(style='thin', color='666666')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    hdr_fill = PatternFill('solid', fgColor='D9D9D9')
    hdr_font = Font(bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    normal_align = Alignment(horizontal='left', vertical='top', wrap_text=False)

    # Egress sheet
    ws_out = wb.create_sheet("Egress IP to investigate")
    egress_headers = ["Direction", "Destination", "Unknown IP", "DNS Resolution", "FQDN found in PCE", "Action"]
    ws_out.append(egress_headers)
    for col in range(1, len(egress_headers) + 1):
        cell = ws_out.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = BORDER
        cell.alignment = hdr_align

    for row in egress_rows.values():
        ws_out.append([row.get(h, "") for h in egress_headers])

    for rr in range(2, ws_out.max_row + 1):
        for cc in range(1, len(egress_headers) + 1):
            cell = ws_out.cell(row=rr, column=cc)
            cell.border = BORDER
            cell.alignment = normal_align

    ws_out.column_dimensions["A"].width = 14
    ws_out.column_dimensions["B"].width = 45
    ws_out.column_dimensions["C"].width = 18
    ws_out.column_dimensions["D"].width = 45
    ws_out.column_dimensions["E"].width = 45
    ws_out.column_dimensions["F"].width = 14  # ~100px

    ws_out.freeze_panes = 'A2'
    ws_out.auto_filter.ref = f"A1:F{ws_out.max_row}" if ws_out.max_row > 1 else "A1:F1"

    # Ingress sheet
    ws_in = wb.create_sheet("Ingress IP to investigate")
    ingress_headers = ["Direction", "Source", "Unknown IP", "Subnet /21"]
    ws_in.append(ingress_headers)
    for col in range(1, len(ingress_headers) + 1):
        cell = ws_in.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = BORDER
        cell.alignment = hdr_align

    for row in ingress_rows.values():
        ws_in.append([row.get(h, "") for h in ingress_headers])

    for rr in range(2, ws_in.max_row + 1):
        for cc in range(1, len(ingress_headers) + 1):
            cell = ws_in.cell(row=rr, column=cc)
            cell.border = BORDER
            cell.alignment = normal_align

    ws_in.column_dimensions["A"].width = 14
    ws_in.column_dimensions["B"].width = 45
    ws_in.column_dimensions["C"].width = 18
    ws_in.column_dimensions["D"].width = 20

    ws_in.freeze_panes = 'A2'
    ws_in.auto_filter.ref = f"A1:D{ws_in.max_row}" if ws_in.max_row > 1 else "A1:D1"

    wb.save(xlsx_path)
    return len(egress_rows), len(ingress_rows)


def _load_to_investigate_ips(xlsx_path: Path) -> List[str]:
    import ipaddress
    import re
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    if "To investigate" not in wb.sheetnames:
        return []
    ws = wb["To investigate"]

    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []

    headers = list(header)
    idx_dir = _toinvest_find_col(headers, ["Direction"])
    idx_ip = _toinvest_find_col(headers, ["Unknown IP"])
    if idx_dir is None or idx_ip is None:
        return []

    ips: List[str] = []
    seen: Set[str] = set()
    ip_pattern = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")
    for row in it:
        if not row:
            continue
        direction = str(row[idx_dir] or "").strip().lower()
        if direction != "flow-out":
            continue
        ip_val = str(row[idx_ip] or "").strip()
        if not ip_val:
            continue
        for candidate in ip_pattern.findall(ip_val):
            try:
                parsed_ip = ipaddress.ip_address(candidate)
            except ValueError:
                continue
            if parsed_ip.version != 4:
                continue
            if candidate.startswith("169.254."):
                continue
            if candidate not in seen:
                seen.add(candidate)
                ips.append(candidate)
    return ips


def _extract_fqdn_by_destination_ip(flow_csv: Path) -> Dict[str, str]:
    rows, cols = load_csv(flow_csv)
    if not rows:
        return {}

    idx_ip = pick(cols, "Destination IP", "Destination IPs", "Dst IP", "Destination_IP")
    if not idx_ip:
        return {}

    idx_fqdn = pick(cols, "Destination FQDN", "Dst FQDN", "Destination_FQDN")
    if not idx_fqdn:
        return {}

    ip_to_fqdn: Dict[str, str] = {}
    for row in rows:
        ip = (row.get(idx_ip) or "").strip()
        if not ip:
            continue
        fqdn = (row.get(idx_fqdn) or "").strip()
        if not fqdn:
            continue
        existing = ip_to_fqdn.get(ip, "")
        if not existing or existing.upper() == "NO_FQDN_FOUND":
            ip_to_fqdn[ip] = fqdn
    return ip_to_fqdn


def _update_to_investigate_fqdn(xlsx_path: Path, fqdn_map: Dict[str, str]) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    if "To investigate" not in wb.sheetnames:
        return 0
    ws = wb["To investigate"]

    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return 0

    headers = list(header)
    idx_dir = _toinvest_find_col(headers, ["Direction"])
    idx_ip = _toinvest_find_col(headers, ["Unknown IP"])
    if idx_dir is None or idx_ip is None:
        return 0

    fqdn_header = "FQDN found in PCE"
    idx_fqdn_col = _toinvest_find_col(headers, [fqdn_header])
    if idx_fqdn_col is None:
        idx_fqdn_col = len(headers)
        ws.cell(row=1, column=idx_fqdn_col + 1, value=fqdn_header)
        ref_cell = ws.cell(row=1, column=idx_ip + 1)
        new_cell = ws.cell(row=1, column=idx_fqdn_col + 1)
        if ref_cell.has_style:
            new_cell._style = ref_cell._style
        ws.column_dimensions[new_cell.column_letter].width = 45

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        dir_val = str(ws.cell(row=row_idx, column=idx_dir + 1).value or "").strip().lower()
        if dir_val != "flow-out":
            continue
        ip_val = str(ws.cell(row=row_idx, column=idx_ip + 1).value or "").strip()
        if not ip_val:
            continue
        fqdn = fqdn_map.get(ip_val, "NO_FQDN_FOUND")
        ws.cell(row=row_idx, column=idx_fqdn_col + 1, value=fqdn)
        updated += 1

    wb.save(xlsx_path)
    return updated


def enrich_unknown_ips_with_pce_fqdn(
    *,
    xlsx_path: Path,
    bin_dir: Path,
    env: Dict[str, str],
    derived_dir: Path,
    start: str,
    end: str,
    tmp_stamp: str,
) -> None:
    ips = _load_to_investigate_ips(xlsx_path)
    if not ips:
        print("[INFO] [To investigate] No Flow-out unknown IPs found for PCE lookup.")
        return

    tmp_prefix = f"_tmp_carto_egress.unknonw.ip_{tmp_stamp}"
    iplist_name = f"{tmp_prefix}-IPL"
    iplist_csv = (derived_dir / f"{tmp_prefix}.csv").resolve()
    iplist_csv.parent.mkdir(parents=True, exist_ok=True)
    with iplist_csv.open("w", encoding="utf-8", newline="") as f:
        f.write("name,description,include\n")
        include = ";".join(ips)
        description = "Temporary IPList for identifying fqdn entries. Do not Use IT Will be automatically deleted !!"
        f.write(f"{iplist_name},{description},{include}\n")

    ipl_export_all = (derived_dir / f"{tmp_prefix}.iplists.csv").resolve()
    href_file = (derived_dir / f"href.iplist_{tmp_prefix}.csv").resolve()
    flow_out = (derived_dir / f"flow_out_{start}-{end}.csv").resolve()

    href_file_created = False
    try:
        ok = run_step("ipl-import-unknown", ["bash", str(bin_dir / "workloader_ipl_import.sh"), str(iplist_csv)], env, Path("."))
        if not ok:
            print("[WARN] [To investigate] IPL import failed; skipping PCE FQDN lookup.")
            return

        ok = run_step(
            "ipl-export-all-unknown",
            ["bash", str(bin_dir / "workloader_ipl_export.sh"), str(ipl_export_all)],
            env,
            Path("."),
        )
        if not ok:
            print("[WARN] [To investigate] IPL export (all) failed; skipping PCE FQDN lookup.")
            return

        rows, cols = load_csv(ipl_export_all)
        name_col = pick(cols, "name", "Name", "NAME")
        href_col = pick(cols, "href", "Href", "HREF")
        href_val = ""
        if name_col and href_col and rows:
            for row in rows:
                if (row.get(name_col) or "").strip() == iplist_name:
                    href_val = (row.get(href_col) or "").strip()
                    if href_val:
                        break
        if not href_val:
            print("[WARN] [To investigate] Missing href in IPL export; skipping PCE FQDN lookup.")
            return

        write_list_semicolon(href_file, [href_val])
        href_file_created = True

        ok = run_step(
            "traffic-out-unknown",
            ["bash", str(bin_dir / "workloader_traffic_out_dst.sh"), str(href_file), start, end, str(flow_out)],
            env,
            Path("."),
        )
        if not ok:
            print("[WARN] [To investigate] traffic-out failed; skipping PCE FQDN lookup.")
            return

        fqdn_map = _extract_fqdn_by_destination_ip(flow_out)
        updated = _update_to_investigate_fqdn(xlsx_path, fqdn_map)
        print(f"[INFO] [To investigate] FQDN updated rows: {updated}")
    finally:
        if href_file_created and href_file.exists():
            run_step(
                "ipl-delete-unknown",
                ["bash", str(bin_dir / "workloader_ipl_delete.sh"), str(href_file)],
                env,
                Path("."),
            )

def main() -> int:
    if "--helpdev" in sys.argv:
        ap = argparse.ArgumentParser("carto_ng_orchestrator.py", formatter_class=argparse.RawTextHelpFormatter)
    else:
        ap = argparse.ArgumentParser(
            "carto_ng_orchestrator.py",
            add_help=False,
            formatter_class=argparse.RawTextHelpFormatter,
        )
        ap.add_argument("-h", "--help", action="help", help="show this help message and exit")
        ap.add_argument("--helpdev", action="help", help="show full help (including dev and frh options)")

    # Scope labels
    ap.add_argument(
        "--role",
        nargs="+",
        help=(
            "Role label filter.\n"
            "  - Single role: --role FRONTEND\n"
            "  - Multiple roles: --role FRONTEND,DATABASE,MG01\n"
            "    Also accepted with spaces: --role FRONTEND, DATABASE, MG01\n"
            "  - Exclusion with --role (quote or escape ! in bash):\n"
            "      --role '!DEFAULT'\n"
            "      --role \\!DEFAULT\n"
            "  - Exclusion with not: syntax: --role not:DEFAULT\n"
            "  - Recommended shell-safe exclusion: --exclude-role DEFAULT"
        ),
    )
    ap.add_argument(
        "--exclude-role",
        nargs="+",
        help=(
            "Shell-safe exclusion filter for role labels.\n"
            "  - Single exclusion: --exclude-role DEFAULT\n"
            "  - Multiple exclusions: --exclude-role FRONTEND,DATABASE\n"
            "    Also accepted with spaces: --exclude-role FRONTEND, DATABASE\n"
            "Equivalent to --role '!DEFAULT', without bash history-expansion issues."
        ),
    )
    for k in ["app", "env", "loc"]:
        ap.add_argument(f"--{k}")
    ap.add_argument("--OS", dest="OS")
    ap.add_argument("--days", type=int, required=True, help="Number of days")
    ap.add_argument(
        "--mail-to",
        default="",
        help="Send a notification email to the provided address(es) (comma/semicolon-separated).",
    )
    ap.add_argument("--one-interface-match", action="store_true", default=True, help=argparse.SUPPRESS)
    ap.add_argument("--dev-flow-stub", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--dev-flow-stub-out", help=argparse.SUPPRESS)
    ap.add_argument("--dev-flow-stub-in", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-label", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-labelgroup", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-iplists", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-wkld-m", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-processes-bulk", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-processes", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-wkld-u", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-ruleset", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-rules-enabled", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-dupecheck", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-dupecheck-filtered", help=argparse.SUPPRESS)
    ap.add_argument("--dev-stub-dupecheck-final", help=argparse.SUPPRESS)
    ap.add_argument("--debug-echo-dupecheck", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--debug-no-scope-filter", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--network-zone", help="IPList including subnet(s) of the zone")
    ap.add_argument("--CreateRules", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--RecertifyRules", action="store_true", help=argparse.SUPPRESS)

    ap.add_argument("--FlowRuleReview", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--CreateNew", action="store_true", default=True, help=argparse.SUPPRESS)

    ap.add_argument("--strategy-egress-bubble", choices=["blacklist", "whitelist", "none"], default=None, help=argparse.SUPPRESS)
    ap.add_argument("--strategy-ingress-bubble", choices=["whitelist", "blacklist", "none"], default=None, help=argparse.SUPPRESS)

    strategy_intra_group = ap.add_argument_group("**Intra-app strategy options:**")
    strategy_intra_group.add_argument(
        "--strategy-intra-app",
        choices=["allow", "finegrained", "blacklist"],
        default=None,
        help="Strategy for intra-app traffic (generates sheet \"Proposed rules\" via propose_rule_for_scope)",
    )
    strategy_intra_group.add_argument(
        "--ports-to-blacklist-intra-app",
        default="",
        help="carto.conf port-list names for intra-app blacklist strategy (e.g. PORTS_TO_CONTROL,PORTS_ADMIN,PORTS_TO_ERADICATE)",
    )
    strategy_egress_group = ap.add_argument_group("**Egress strategy options:**")
    strategy_egress_group.add_argument(
        "--strategy-egress",
        choices=["allow", "finegrained", "blacklist"],
        default=None,
        help="Strategy for outbound traffic (generates Proposed rules rows).",
    )
    strategy_egress_group.add_argument(
        "--ports-to-blacklist-egress",
        default="",
        help="carto.conf port-list names for egress blacklist strategy",
    )
    strategy_ingress_group = ap.add_argument_group("**Ingress strategy options:**")
    strategy_ingress_group.add_argument(
        "--strategy-ingress",
        choices=["allow", "finegrained", "blacklist"],
        default=None,
        help="Strategy for inbound traffic (generates Proposed rules rows).",
    )
    strategy_ingress_group.add_argument(
        "--ports-to-blacklist-ingress",
        default="",
        help="carto.conf port-list names for ingress blacklist strategy",
    )
    ap.add_argument(
        "--excel-stream-update",
        action="store_true",
        help="Low-mem Excel mode: use this mode in case of memory issues due to huge amount of flows",
    )
    ap.add_argument("--add-elected-iplist-column", action="store_true", default=True, help=argparse.SUPPRESS)
    ap.add_argument("--enable-umgd-app-label-rules-sheet", action="store_true", default=True, help=argparse.SUPPRESS)

    ap.add_argument(
        "--enable-to-investigate",
        action="store_true",
        default=True,
        help=argparse.SUPPRESS,
    )
    ap.add_argument(
        "--skip-pce-fqdn-enrichment",
        action="store_true",
        default=False,
        help="Disable PCE FQDN enrichment for the 'To investigate' sheet.",
    )
    ap.add_argument("--dns-timeout", type=float, default=5.0, help=argparse.SUPPRESS)

    # Legacy (backward compatible): when only one direction is in blacklist mode, this applies to that direction.
    ap.add_argument("--ports-to-blacklist", default="", help=argparse.SUPPRESS)
    # Optional: replace some peer app labels (configured in carto.conf AVOID_LABEL_PAIRS)
    # by best-matching IPLISTS (using conf priority) when building Proposed rules.
    # Non-regression: disabled by default unless this flag is set.
    ap.add_argument("--enable-avoid-label-pairs", action="store_true", default=True, help=argparse.SUPPRESS)
    # Default behavior: enabled in propose_rule_for_scope. Use this flag to disable.
    ap.add_argument("--no-mark-potential-core-service", dest="mark_potential_core_service", action="store_false", default=True,
                    help=argparse.SUPPRESS)

    ap.add_argument("--strategy-intra-bubble", choices=["restrict-by-role", "open", "none"], default=None, help=argparse.SUPPRESS)

    # NEW v2.2: Flow→Rule Hits integration flags
    ap.add_argument("--FlowRuleHits", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--frh-filter-direction", choices=["ingress", "egress"], default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-filter-proto", choices=["TCP", "UDP", "ICMP"], default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-filter-port", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-ruleset-name-contains", type=str, default="", help=argparse.SUPPRESS)
    ap.add_argument("--frh-exclude-all-workloads-rules", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--frh-prefer-raw", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--frh-limit-flows", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-debug", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--frh-debug-matches-only", action="store_true", help=argparse.SUPPRESS)
    ap.add_argument("--frh-debug-max-rows", type=int, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-debug-sample-rate", type=float, default=None, help=argparse.SUPPRESS)
    ap.add_argument("--frh-log-level", type=str, default="INFO", help=argparse.SUPPRESS)

    args = ap.parse_args()
    exec_start_dt = datetime.now()
    mail_recipients = parse_recipients(args.mail_to)
    MAIL_CONTEXT.update(
        {
            "enabled": bool(mail_recipients),
            "recipients": mail_recipients,
            "status": "FAIL",
            "app": getattr(args, "app", "") or "",
            "env": getattr(args, "env", "") or "",
        }
    )

    # Guard: mutually exclusive high-level branches
    if args.FlowRuleReview and args.CreateNew:
        raise SystemExit("Choose only one: --FlowRuleReview or --CreateNew")


    conf = load_conf("carto.conf")
    MAIL_CONTEXT["conf"] = conf
    exe = conf.get("EXECUTABLE", "").strip(); cfg = conf.get("EXECUTABLE_CONFIG_FILE", "").strip()
    root = Path(conf.get("EXPORT_ROOT", "./RUNS")); date_fmt = conf.get("DATE_FMT", "%Y%m%d-%H%M%S")
    if not exe or not os.path.isfile(exe):
        raise SystemExit(f"EXECUTABLE not found: {exe}")
    if not os.access(exe, os.X_OK):
        raise SystemExit(f"EXECUTABLE not executable: {exe}")
    if not cfg or not os.path.isfile(cfg):
        raise SystemExit(f"EXECUTABLE_CONFIG_FILE not found: {cfg}")

    start, end = compute_window_days(args.days)
    print(f"Window from --days={args.days}: start={start} end={end}")

    base_ts = now_stamp(date_fmt)
    app = sanitize_token(getattr(args, "app", "") or "")
    envl = sanitize_token(getattr(args, "env", "") or "")
    role = build_role_filter_arg(getattr(args, "role", ""), getattr(args, "exclude_role", ""))
    role_for_path = sanitize_token(role)
    label_suffix = "-".join([x for x in (app, envl, role_for_path) if x])
    run_ts = f"{base_ts}" if not label_suffix else f"{base_ts}_{label_suffix}"
    run_dir = root/run_ts
    raw = run_dir/"raw"; der = run_dir/"derived"; xls = run_dir/"excel"
    raw.mkdir(parents=True, exist_ok=True); der.mkdir(parents=True, exist_ok=True); xls.mkdir(parents=True, exist_ok=True)

    log = run_dir/'log'
    log.mkdir(parents=True, exist_ok=True)
    log_file = log / f"orchestrator_{run_ts}.log"
    stdout_log_file = log / f"orchestrator_{run_ts}.stdout.log"
    MAIL_CONTEXT["log_path"] = str(stdout_log_file)
    if MAIL_CONTEXT.get("enabled"):
        try:
            MAIL_CONTEXT["tee_state"] = start_stdout_tee(stdout_log_file)
        except Exception as exc:
            MAIL_CONTEXT["enabled"] = False
            print(f"[WARN] Email log capture disabled: {exc}")
    atexit.register(finalize_email_notification)
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
    fh = logging.FileHandler(str(log_file), encoding='utf-8'); fh.setFormatter(fmt)
    ch = logging.StreamHandler(); ch.setFormatter(fmt)
    for h in list(logger.handlers): logger.removeHandler(h)
    logger.addHandler(fh); logger.addHandler(ch)
    logger.info(f"Run started: ts={run_ts} scope app={app} env={envl} role={role} days={args.days}")

    env: Dict[str, str] = dict(os.environ)
    env.update(load_env_file(Path(".env")))
    env.update({
        "EXECUTABLE": exe,
        "CFG": cfg,
        "PCE_NAME": conf.get("PCE_NAME", ""),
        "BASE_SLEEP": conf.get("RETRY_BASE_SLEEP", "3"),
        "BACKOFF": conf.get("RETRY_BACKOFF_FACTOR", "2"),
        "MAX_SLEEP": conf.get("RETRY_MAX_SLEEP", "60"),
        "JITTER": conf.get("RETRY_JITTER_PCT", "20"),
        "TIMEOUT_SEC": conf.get("TIMEOUT_SEC", "2700"),
        "MAX_ATTEMPTS": conf.get("MAX_ATTEMPTS", "5"),
    })

    bin_dir = Path(__file__).parent/"bin"

    def dev_copy(src: Optional[str], dst: Path, label: str) -> bool:
        if not src:
            return False
        try:
            ensure_dir(dst)
            import shutil
            shutil.copy2(src, dst)
            print(f"DEV-STUB: copy {label} '{src}' -> '{dst}'")
            return True
        except Exception as e:
            print(f"ERROR: DEV-STUB copy failed for {label}: {e}")
            try:
                logger.warning(f"DEV-STUB copy failed for {label}: {e}")
            except Exception:
                pass
            return False

    # label-export
    if args.dev_stub_label:
        dev_copy(args.dev_stub_label, raw/"export_label.csv", "export_label.csv"); DUR["label-export"] = DUR.get("label-export", 0.0)
    else:
        ok = run_step("label-export", ["bash", str(bin_dir/"workloader_label.sh"), str(raw/"export_label.csv")], env, Path("."))
        if not ok:
            print("WARN: label-export failed; continuing if file exists")

    labels = read_labels_csv(raw/"export_label.csv")
    filters = {
        k: normalize_arg_value(getattr(args, k))
        for k in ["app", "env", "loc", "role", "OS"]
        if normalize_arg_value(getattr(args, k, ""))
    }
    if role:
        filters["role"] = role
    hrefs_labels = find_label_hrefs(labels, filters)
    include_file = der/"include_labels_semicolon.csv"; write_list_semicolon(include_file, hrefs_labels)

    # flows out
    fout = raw/f"flows_out_{start}_{end}.csv"
    if args.dev_flow_stub and args.dev_flow_stub_out:
        dev_copy(args.dev_flow_stub_out, fout, "flows_out"); DUR["traffic-out"] = DUR.get("traffic-out", 0.0)
    else:
        run_step("traffic-out", ["bash", str(bin_dir/"workloader_traffic_out.sh"), str(include_file), start, end, str(fout)], env, Path("."))

    # labelgroup-export
    if args.dev_stub_labelgroup:
        dev_copy(args.dev_stub_labelgroup, raw/"export_labelgroup.csv", "export_labelgroup.csv"); DUR["labelgroup-export"] = DUR.get("labelgroup-export", 0.0)
    else:
        run_step("labelgroup-export", ["bash", str(bin_dir/"workloader_labelgroup.sh"), str(raw/"export_labelgroup.csv")], env, Path("."))

    # iplists export
    ipl = raw/"export_iplists.csv"
    if args.dev_stub_iplists:
        dev_copy(args.dev_stub_iplists, ipl, "export_iplists.csv"); DUR["ipl-export"] = DUR.get("ipl-export", 0.0)
    else:
        run_step("ipl-export", ["bash", str(bin_dir/"workloader_ipl_export.sh"), str(ipl)], env, Path("."))

    # flows in
    fin = raw/f"flows_in_{start}_{end}.csv"
    def build_excl_files(conf_map: Dict[str, str], out_dir: Path) -> Tuple[Optional[Path], Optional[Path]]:
        lbls = (conf_map.get("HREF_LABEL_QUALYS", "") or "").strip()
        ipls = (conf_map.get("HREF_IPLISTS_QUALYS", "") or "").strip()
        p_lbl = None; p_ipl = None
        if lbls:
            vals = [x.strip() for x in lbls.split(";") if x.strip()]
            if vals:
                p_lbl = out_dir / "exclude_flow_in_qualys.labels.csv"
                write_list_newline(p_lbl, list(dict.fromkeys(vals)))
        if ipls:
            vals = [x.strip() for x in ipls.split(";") if x.strip()]
            if vals:
                p_ipl = out_dir / "exclude_flow_in_qualys.iplists.csv"
                write_list_newline(p_ipl, list(dict.fromkeys(vals)))
        return p_lbl, p_ipl

    excl_lbl, excl_ipl = build_excl_files(conf, der)
    if args.dev_flow_stub and args.dev_flow_stub_in:
        dev_copy(args.dev_flow_stub_in, fin, "flows_in"); DUR["traffic-in"] = DUR.get("traffic-in", 0.0)
    else:
        run_step("traffic-in",
                 ["bash", str(bin_dir/"workloader_traffic_in.sh"),
                  str(include_file), start, end, str(fin), str(excl_lbl or ""), str(excl_ipl or "")],
                 env, Path("."))

    # wkld.m export
    wkld_m_path = raw/"export_wkld.m.csv"
    if args.dev_stub_wkld_m:
        dev_copy(args.dev_stub_wkld_m, wkld_m_path, "export_wkld.m.csv"); DUR["wkld-export-m"] = DUR.get("wkld-export-m", 0.0)
    else:
        run_step("wkld-export-m", ["bash", str(bin_dir/"workloader_wkld_m.sh"), str(wkld_m_path)], env, Path("."))

    # processes export
    final_proc = raw/"export_processes.csv"
    if args.dev_stub_processes:
        dev_copy(args.dev_stub_processes, final_proc, "export_processes.csv"); DUR["process-export-aggregate"] = DUR.get("process-export-aggregate", 0.0)
    elif args.dev_stub_processes_bulk:
        bulk_out = raw/"export_processes.bulk.csv"
        dev_copy(args.dev_stub_processes_bulk, bulk_out, "export_processes.bulk.csv")
        ensure_dir(final_proc)
        if bulk_out.exists() and bulk_out.stat().st_size > 0:
            with bulk_out.open("r", encoding="utf-8") as fi, final_proc.open("w", encoding="utf-8", newline="") as fo:
                fo.write(fi.read())
        else:
            with final_proc.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=['hostname','process_path','service_name','port','proto','role','app','env','loc','OS','href'])
                w.writeheader()
        DUR["process-export-aggregate"] = DUR.get("process-export-aggregate", 0.0)
    else:
        try:
            hrefs_scope = find_managed_wkld_hrefs_for_filters(wkld_m_path, filters)
        except Exception as e:
            print(f"WARN: cannot compute managed hrefs for scope: {e}")
            hrefs_scope = []
        list_href_file = der/"list_href.managed.csv"
        write_list_newline(list_href_file, hrefs_scope)
        bulk_out = raw/"export_processes.bulk.csv"
        ok_bulk = run_step("process-export-bulk", ["bash", str(bin_dir/"workloader_process_bulk.sh"), str(list_href_file), str(bulk_out)], env, Path("."))
        if not ok_bulk:
            print("WARN: process-export-bulk failed; attempting single; continuing if file exists")
            try:
                logger.warning("process-export-bulk failed; attempting single; continuing if file exists")
            except Exception:
                pass
        ensure_dir(final_proc)
        if bulk_out.exists() and bulk_out.stat().st_size > 0:
            with bulk_out.open("r", encoding="utf-8") as fi, final_proc.open("w", encoding="utf-8", newline="") as fo:
                fo.write(fi.read())
        else:
            ok_single = run_step("process-export-single", ["bash", str(bin_dir/"workloader_process.sh"), str(list_href_file), str(final_proc)], env, Path("."))
            if not ok_single or (not final_proc.exists()) or final_proc.stat().st_size == 0:
                with final_proc.open("w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=['hostname','process_path','service_name','port','proto','role','app','env','loc','OS','href'])
                    w.writeheader()
        DUR["process-export-aggregate"] = DUR.get("process-export-aggregate", 0.0)

    # wkld.u export
    if args.dev_stub_wkld_u:
        dev_copy(args.dev_stub_wkld_u, raw/"export_wkld.u.csv", "export_wkld.u.csv"); DUR["wkld-export-u"] = DUR.get("wkld-export-u", 0.0)
    else:
        run_step("wkld-export-u", ["bash", str(bin_dir/"workloader_wkld_u.sh"), str(raw/"export_wkld.u.csv")], env, Path("."))

    # ruleset export
    if args.dev_stub_ruleset:
        dev_copy(args.dev_stub_ruleset, raw/"export_ruleset.csv", "export_ruleset.csv"); DUR["ruleset-export"] = DUR.get("ruleset-export", 0.0)
    else:
        run_step("ruleset-export", ["bash", str(bin_dir/"workloader_ruleset.sh"), str(raw/"export_ruleset.csv")], env, Path("."))

    # rules.enabled export + list href enabled
    rs_path = raw/"export_ruleset.csv"; rows_rs, fn_rs = load_csv(rs_path); hrefs_enabled: List[str] = []
    if rows_rs:
        c_href = pick(fn_rs, "href"); c_en = pick(fn_rs, "enabled")
        for r in rows_rs:
            if is_truthy(r.get(c_en, '')):
                h = norm_href(r.get(c_href) or '')
                if h:
                    hrefs_enabled.append(h)
    list_enabled = der/"list_href.ruleset.enabled.csv"; write_list_newline(list_enabled, sorted(set(hrefs_enabled)))
    if args.dev_stub_rules_enabled:
        dev_copy(args.dev_stub_rules_enabled, raw/"export_rules.enabled.csv", "export_rules.enabled.csv"); DUR["rule-export-enabled-rs"] = DUR.get("rule-export-enabled-rs", 0.0)
    else:
        run_step("rule-export-enabled-rs", ["bash", str(bin_dir/"workloader_rule_enabled.sh"), str(list_enabled), str(raw/"export_rules.enabled.csv")], env, Path("."))

    # zone filter (optional)
    if getattr(args, "network_zone", None):
        run_step("zone-filter", [
            "python3", str(Path(__file__).parent/"modules"/"zone_filter.py"),
            "--input-raw", str(raw),
            "--derived-dir", str(der),
            "--zone-iplist", str(args.network_zone),
            "--start", start, "--end", end
        ], env, Path("."))

    # dupecheck
    dupe_raw = raw/"export_dupecheck.csv"
    if args.dev_stub_dupecheck:
        dev_copy(args.dev_stub_dupecheck, dupe_raw, "export_dupecheck.csv"); DUR["dupecheck"] = DUR.get("dupecheck", 0.0)
    else:
        run_step("dupecheck", ["bash", str(bin_dir/"workloader_dupecheck.sh"), str(dupe_raw), "true" if args.one_interface_match else ""], env, Path("."))

    dup_rows, dup_fn = load_csv(dupe_raw)
    # optional filtered dupecheck stub copy
    if args.dev_stub_dupecheck_filtered:
        stub_f = Path(args.dev_stub_dupecheck_filtered)
        dst_f = der / "export_dupecheck.filtered.csv"
        ensure_dir(dst_f)
        copied = False
        try:
            import shutil
            shutil.copy2(stub_f, dst_f)
            print(f"DEV-STUB: copy export_dupecheck.filtered.csv '{stub_f}' -> '{dst_f}'")
            copied = True
        except Exception as e:
            print(f"[WARN] DEV-STUB copy export_dupecheck.filtered.csv failed: {e}; continue with raw/exploded path")
            try:
                logger.warning(f"DEV-STUB copy export_dupecheck.filtered.csv failed: {e}; continue with raw/exploded path")
            except Exception:
                pass
        if copied:
            dup_rows, dup_fn = load_csv(dst_f)

    if args.debug_echo_dupecheck:
        print(f"[DEBUG] dupe rows: {len(dup_rows)} cols: {dup_fn}")
    fn_low = [f.lower() for f in (dup_fn or [])]
    has_reason = 'reason' in fn_low
    has_href_managed = 'href_managed' in fn_low
    has_final_cols = has_href_managed and (('hostname_managed' in fn_low) or ('role_managed' in fn_low) or ('app_managed' in fn_low))

    if has_final_cols:
        base_rows = dup_rows
        print('INFO: Using filtered/final dupecheck rows directly (final-like schema).')
        for r in base_rows:
            r['href_managed'] = norm_href(r.get('href_managed',''))
    elif has_href_managed:
        print('INFO: Using filtered dupecheck rows (exploded schema), enriching with managed attributes...')
        base_rows = join_with_managed(dup_rows, *load_csv(raw/"export_wkld.m.csv"))
        for r in base_rows:
            r['href_managed'] = norm_href(r.get('href_managed',''))
    elif has_reason:
        print('INFO: Using raw dupecheck rows (reason-based), exploding then enriching...')
        exp_rows = explode_reason_to_matches(dup_rows, dup_fn)
        base_rows = join_with_managed(exp_rows, *load_csv(raw/"export_wkld.m.csv"))
        for r in base_rows:
            r['href_managed'] = norm_href(r.get('href_managed',''))
    else:
        print('WARN: dupecheck rows have unexpected schema (no reason / no href_managed); Workloads may be empty.')
        base_rows = []

    hrefs_scope_set = set(norm_href(h) for h in find_managed_wkld_hrefs_for_filters(raw/"export_wkld.m.csv", filters))
    if has_final_cols and not hrefs_scope_set:
        hrefs_scope_set = {norm_href(r.get('href_managed','')) for r in base_rows if r.get('href_managed')}
    if not args.debug_no_scope_filter and hrefs_scope_set:
        base_rows = [r for r in base_rows if norm_href(r.get('href_managed','')) in hrefs_scope_set]

    all_rows, counts = enrich_workloads(base_rows, filters, raw/"export_wkld.m.csv", raw/"export_wkld.u.csv", debug=args.debug_echo_dupecheck)
    lg_rows = read_labelgroups_csv(raw/"export_labelgroup.csv")
    final_csv = build_final_csv(der, all_rows)
    eff_scope = set(r.get('href_managed','') for r in all_rows)
    scope_values = gather_scope_label_values(raw/"export_wkld.m.csv", eff_scope if eff_scope else hrefs_scope_set)
    lbl_groups_incl_scope = build_labelgroup_inclusions_for_scope(lg_rows, scope_values)
    prows, pfn = load_csv(raw/"export_processes.csv")
    processes_rows: List[Dict[str, str]] = []
    if prows:
        def pickp(cols: List[str], *c: str) -> str:
            low = {x.lower(): x for x in cols}
            for k in c:
                if k.lower() in low:
                    return low[k.lower()]
            return ''
        c_host = pickp(pfn, 'hostname','host','name','host_name')
        c_path = pickp(pfn, 'process_path','path','exe','binary','command')
        c_svc = pickp(pfn, 'service_name','service','svc_name')
        c_port = pickp(pfn, 'port','listen_port')
        c_proto= pickp(pfn, 'proto','protocol')
        c_role = pickp(pfn, 'role')
        c_app = pickp(pfn, 'app','application')
        c_env = pickp(pfn, 'env','environment')
        c_loc = pickp(pfn, 'loc','location')
        c_os = pickp(pfn, 'OS','os')
        c_href = pickp(pfn, 'href')
        for r in prows:
            processes_rows.append({
                'hostname': r.get(c_host, ''),
                'process_path': r.get(c_path, ''),
                'service_name': r.get(c_svc, ''),
                'port': r.get(c_port, ''),
                'proto': r.get(c_proto, ''),
                'role': r.get(c_role, ''),
                'app': r.get(c_app, ''),
                'env': r.get(c_env, ''),
                'loc': r.get(c_loc, ''),
                'OS': r.get(c_os, ''),
                'href': r.get(c_href, '')
            })

    final_xlsx = build_final_excel(
        xls,
        raw,
        all_rows,
        processes_rows,
        start,
        end,
        app,
        envl,
        role,
        lbl_groups_incl_scope,
        counts,
        exec_start_dt,
        include_flow_sheets=not args.excel_stream_update,
        add_elected_iplist_column=args.add_elected_iplist_column,
        enable_umgd_app_label_rules_sheet=args.enable_umgd_app_label_rules_sheet,
        network_zone_name=args.network_zone,
    )
    print(f"POST-PROCESS: final CSV -> {final_csv}")
    print(f"POST-PROCESS: final XLSX -> {final_xlsx}")
    try:
        logger.info(f"POST-PROCESS: final XLSX -> {final_xlsx}")
    except Exception:
        pass

    # NEW HOOK: Services Catalogue (export_services.py)
    try:
        enable_services = (conf.get("ENABLE_SERVICES_CATALOG", "") or "").strip()
    except Exception:
        enable_services = ""
    if is_truthy(enable_services) or args.FlowRuleReview or args.CreateNew:
        run_step(
            "export-services",
            [
                "python3", str(Path(__file__).parent/"modules"/"export_services.py"),
                "--input-raw", str(raw),
                "--derived-dir", str(der),
                "--conf", "carto.conf",
                "--start", start, "--end", end,
                "--excel", str(final_xlsx),
            ],
            env, Path(".")
        )

    # Optional modules after Excel is produced — they append sheets

    # ------------------------------ Branches: FlowRuleReview / CreateNew ------------------------------
    # IMPORTANT: we do NOT change the scope applicability logic (scope_rules_applicability.py). We only
    # run downstream analysis that matches flows against the already computed 'Scope Applicable Rules'.
    if args.FlowRuleReview or args.CreateNew:
        # 1) Compute Scope Applicable Rules + Ruleset Effectiveness (append sheets into Excel)
        ok = run_step(
            "scope-applicable-rules",
            [
                "python3", str(Path(__file__).parent/"modules"/"flows_to_rules.py"),
                "--input-raw", str(raw),
                "--derived-dir", str(der),
                "--conf", "carto.conf",
                "--start", start, "--end", end,
                "--excel", str(final_xlsx),
                "--scope-app", str(app or ""),
                "--scope-env", str(envl or ""),
                "--scope-role", str(role or ""),
                "--log-level", "INFO",
            ],
            env, Path(".")
        )
        if not ok:
            print("[ERROR] scope-applicable-rules failed; aborting FlowRuleReview/CreateNew.")
            return 2

        # 2) Flow↔Rule review (writes Flow-Rule Match + Proposed Rules + Action Plan)
        pr_cmd = [
            "python3", str(Path(__file__).parent/"modules"/"propose_rule_for_scope.py"),
            "--input-raw", str(raw),
            "--derived-dir", str(der),
            "--conf", "carto.conf",
            "--start", start, "--end", end,
            "--excel", str(final_xlsx),
            "--log-level", "INFO",
        ]
        # Reuse existing debug toggles to enable verbose mode (keeps CLI stable)
        if args.dev_flow_stub or args.debug_echo_dupecheck or args.debug_no_scope_filter:
            pr_cmd += ["--debug"]

        # Proposed rules strategies
        if getattr(args, "strategy_intra_app", None):
            pr_cmd += ["--strategy-intra-app", args.strategy_intra_app]
        if getattr(args, "strategy_ingress", None):
            pr_cmd += ["--strategy-ingress", args.strategy_ingress]
        if getattr(args, "strategy_egress", None):
            pr_cmd += ["--strategy-egress", args.strategy_egress]
        if getattr(args, "network_zone", None):
            pr_cmd += ["--network-zone", args.network_zone]

        # Ports-to-blacklist (legacy + per-direction). Per-direction flags take precedence in the module.
        for flag, val in [
            ("--ports-to-blacklist-intra-app", getattr(args, "ports_to_blacklist_intra_app", "")),
            ("--ports-to-blacklist-egress", getattr(args, "ports_to_blacklist_egress", "")),
            ("--ports-to-blacklist-ingress", getattr(args, "ports_to_blacklist_ingress", "")),
            ("--ports-to-blacklist", getattr(args, "ports_to_blacklist", "")),  # legacy
        ]:
            if str(val).strip():
                pr_cmd += [flag, str(val).strip()]
        if getattr(args, "enable_avoid_label_pairs", False):
            pr_cmd.append("--enable-avoid-label-pairs")
        if not getattr(args, "mark_potential_core_service", True):
            pr_cmd.append("--no-mark-potential-core-service")

        ok = run_step("flow-rule-review", pr_cmd, env, Path("."))
        if not ok:
            print("[ERROR] flow-rule-review failed; aborting FlowRuleReview/CreateNew.")
            return 2

        # Branch 2: CreateNew stops at Flow-Rule Match — remove proposal/action sheets
        if args.CreateNew:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(final_xlsx))
                removed = False
                for sname in ["Proposed Rules", "Action Plan"]:
                    if sname in wb.sheetnames:
                        del wb[sname]
                        removed = True
                if removed:
                    wb.save(str(final_xlsx))
                    print("[INFO] CreateNew: removed sheets Proposed Rules / Action Plan (stopping at Flow-Rule Match).")
            except Exception as e:
                print(f"[WARN] CreateNew: could not remove Proposed Rules / Action Plan: {e}")

    if args.RecertifyRules:
        subprocess.run([
            "python3", str(Path(__file__).parent/"modules"/"recertify_rules.py"),
            "--input-raw", str(raw),
            "--derived-dir", str(der),
            "--excel", str(final_xlsx),
            "--conf", "carto.conf",
            "--start", start, "--end", end
        ])

    if args.CreateRules:
        call = [
            "python3", str(Path(__file__).parent/"modules"/"propose_rules_for_app.py"),
            "--input-raw", str(raw),
            "--derived-dir", str(der),
            "--excel", str(final_xlsx),
            "--conf", "carto.conf",
            "--start", start, "--end", end
        ]
        if args.strategy_egress_bubble: call += ["--strategy-egress-bubble", args.strategy_egress_bubble]
        if args.strategy_ingress_bubble: call += ["--strategy-ingress-bubble", args.strategy_ingress_bubble]
        if args.strategy_intra_bubble: call += ["--strategy-intra-bubble", args.strategy_intra_bubble]
        if args.network_zone: call += ["--network-zone", args.network_zone]
        subprocess.run(call)

    # NEW v2.2: Flow→Rule Hits integration (optional)
    if args.FlowRuleHits:
        frh = [
            "python3", str(Path(__file__).parent/"modules"/"flows_to_rules.py"),
            "--input-raw", str(raw),
            "--derived-dir", str(der),
            "--conf", "carto.conf",
            "--start", start, "--end", end,
            "--excel", str(final_xlsx),
            "--scope-app", str(app or ""),
            "--scope-env", str(envl or ""),
            "--scope-role", str(role or ""),
            "--log-level", args.frh_log_level
        ]
        # default filters if explicitly provided
        if args.frh_filter_direction:
            frh += ["--filter-direction", args.frh_filter_direction]
        if args.frh_filter_proto:
            frh += ["--filter-proto", args.frh_filter_proto]
        if args.frh_filter_port is not None:
            frh += ["--filter-port", str(args.frh_filter_port)]
        # ruleset filtering
        if args.frh_ruleset_name_contains:
            frh += ["--ruleset-name-contains", args.frh_ruleset_name_contains]
        else:
            # Default to scope tokens (app;env;role) if provided
            default_tokens = ";".join([x for x in [app, envl, role] if x])
            if default_tokens:
                frh += ["--ruleset-name-contains", default_tokens]
        if args.frh_exclude_all_workloads_rules:
            frh += ["--exclude-all-workloads-rules"]
        if args.frh_prefer_raw:
            frh += ["--prefer-raw"]
        if args.frh_limit_flows is not None:
            frh += ["--limit-flows", str(args.frh_limit_flows)]
        # debug flags
        if args.frh_debug:
            frh += ["--debug"]
            if args.frh_debug_matches_only:
                frh += ["--debug-matches-only"]
            if args.frh_debug_max_rows is not None:
                frh += ["--debug-max-rows", str(args.frh_debug_max_rows)]
            if args.frh_debug_sample_rate is not None:
                frh += ["--debug-sample-rate", str(args.frh_debug_sample_rate)]

        # Run and record duration in summary
        run_step("flow-rule-hits", frh, env, Path("."))



    # Optional: Build 'To investigate' sheet (trivial branch only)
    if args.enable_to_investigate and not args.excel_stream_update:
        try:
            n_rows = build_to_investigate_sheet(final_xlsx, dns_timeout=float(args.dns_timeout))
            print(f"[INFO] [To investigate] rows: {n_rows}")
        except Exception as e:
            print(f"[WARN] [To investigate] skipped due to: {e}")

    # FINAL STEP (optional): inject Flow-out/Flow-in sheets using streaming to avoid OOM.
    # This is only needed when Flow sheets were skipped during initial Excel build to keep openpyxl updates low-memory.
    if args.excel_stream_update:
        flow_out_path = raw / f"flows_out_{start}_{end}.csv"
        flow_in_path = raw / f"flows_in_{start}_{end}.csv"
        # Resolve excel_stream_update script path (prefer modules/, fallback to repo root)
        excel_stream_script = Path(__file__).parent / 'modules' / 'excel_stream_update.py'
        if not excel_stream_script.exists():
            alt = Path(__file__).parent / 'excel_stream_update.py'
            if alt.exists():
                excel_stream_script = alt
            else:
                excel_stream_script = Path('excel_stream_update.py')

        cmd_excel_stream = [
            'python3',
            str(excel_stream_script),
            '--input-xlsx',
            str(final_xlsx),
            '--flows-out',
            str(flow_out_path),
            '--flows-in',
            str(flow_in_path),
            '--conf',
            'carto.conf',
        ]
        if args.add_elected_iplist_column:
            cmd_excel_stream.append('--add-elected-iplist-column')

        if args.enable_to_investigate:
            cmd_excel_stream.append('--enable-to-investigate')
            cmd_excel_stream.extend(['--dns-timeout', str(args.dns_timeout)])
        cmd_excel_stream += [
            '--log-level',
            str(getattr(args, 'log_level', 'INFO')),
        ]
        ok = run_step(
            'excel-stream-update',
            cmd_excel_stream,
            env,
            Path('.'),
        )
        if not ok:
            print("[ERROR] excel-stream-update failed; final Excel may be missing Flow-out/Flow-in sheets.")
            return 2

    # Final step (optional): enrich unknown Flow-out IPs with FQDN from PCE flows.
    if args.skip_pce_fqdn_enrichment:
        print("[INFO] [To investigate] PCE FQDN enrichment disabled (--skip-pce-fqdn-enrichment).")
    else:
        try:
            enrich_unknown_ips_with_pce_fqdn(
                xlsx_path=final_xlsx,
                bin_dir=bin_dir,
                env=env,
                derived_dir=der,
                start=start,
                end=end,
                tmp_stamp=base_ts,
            )
        except Exception as e:
            print(f"[WARN] [To investigate] PCE FQDN enrichment skipped due to: {e}")

    try:
        n_egress, n_ingress = build_to_investigate_ip_sheets(final_xlsx)
        if n_egress or n_ingress:
            print(f"[INFO] [To investigate] Egress/IP sheets rows: egress={n_egress} ingress={n_ingress}")
    except Exception as e:
        print(f"[WARN] [To investigate] IP sheets skipped due to: {e}")

    print("==== EXECUTION SUMMARY (durations) ====")
    for k in sorted(DUR.keys()):
        print(f" - {k:28s}: {DUR[k]:6.1f}s")

    try:
        final_xlsx = rename_final_excel(final_xlsx, app, envl)
        rel_final_xlsx = os.path.relpath(final_xlsx, Path.cwd())
        print(f"[SUCCESS] Execution completed. Output Excel: {rel_final_xlsx}")
        MAIL_CONTEXT["excel_path"] = rel_final_xlsx
    except Exception as e:
        print(f"[WARN] Unable to rename the final Excel file: {e}")
        MAIL_CONTEXT["excel_path"] = str(final_xlsx)
    MAIL_CONTEXT["status"] = "SUCCESS"
    return 0

if __name__ == "__main__":
    sys.exit(main())
